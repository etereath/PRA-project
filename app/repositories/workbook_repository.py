from __future__ import annotations

import ast
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.enums import (
    ListingStrategy,
    PricingMethod,
    PricingSource,
    RoundingRule,
    TaskActionType,
    TaskOriginType,
    TaskStatus,
)
from app.exceptions import TableValidationError, TableValidationIssue, ValidationError
from app.models import (
    ColdStorageStatus,
    ExecutionLog,
    HarvestForecast,
    ListingRule,
    PackingCapacityPlan,
    PriceForecast,
    PriceRule,
    Product,
    Task,
)
from app.services.harvest_forecast import forecast_group_key
from app.utils import parse_bool, parse_date, parse_datetime, parse_decimal, parse_int, serialize_decimal


PRODUCT_HEADERS = [
    "internal_sku",
    "product_name",
    "grade",
    "stem_length",
    "unit",
    "base_cost",
    "current_stock",
    "sale_enabled",
    "last_price",
    "recommended_price",
    "remark",
    "feature_season",
    "feature_color",
]

PRICE_RULE_HEADERS = [
    "rule_id",
    "rule_name",
    "variety_filter",
    "grade_filter",
    "platform_filter",
    "pricing_method",
    "markup_value",
    "min_price",
    "rounding_rule",
    "rounding_step",
    "active",
    "priority",
    "remark",
]

LISTING_RULE_HEADERS = [
    "rule_id",
    "rule_name",
    "variety_filter",
    "grade_filter",
    "platform_filter",
    "stock_threshold",
    "listing_strategy",
    "active",
    "priority",
    "remark",
]

HARVEST_FORECAST_HEADERS = [
    "forecast_id",
    "forecast_date",
    "target_trade_date",
    "variety",
    "grade",
    "predicted_harvest_qty",
    "lower_bound_qty",
    "upper_bound_qty",
    "confidence",
    "source",
    "generated_at",
    "note",
]

PRICE_FORECAST_HEADERS = [
    "forecast_id",
    "forecast_date",
    "target_trade_date",
    "variety",
    "grade",
    "recommended_price",
    "lower_bound_price",
    "upper_bound_price",
    "confidence",
    "source",
    "generated_at",
    "note",
]

CAPACITY_PLAN_HEADERS = [
    "trade_date",
    "normal_packing_capacity_qty",
    "temp_worker_capacity_qty",
    "confirmed_temp_worker_count",
    "confirmed_packing_capacity_qty",
    "allocation_rule",
    "active",
    "note",
]

COLD_STORAGE_STATUS_HEADERS = [
    "trade_date",
    "total_capacity_qty",
    "current_occupied_qty",
    "expected_inbound_qty",
    "expected_outbound_qty",
    "warning_threshold_qty",
    "projected_occupied_qty",
    "remaining_capacity_qty",
    "active",
    "note",
]

PLATFORM_MAPPING_HEADERS = [
    "mapping_id",
    "internal_sku",
    "platform_name",
    "platform_product_id",
    "platform_product_name",
    "search_keyword",
    "mapping_status",
    "last_verified_at",
    "remark",
]

TASK_HEADERS = [
    "task_id",
    "internal_sku",
    "platform_name",
    "action_type",
    "priority",
    "task_status",
    "created_at",
    "origin_type",
    "origin_ref_id",
    "approval_policy",
    "policy_version",
    "expected_old_price",
    "target_price",
    "target_inventory",
    "target_status",
    "pricing_source",
    "decision_trace",
    "result_message",
    "required_by",
]

EXECUTION_LOG_HEADERS = [
    "log_id",
    "task_id",
    "executor_name",
    "start_time",
    "end_time",
    "success_flag",
    "error_code",
    "error_message",
    "raw_output",
    "ai_model_version",
    "ai_summary",
]

TABLE_HEADERS = {
    "products": PRODUCT_HEADERS,
    "price_rules": PRICE_RULE_HEADERS,
    "listing_rules": LISTING_RULE_HEADERS,
    "harvest_forecasts": HARVEST_FORECAST_HEADERS,
    "price_forecasts": PRICE_FORECAST_HEADERS,
    "capacity_plans": CAPACITY_PLAN_HEADERS,
    "cold_storage_status": COLD_STORAGE_STATUS_HEADERS,
    "platform_mappings": PLATFORM_MAPPING_HEADERS,
}


def create_template_workbooks(output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    return [
        _create_template(output_dir / "products_template.xlsx", PRODUCT_HEADERS),
        _create_template(output_dir / "price_rules_template.xlsx", PRICE_RULE_HEADERS),
        _create_template(output_dir / "listing_rules_template.xlsx", LISTING_RULE_HEADERS),
        _create_template(output_dir / "harvest_forecasts_template.xlsx", HARVEST_FORECAST_HEADERS),
        _create_template(output_dir / "price_forecasts_template.xlsx", PRICE_FORECAST_HEADERS),
        _create_template(output_dir / "capacity_plans_template.xlsx", CAPACITY_PLAN_HEADERS),
        _create_template(output_dir / "cold_storage_status_template.xlsx", COLD_STORAGE_STATUS_HEADERS),
        _create_template(output_dir / "platform_mappings_template.xlsx", PLATFORM_MAPPING_HEADERS),
    ]


def _create_template(path: Path, headers: list[str]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(headers)
    workbook.save(path)
    return path


def load_products(path: Path) -> list[Product]:
    rows = _read_rows(path, PRODUCT_HEADERS)
    products: list[Product] = []
    seen_skus: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        sku = _required_text(row["internal_sku"], "internal_sku", row_number)
        if sku in seen_skus:
            raise ValidationError(f"products row {row_number}: duplicate internal_sku '{sku}'")
        seen_skus.add(sku)
        metadata = {
            "feature_season": row.get("feature_season") or "",
            "feature_color": row.get("feature_color") or "",
        }
        current_stock = _parse_int_or_default(
            row.get("current_stock"),
            field_name=f"products row {row_number} current_stock",
            default_value=0,
        )
        products.append(
            Product(
                internal_sku=sku,
                product_name=_required_text(row["product_name"], "product_name", row_number),
                grade=_required_text(row["grade"], "grade", row_number),
                stem_length=_required_text(row["stem_length"], "stem_length", row_number),
                unit=_required_text(row["unit"], "unit", row_number),
                base_cost=parse_decimal(row["base_cost"], f"products row {row_number} base_cost"),
                current_stock=current_stock,
                sale_enabled=parse_bool(row["sale_enabled"], f"products row {row_number} sale_enabled"),
                last_price=parse_decimal(row["last_price"], f"products row {row_number} last_price")
                if row.get("last_price") not in ("", None)
                else None,
                recommended_price=parse_decimal(
                    row["recommended_price"], f"products row {row_number} recommended_price"
                )
                if row.get("recommended_price") not in ("", None)
                else None,
                remark=str(row.get("remark") or ""),
                metadata=metadata,
            )
        )
    return products


def load_price_rules(path: Path) -> list[PriceRule]:
    rows = _read_rows(path, PRICE_RULE_HEADERS)
    return [_price_rule_from_row(row, row_number) for row_number, row in enumerate(rows, start=2)]


def load_price_rule(path: Path, rule_id: str) -> PriceRule:
    selected_id = str(rule_id or "").strip()
    matches = [
        (row_number, row)
        for row_number, row in enumerate(_read_rows(path, PRICE_RULE_HEADERS), start=2)
        if str(row.get("rule_id") or "").strip() == selected_id
    ]
    if not selected_id or not matches:
        raise ValidationError(f"未找到价格规则：{selected_id or '-'}")
    if len(matches) > 1:
        raise ValidationError(f"价格规则 ID 重复：{selected_id}")
    row_number, row = matches[0]
    return _price_rule_from_row(row, row_number)


def _price_rule_from_row(row: dict[str, object], row_number: int) -> PriceRule:
    return PriceRule(
        rule_id=_required_text(row["rule_id"], "rule_id", row_number),
        rule_name=_required_text(row["rule_name"], "rule_name", row_number),
        variety_filter=_required_text(row["variety_filter"], "variety_filter", row_number),
        grade_filter=_required_text(row["grade_filter"], "grade_filter", row_number),
        platform_filter=_required_text(row["platform_filter"], "platform_filter", row_number),
        pricing_method=PricingMethod(_required_text(row["pricing_method"], "pricing_method", row_number)),
        markup_value=_parse_required_non_zero_decimal(
            row["markup_value"], f"price_rules row {row_number} markup_value"
        ),
        min_price=parse_decimal(row["min_price"], f"price_rules row {row_number} min_price")
        if row.get("min_price") not in ("", None)
        else None,
        rounding_rule=RoundingRule(_required_text(row["rounding_rule"], "rounding_rule", row_number)),
        rounding_step=parse_decimal(row["rounding_step"], f"price_rules row {row_number} rounding_step")
        if row.get("rounding_step") not in ("", None)
        else None,
        active=parse_bool(row["active"], f"price_rules row {row_number} active"),
        priority=parse_int(row["priority"], f"price_rules row {row_number} priority"),
        remark=str(row.get("remark") or ""),
    )


def _parse_required_non_zero_decimal(value: object, field_context: str):
    if value in ("", None):
        raise ValidationError(f"{field_context} is required")
    parsed = parse_decimal(value, field_context)
    if parsed == 0:
        raise ValidationError(f"{field_context} must not be zero")
    return parsed


def load_listing_rules(path: Path) -> list[ListingRule]:
    rows = _read_rows(path, LISTING_RULE_HEADERS)
    return [_listing_rule_from_row(row, row_number) for row_number, row in enumerate(rows, start=2)]


def load_listing_rule(path: Path, rule_id: str) -> ListingRule:
    selected_id = str(rule_id or "").strip()
    matches = [
        (row_number, row)
        for row_number, row in enumerate(_read_rows(path, LISTING_RULE_HEADERS), start=2)
        if str(row.get("rule_id") or "").strip() == selected_id
    ]
    if not selected_id or not matches:
        raise ValidationError(f"未找到上下架规则：{selected_id or '-'}")
    if len(matches) > 1:
        raise ValidationError(f"上下架规则 ID 重复：{selected_id}")
    row_number, row = matches[0]
    return _listing_rule_from_row(row, row_number)


def _listing_rule_from_row(row: dict[str, object], row_number: int) -> ListingRule:
    return ListingRule(
        rule_id=_required_text(row["rule_id"], "rule_id", row_number),
        rule_name=_required_text(row["rule_name"], "rule_name", row_number),
        variety_filter=_required_text(row["variety_filter"], "variety_filter", row_number),
        grade_filter=_required_text(row["grade_filter"], "grade_filter", row_number),
        platform_filter=_required_text(row["platform_filter"], "platform_filter", row_number),
        stock_threshold=parse_decimal(row["stock_threshold"], f"listing_rules row {row_number} stock_threshold"),
        listing_strategy=ListingStrategy(_required_text(row["listing_strategy"], "listing_strategy", row_number)),
        active=parse_bool(row["active"], f"listing_rules row {row_number} active"),
        priority=parse_int(row["priority"], f"listing_rules row {row_number} priority"),
        remark=str(row.get("remark") or ""),
    )


def load_harvest_forecasts(path: Path) -> list[HarvestForecast]:
    rows = _read_rows(path, HARVEST_FORECAST_HEADERS)
    forecasts: list[HarvestForecast] = []
    for row_number, row in enumerate(rows, start=2):
        variety = _required_text(row["variety"], "variety", row_number)
        grade = _required_text(row["grade"], "grade", row_number)
        forecasts.append(
            HarvestForecast(
                forecast_id=_required_text(row["forecast_id"], "forecast_id", row_number),
                forecast_date=parse_date(row["forecast_date"], f"harvest_forecasts row {row_number} forecast_date"),
                target_trade_date=parse_date(
                    row["target_trade_date"], f"harvest_forecasts row {row_number} target_trade_date"
                ),
                forecast_group_key=forecast_group_key(variety, grade),
                variety=variety,
                grade=grade,
                predicted_harvest_qty=parse_int(
                    row["predicted_harvest_qty"],
                    f"harvest_forecasts row {row_number} predicted_harvest_qty",
                ),
                lower_bound_qty=parse_int(row["lower_bound_qty"], f"harvest_forecasts row {row_number} lower_bound_qty")
                if row.get("lower_bound_qty") not in ("", None)
                else None,
                upper_bound_qty=parse_int(row["upper_bound_qty"], f"harvest_forecasts row {row_number} upper_bound_qty")
                if row.get("upper_bound_qty") not in ("", None)
                else None,
                confidence=parse_decimal(row["confidence"], f"harvest_forecasts row {row_number} confidence")
                if row.get("confidence") not in ("", None)
                else None,
                source=str(row.get("source") or "manual"),
                generated_at=parse_datetime(row["generated_at"], f"harvest_forecasts row {row_number} generated_at")
                if row.get("generated_at") not in ("", None)
                else None,
                note=str(row.get("note") or ""),
            )
        )
    return forecasts


def load_price_forecasts(path: Path) -> list[PriceForecast]:
    rows = _read_rows(path, PRICE_FORECAST_HEADERS)
    forecasts: list[PriceForecast] = []
    for row_number, row in enumerate(rows, start=2):
        variety = _required_text(row["variety"], "variety", row_number)
        grade = _required_text(row["grade"], "grade", row_number)
        forecasts.append(
            PriceForecast(
                forecast_id=_required_text(row["forecast_id"], "forecast_id", row_number),
                forecast_date=parse_date(row["forecast_date"], f"price_forecasts row {row_number} forecast_date"),
                target_trade_date=parse_date(
                    row["target_trade_date"], f"price_forecasts row {row_number} target_trade_date"
                ),
                forecast_group_key=forecast_group_key(variety, grade),
                variety=variety,
                grade=grade,
                recommended_price=parse_decimal(
                    row["recommended_price"],
                    f"price_forecasts row {row_number} recommended_price",
                ),
                lower_bound_price=parse_decimal(
                    row["lower_bound_price"], f"price_forecasts row {row_number} lower_bound_price"
                )
                if row.get("lower_bound_price") not in ("", None)
                else None,
                upper_bound_price=parse_decimal(
                    row["upper_bound_price"], f"price_forecasts row {row_number} upper_bound_price"
                )
                if row.get("upper_bound_price") not in ("", None)
                else None,
                confidence=parse_decimal(row["confidence"], f"price_forecasts row {row_number} confidence")
                if row.get("confidence") not in ("", None)
                else None,
                source=str(row.get("source") or "manual"),
                generated_at=parse_datetime(row["generated_at"], f"price_forecasts row {row_number} generated_at")
                if row.get("generated_at") not in ("", None)
                else None,
                note=str(row.get("note") or ""),
            )
        )
    return forecasts


def load_capacity_plans(path: Path) -> list[PackingCapacityPlan]:
    rows = _read_capacity_plan_rows(path)
    if not rows:
        raise ValidationError(f"{path.name}: capacity plan requires at least one row")
    plans: list[PackingCapacityPlan] = []
    for row_number, row in enumerate(rows, start=2):
        confirmed_override = None
        if row.get("confirmed_packing_capacity_qty") not in ("", None):
            confirmed_override = _parse_non_negative_int_or_default(
                row.get("confirmed_packing_capacity_qty"),
                field_name=f"capacity_plans row {row_number} confirmed_packing_capacity_qty",
                default_value=250,
            )
        plans.append(
            PackingCapacityPlan(
                trade_date=parse_date(row["trade_date"], f"capacity_plans row {row_number} trade_date"),
                normal_packing_capacity_qty=_parse_non_negative_int_or_default(
                    row.get("normal_packing_capacity_qty"),
                    field_name=f"capacity_plans row {row_number} normal_packing_capacity_qty",
                    default_value=250,
                ),
                temp_worker_capacity_qty=_parse_non_negative_int_or_default(
                    row.get("temp_worker_capacity_qty"),
                    field_name=f"capacity_plans row {row_number} temp_worker_capacity_qty",
                    default_value=100,
                ),
                confirmed_temp_worker_count=_parse_non_negative_int_or_default(
                    row.get("confirmed_temp_worker_count"),
                    field_name=f"capacity_plans row {row_number} confirmed_temp_worker_count",
                    default_value=0,
                ),
                confirmed_packing_capacity_qty_override=confirmed_override,
                allocation_rule=str(row.get("allocation_rule") or "proportional_by_forecast"),
                active=parse_bool(row.get("active", True), f"capacity_plans row {row_number} active"),
                note=str(row.get("note") or ""),
            )
        )
    return plans


def load_capacity_plan(path: Path) -> PackingCapacityPlan:
    plans = load_capacity_plans(path)
    for plan in plans:
        if plan.active:
            return plan
    return plans[0]


def load_cold_storage_statuses(path: Path) -> list[ColdStorageStatus]:
    rows = _read_cold_storage_status_rows(path)
    if not rows:
        raise ValidationError(f"{path.name}: cold storage status requires at least one row")
    statuses: list[ColdStorageStatus] = []
    for row_number, row in enumerate(rows, start=2):
        projected_override = None
        if row.get("projected_occupied_qty") not in ("", None):
            projected_override = _parse_non_negative_int_or_default(
                row.get("projected_occupied_qty"),
                field_name=f"cold_storage_status row {row_number} projected_occupied_qty",
                default_value=0,
            )
        remaining_override = None
        if row.get("remaining_capacity_qty") not in ("", None):
            remaining_override = _parse_int_or_default(
                row.get("remaining_capacity_qty"),
                field_name=f"cold_storage_status row {row_number} remaining_capacity_qty",
                default_value=0,
            )
        total_capacity = _parse_non_negative_int_or_default(
            row.get("total_capacity_qty"),
            field_name=f"cold_storage_status row {row_number} total_capacity_qty",
            default_value=500,
        )
        if total_capacity <= 0:
            raise ValidationError(f"cold_storage_status row {row_number} total_capacity_qty must be greater than 0")
        statuses.append(
            ColdStorageStatus(
                trade_date=parse_date(row["trade_date"], f"cold_storage_status row {row_number} trade_date"),
                total_capacity_qty=total_capacity,
                current_occupied_qty=_parse_non_negative_int_or_default(
                    row.get("current_occupied_qty"),
                    field_name=f"cold_storage_status row {row_number} current_occupied_qty",
                    default_value=0,
                ),
                expected_inbound_qty=_parse_non_negative_int_or_default(
                    row.get("expected_inbound_qty"),
                    field_name=f"cold_storage_status row {row_number} expected_inbound_qty",
                    default_value=0,
                ),
                expected_outbound_qty=_parse_non_negative_int_or_default(
                    row.get("expected_outbound_qty"),
                    field_name=f"cold_storage_status row {row_number} expected_outbound_qty",
                    default_value=0,
                ),
                warning_threshold_qty=_parse_non_negative_int_or_default(
                    row.get("warning_threshold_qty"),
                    field_name=f"cold_storage_status row {row_number} warning_threshold_qty",
                    default_value=50,
                ),
                projected_occupied_qty_override=projected_override,
                remaining_capacity_qty_override=remaining_override,
                active=parse_bool(row.get("active", True), f"cold_storage_status row {row_number} active"),
                note=str(row.get("note") or ""),
            )
        )
    return statuses


def load_cold_storage_status(path: Path) -> ColdStorageStatus:
    statuses = load_cold_storage_statuses(path)
    for status in statuses:
        if status.active:
            return status
    return statuses[0]


def export_tasks(path: Path, tasks: Iterable[Task]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "tasks"
    sheet.append(TASK_HEADERS)
    for task in tasks:
        record = task.to_record()
        sheet.append(
            [
                record["task_id"],
                record["internal_sku"],
                record["platform_name"],
                record["action_type"],
                record["priority"],
                record["task_status"],
                record["created_at"],
                record["origin_type"],
                record["origin_ref_id"],
                record["approval_policy"],
                record["policy_version"],
                serialize_decimal(task.expected_old_price),
                serialize_decimal(task.target_price),
                task.target_inventory,
                record["target_status"],
                record["pricing_source"],
                str(record["decision_trace"]),
                record["result_message"],
                record["required_by"],
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def load_tasks(path: Path) -> list[Task]:
    rows = _read_task_rows(path)
    tasks: list[Task] = []
    for row_number, row in enumerate(rows, start=2):
        decision_trace: dict[str, object] = {}
        raw_trace = row.get("decision_trace")
        if raw_trace not in ("", None):
            parsed_trace = ast.literal_eval(str(raw_trace))
            if isinstance(parsed_trace, dict):
                decision_trace = parsed_trace
        origin_type = TaskOriginType(
            _required_text(
                row["origin_type"],
                "origin_type",
                row_number,
            )
        )
        origin_ref_id = (
            str(row["origin_ref_id"]).strip()
            if row.get("origin_ref_id") not in ("", None)
            else None
        )
        if (
            origin_type is TaskOriginType.AUTOMATION
            and origin_ref_id is None
        ):
            raise ValidationError(
                f"row {row_number}: AUTOMATION requires origin_ref_id"
            )
        tasks.append(
            Task(
                task_id=_required_text(row["task_id"], "task_id", row_number),
                internal_sku=_required_text(row["internal_sku"], "internal_sku", row_number),
                platform_name=_required_text(row["platform_name"], "platform_name", row_number),
                action_type=TaskActionType(_required_text(row["action_type"], "action_type", row_number)),
                priority=parse_int(row["priority"], f"tasks row {row_number} priority"),
                task_status=TaskStatus(_required_text(row["task_status"], "task_status", row_number)),
                created_at=datetime.fromisoformat(_required_text(row["created_at"], "created_at", row_number)),
                origin_type=origin_type,
                origin_ref_id=origin_ref_id,
                approval_policy=str(
                    row.get("approval_policy") or "UNSPECIFIED"
                ),
                policy_version=(
                    str(row["policy_version"]).strip()
                    if row.get("policy_version") not in ("", None)
                    else None
                ),
                expected_old_price=parse_decimal(
                    row["expected_old_price"], f"tasks row {row_number} expected_old_price"
                )
                if row.get("expected_old_price") not in ("", None)
                else None,
                target_price=parse_decimal(row["target_price"], f"tasks row {row_number} target_price")
                if row.get("target_price") not in ("", None)
                else None,
                target_inventory=parse_int(
                    row["target_inventory"],
                    f"tasks row {row_number} target_inventory",
                )
                if row.get("target_inventory") not in ("", None)
                else None,
                target_status=str(row["target_status"]) if row.get("target_status") not in ("", None) else None,
                pricing_source=PricingSource(str(row["pricing_source"]))
                if row.get("pricing_source") not in ("", None)
                else None,
                decision_trace=decision_trace,
                result_message=str(row.get("result_message") or ""),
                required_by=datetime.fromisoformat(str(row["required_by"]))
                if row.get("required_by") not in ("", None)
                else None,
            )
        )
    return tasks


def export_execution_logs(path: Path, logs: Iterable[ExecutionLog]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "execution_logs"
    sheet.append(EXECUTION_LOG_HEADERS)
    for log in logs:
        sheet.append(
            [
                log.log_id or uuid4().hex[:12],
                log.task_id,
                log.executor_name,
                log.start_time.isoformat(),
                log.end_time.isoformat() if log.end_time is not None else None,
                log.success_flag,
                log.error_code,
                log.error_message,
                log.raw_output,
                log.ai_model_version,
                log.ai_summary,
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def load_table_records(table_name: str, path: Path) -> list[dict[str, object]]:
    headers = get_table_headers(table_name)
    if table_name == "capacity_plans":
        return _read_capacity_plan_rows(path)
    if table_name == "cold_storage_status":
        return _read_cold_storage_status_rows(path)
    return _read_rows(path, headers)


def save_table_records(table_name: str, path: Path, rows: list[dict[str, object]]) -> Path:
    headers = get_table_headers(table_name)
    _validate_table_records(table_name, rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def get_table_headers(table_name: str) -> list[str]:
    try:
        return TABLE_HEADERS[table_name]
    except KeyError as exc:
        raise ValidationError(f"unsupported table '{table_name}'") from exc


def _validate_table_records(table_name: str, rows: list[dict[str, object]]) -> None:
    issues = _collect_table_validation_issues(table_name, rows)
    if issues:
        raise TableValidationError(table_name, issues)

    headers = get_table_headers(table_name)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        temp_path = Path(handle.name)
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "data"
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        workbook.save(temp_path)
        if table_name == "products":
            load_products(temp_path)
        elif table_name == "price_rules":
            load_price_rules(temp_path)
        elif table_name == "listing_rules":
            load_listing_rules(temp_path)
        elif table_name == "harvest_forecasts":
            load_harvest_forecasts(temp_path)
        elif table_name == "price_forecasts":
            load_price_forecasts(temp_path)
        elif table_name == "capacity_plans":
            load_capacity_plan(temp_path)
        elif table_name == "cold_storage_status":
            load_cold_storage_status(temp_path)
        elif table_name == "platform_mappings":
            load_table_records("platform_mappings", temp_path)
    finally:
        temp_path.unlink(missing_ok=True)


def _collect_table_validation_issues(table_name: str, rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    if table_name == "products":
        return _collect_product_issues(rows)
    if table_name == "price_rules":
        return _collect_price_rule_issues(rows)
    if table_name == "listing_rules":
        return _collect_listing_rule_issues(rows)
    if table_name == "harvest_forecasts":
        return _collect_harvest_forecast_issues(rows)
    if table_name == "price_forecasts":
        return _collect_price_forecast_issues(rows)
    if table_name == "capacity_plans":
        return _collect_capacity_plan_issues(rows)
    if table_name == "cold_storage_status":
        return _collect_cold_storage_issues(rows)
    if table_name == "platform_mappings":
        return _collect_platform_mapping_issues(rows)
    raise ValidationError(f"unsupported table '{table_name}'")


def _collect_platform_mapping_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    seen_mapping_ids: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=2):
        mapping_id = str(row.get("mapping_id") or "").strip()
        platform_name = str(row.get("platform_name") or "").strip()
        if not mapping_id:
            issues.append(TableValidationIssue(row_number, "mapping_id", "该字段必填"))
        elif mapping_id in seen_mapping_ids:
            issues.append(TableValidationIssue(row_number, "mapping_id", f"与第 {seen_mapping_ids[mapping_id]} 行重复"))
        else:
            seen_mapping_ids[mapping_id] = row_number
        if not platform_name:
            issues.append(TableValidationIssue(row_number, "platform_name", "该字段必填"))
    return issues


def _collect_product_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    seen_skus: dict[str, int] = {}
    required_fields = ["internal_sku", "product_name", "grade", "stem_length", "unit", "base_cost", "sale_enabled"]

    for row_number, row in enumerate(rows, start=2):
        for field_name in required_fields:
            value = row.get(field_name)
            if value is None or str(value).strip() == "":
                issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))

        sku = str(row.get("internal_sku") or "").strip()
        if sku:
            if sku in seen_skus:
                issues.append(TableValidationIssue(row_number, "internal_sku", f"与第 {seen_skus[sku]} 行重复"))
            else:
                seen_skus[sku] = row_number

        _try_parse_decimal(row.get("base_cost"), row_number, "base_cost", issues, required=True)
        _try_parse_int(row.get("current_stock"), row_number, "current_stock", issues, required=False)
        _try_parse_bool(row.get("sale_enabled"), row_number, "sale_enabled", issues, required=True)
        _try_parse_decimal(row.get("last_price"), row_number, "last_price", issues, required=False)
        _try_parse_decimal(row.get("recommended_price"), row_number, "recommended_price", issues, required=False)

    return issues


def _collect_price_rule_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    required_fields = [
        "rule_id",
        "rule_name",
        "variety_filter",
        "grade_filter",
        "platform_filter",
        "pricing_method",
        "markup_value",
        "rounding_rule",
    ]

    for row_number, row in enumerate(rows, start=2):
        for field_name in required_fields:
            value = row.get(field_name)
            if value is None or str(value).strip() == "":
                issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))

        _try_enum(row.get("pricing_method"), PricingMethod, row_number, "pricing_method", issues)
        grade_filter = str(row.get("grade_filter") or "").strip().upper()
        if grade_filter not in {"*", "A", "B", "C", "D", "E", "0"}:
            issues.append(TableValidationIssue(row_number, "grade_filter", "必须为 * / A / B / C / D / E / 0"))
        _try_parse_decimal(row.get("markup_value"), row_number, "markup_value", issues, required=True)
        try:
            if row.get("markup_value") not in ("", None) and parse_decimal(row.get("markup_value"), "") == 0:
                issues.append(TableValidationIssue(row_number, "markup_value", "不能为 0；正数表示涨价，负数表示降价"))
        except ValidationError:
            pass
        _try_parse_decimal(row.get("min_price"), row_number, "min_price", issues, required=False)
        _try_enum(row.get("rounding_rule"), RoundingRule, row_number, "rounding_rule", issues)
        _try_parse_decimal(row.get("rounding_step"), row_number, "rounding_step", issues, required=False)
        _try_parse_bool(row.get("active"), row_number, "active", issues, required=True)
        _try_parse_int(row.get("priority"), row_number, "priority", issues, required=True)

    return issues


def _collect_listing_rule_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    required_fields = [
        "rule_id",
        "rule_name",
        "variety_filter",
        "grade_filter",
        "platform_filter",
        "stock_threshold",
        "listing_strategy",
        "active",
        "priority",
    ]

    for row_number, row in enumerate(rows, start=2):
        for field_name in required_fields:
            value = row.get(field_name)
            if value is None or str(value).strip() == "":
                issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))

        grade_filter = str(row.get("grade_filter") or "").strip().upper()
        if grade_filter not in {"*", "A", "B", "C", "D", "E", "0"}:
            issues.append(TableValidationIssue(row_number, "grade_filter", "必须为 * / A / B / C / D / E / 0"))
        _try_parse_decimal(row.get("stock_threshold"), row_number, "stock_threshold", issues, required=True)
        try:
            if row.get("stock_threshold") not in ("", None) and parse_decimal(row.get("stock_threshold"), "") < 0:
                issues.append(TableValidationIssue(row_number, "stock_threshold", "必须大于或等于 0"))
        except ValidationError:
            pass
        _try_enum(row.get("listing_strategy"), ListingStrategy, row_number, "listing_strategy", issues)
        _try_parse_bool(row.get("active"), row_number, "active", issues, required=True)
        _try_parse_int(row.get("priority"), row_number, "priority", issues, required=True)
        try:
            if row.get("priority") not in ("", None) and parse_int(row.get("priority"), "") < 0:
                issues.append(TableValidationIssue(row_number, "priority", "必须大于或等于 0"))
        except ValidationError:
            pass

    return issues


def _collect_harvest_forecast_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    required_fields = ["forecast_id", "forecast_date", "target_trade_date", "variety", "grade", "predicted_harvest_qty"]
    seen_groups: dict[tuple[str, str, str], int] = {}

    for row_number, row in enumerate(rows, start=2):
        for field_name in required_fields:
            value = row.get(field_name)
            if value is None or str(value).strip() == "":
                issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        _try_parse_date(row.get("forecast_date"), row_number, "forecast_date", issues, required=True)
        _try_parse_date(row.get("target_trade_date"), row_number, "target_trade_date", issues, required=True)
        _try_parse_int(row.get("predicted_harvest_qty"), row_number, "predicted_harvest_qty", issues, required=True)
        _try_parse_int(row.get("lower_bound_qty"), row_number, "lower_bound_qty", issues, required=False)
        _try_parse_int(row.get("upper_bound_qty"), row_number, "upper_bound_qty", issues, required=False)
        _try_parse_decimal(row.get("confidence"), row_number, "confidence", issues, required=False)
        _try_parse_datetime(row.get("generated_at"), row_number, "generated_at", issues, required=False)
        group = (
            str(row.get("target_trade_date") or "").strip(),
            str(row.get("variety") or "").strip(),
            str(row.get("grade") or "").strip(),
        )
        if all(group):
            if group in seen_groups:
                issues.append(TableValidationIssue(row_number, "forecast_group_key", f"与第 {seen_groups[group]} 行重复"))
            else:
                seen_groups[group] = row_number
    return issues


def _collect_price_forecast_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    required_fields = [
        "forecast_id",
        "forecast_date",
        "target_trade_date",
        "variety",
        "grade",
        "recommended_price",
    ]
    seen_groups: dict[tuple[str, str, str], int] = {}

    for row_number, row in enumerate(rows, start=2):
        for field_name in required_fields:
            value = row.get(field_name)
            if value is None or str(value).strip() == "":
                issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        _try_parse_date(row.get("forecast_date"), row_number, "forecast_date", issues, required=True)
        _try_parse_date(row.get("target_trade_date"), row_number, "target_trade_date", issues, required=True)
        _try_parse_decimal(row.get("recommended_price"), row_number, "recommended_price", issues, required=True)
        _try_parse_decimal(row.get("lower_bound_price"), row_number, "lower_bound_price", issues, required=False)
        _try_parse_decimal(row.get("upper_bound_price"), row_number, "upper_bound_price", issues, required=False)
        _try_parse_decimal(row.get("confidence"), row_number, "confidence", issues, required=False)
        _try_parse_datetime(row.get("generated_at"), row_number, "generated_at", issues, required=False)
        group = (
            str(row.get("target_trade_date") or "").strip(),
            str(row.get("variety") or "").strip(),
            str(row.get("grade") or "").strip(),
        )
        if all(group):
            if group in seen_groups:
                issues.append(TableValidationIssue(row_number, "forecast_group_key", f"与第 {seen_groups[group]} 行重复"))
            else:
                seen_groups[group] = row_number
    return issues


def _collect_capacity_plan_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    if not rows:
        issues.append(TableValidationIssue(2, "trade_date", "该字段必填"))
        return issues
    active_trade_dates: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=2):
        _try_parse_date(row.get("trade_date"), row_number, "trade_date", issues, required=True)
        _try_parse_non_negative_int(
            row.get("normal_packing_capacity_qty"),
            row_number,
            "normal_packing_capacity_qty",
            issues,
            required=False,
        )
        _try_parse_non_negative_int(row.get("temp_worker_capacity_qty"), row_number, "temp_worker_capacity_qty", issues, required=False)
        _try_parse_non_negative_int(row.get("confirmed_temp_worker_count"), row_number, "confirmed_temp_worker_count", issues, required=False)
        _try_parse_non_negative_int(
            row.get("confirmed_packing_capacity_qty"),
            row_number,
            "confirmed_packing_capacity_qty",
            issues,
            required=False,
        )
        active_value = row.get("active", True)
        try:
            is_active = parse_bool(active_value, f"capacity_plans row {row_number} active")
        except ValidationError:
            issues.append(TableValidationIssue(row_number, "active", "请输入是或否"))
            is_active = False
        trade_date_value = str(row.get("trade_date") or "").strip()
        if is_active and trade_date_value:
            if trade_date_value in active_trade_dates:
                issues.append(
                    TableValidationIssue(
                        row_number,
                        "trade_date",
                        f"与第 {active_trade_dates[trade_date_value]} 行存在同一业务日期的启用计划",
                    )
                )
            else:
                active_trade_dates[trade_date_value] = row_number
    return issues


def _collect_cold_storage_issues(rows: list[dict[str, object]]) -> list[TableValidationIssue]:
    issues: list[TableValidationIssue] = []
    if not rows:
        issues.append(TableValidationIssue(2, "trade_date", "该字段必填"))
        return issues
    active_trade_dates: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=2):
        _try_parse_date(row.get("trade_date"), row_number, "trade_date", issues, required=True)
        _try_parse_positive_int(row.get("total_capacity_qty"), row_number, "total_capacity_qty", issues, required=False)
        _try_parse_non_negative_int(row.get("current_occupied_qty"), row_number, "current_occupied_qty", issues, required=False)
        _try_parse_non_negative_int(row.get("expected_inbound_qty"), row_number, "expected_inbound_qty", issues, required=False)
        _try_parse_non_negative_int(row.get("expected_outbound_qty"), row_number, "expected_outbound_qty", issues, required=False)
        _try_parse_non_negative_int(row.get("warning_threshold_qty"), row_number, "warning_threshold_qty", issues, required=False)
        _try_parse_non_negative_int(row.get("projected_occupied_qty"), row_number, "projected_occupied_qty", issues, required=False)
        _try_parse_int(row.get("remaining_capacity_qty"), row_number, "remaining_capacity_qty", issues, required=False)
        active_value = row.get("active", True)
        try:
            is_active = parse_bool(active_value, f"cold_storage_status row {row_number} active")
        except ValidationError:
            issues.append(TableValidationIssue(row_number, "active", "请输入是或否"))
            is_active = False
        trade_date_value = str(row.get("trade_date") or "").strip()
        if is_active and trade_date_value:
            if trade_date_value in active_trade_dates:
                issues.append(
                    TableValidationIssue(
                        row_number,
                        "trade_date",
                        f"与第 {active_trade_dates[trade_date_value]} 行存在同一业务日期的启用冷库状态",
                    )
                )
            else:
                active_trade_dates[trade_date_value] = row_number
    return issues


def _try_enum(value: object, enum_cls, row_number: int, field_name: str, issues: list[TableValidationIssue]) -> None:
    if value in (None, ""):
        return
    try:
        enum_cls(str(value).strip())
    except ValueError:
        allowed = ", ".join(item.value for item in enum_cls)
        issues.append(TableValidationIssue(row_number, field_name, f"值无效，可选：{allowed}"))


def _try_parse_decimal(
    value: object,
    row_number: int,
    field_name: str,
    issues: list[TableValidationIssue],
    *,
    required: bool,
) -> None:
    if value in (None, ""):
        if required:
            issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        return
    try:
        parse_decimal(value, f"row {row_number} {field_name}")
    except ValidationError:
        issues.append(TableValidationIssue(row_number, field_name, "请输入数字"))


def _try_parse_int(
    value: object,
    row_number: int,
    field_name: str,
    issues: list[TableValidationIssue],
    *,
    required: bool,
) -> None:
    if value in (None, ""):
        if required:
            issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        return
    try:
        parse_int(value, f"row {row_number} {field_name}")
    except ValidationError:
        issues.append(TableValidationIssue(row_number, field_name, "请输入整数"))


def _try_parse_non_negative_int(
    value: object,
    row_number: int,
    field_name: str,
    issues: list[TableValidationIssue],
    *,
    required: bool,
) -> None:
    if value in (None, ""):
        if required:
            issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        return
    try:
        parsed = parse_int(value, f"row {row_number} {field_name}")
    except ValidationError:
        issues.append(TableValidationIssue(row_number, field_name, "请输入整数"))
        return
    if parsed < 0:
        issues.append(TableValidationIssue(row_number, field_name, "不能小于 0"))


def _try_parse_positive_int(
    value: object,
    row_number: int,
    field_name: str,
    issues: list[TableValidationIssue],
    *,
    required: bool,
) -> None:
    if value in (None, ""):
        if required:
            issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        return
    try:
        parsed = parse_int(value, f"row {row_number} {field_name}")
    except ValidationError:
        issues.append(TableValidationIssue(row_number, field_name, "请输入整数"))
        return
    if parsed <= 0:
        issues.append(TableValidationIssue(row_number, field_name, "必须大于 0"))


def _try_parse_bool(
    value: object,
    row_number: int,
    field_name: str,
    issues: list[TableValidationIssue],
    *,
    required: bool,
) -> None:
    if value in (None, ""):
        if required:
            issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        return
    try:
        parse_bool(value, f"row {row_number} {field_name}")
    except ValidationError:
        issues.append(TableValidationIssue(row_number, field_name, "请输入 true/false、1/0、yes/no"))


def _try_parse_date(
    value: object,
    row_number: int,
    field_name: str,
    issues: list[TableValidationIssue],
    *,
    required: bool,
) -> None:
    if value in (None, ""):
        if required:
            issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        return
    try:
        parse_date(value, f"row {row_number} {field_name}")
    except ValidationError:
        issues.append(TableValidationIssue(row_number, field_name, "请输入 YYYY-MM-DD 日期"))


def _try_parse_datetime(
    value: object,
    row_number: int,
    field_name: str,
    issues: list[TableValidationIssue],
    *,
    required: bool,
) -> None:
    if value in (None, ""):
        if required:
            issues.append(TableValidationIssue(row_number, field_name, "该字段必填"))
        return
    try:
        parse_datetime(value, f"row {row_number} {field_name}")
    except ValidationError:
        issues.append(TableValidationIssue(row_number, field_name, "请输入 ISO 时间"))

def _read_rows(path: Path, expected_headers: list[str]) -> list[dict[str, object]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if header_row != expected_headers:
        raise ValidationError(f"{path.name}: invalid headers. Expected {expected_headers}, got {header_row}")
    rows: list[dict[str, object]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        if all(_is_blank_cell(value) for value in raw_row):
            continue
        rows.append(dict(zip(expected_headers, raw_row, strict=True)))
    return rows


def _read_capacity_plan_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    legacy_headers = [
        "trade_date",
        "normal_packing_capacity_qty",
        "temp_worker_capacity_qty",
        "confirmed_temp_worker_count",
        "allocation_rule",
        "note",
    ]
    if header_row == CAPACITY_PLAN_HEADERS:
        accepted_headers = CAPACITY_PLAN_HEADERS
    elif header_row == legacy_headers:
        accepted_headers = legacy_headers
    else:
        raise ValidationError(f"{path.name}: invalid headers. Expected {CAPACITY_PLAN_HEADERS}, got {header_row}")

    rows: list[dict[str, object]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        if all(_is_blank_cell(value) for value in raw_row):
            continue
        row = dict(zip(accepted_headers, raw_row, strict=True))
        row.setdefault("confirmed_packing_capacity_qty", "")
        row.setdefault("active", True)
        rows.append(row)
    return rows


def _read_cold_storage_status_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    legacy_headers = [
        "trade_date",
        "cold_storage_total_capacity_qty",
        "cold_storage_current_qty",
        "note",
    ]
    if header_row == COLD_STORAGE_STATUS_HEADERS:
        accepted_headers = COLD_STORAGE_STATUS_HEADERS
    elif header_row == legacy_headers:
        accepted_headers = legacy_headers
    else:
        raise ValidationError(
            f"{path.name}: invalid headers. Expected {COLD_STORAGE_STATUS_HEADERS}, got {header_row}"
        )

    rows: list[dict[str, object]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        if all(_is_blank_cell(value) for value in raw_row):
            continue
        row = dict(zip(accepted_headers, raw_row, strict=True))
        if accepted_headers == legacy_headers:
            row = {
                "trade_date": row.get("trade_date"),
                "total_capacity_qty": row.get("cold_storage_total_capacity_qty"),
                "current_occupied_qty": row.get("cold_storage_current_qty"),
                "expected_inbound_qty": 0,
                "expected_outbound_qty": 0,
                "warning_threshold_qty": 50,
                "projected_occupied_qty": "",
                "remaining_capacity_qty": "",
                "active": True,
                "note": row.get("note") or "",
            }
        else:
            row.setdefault("expected_inbound_qty", 0)
            row.setdefault("expected_outbound_qty", 0)
            row.setdefault("warning_threshold_qty", 50)
            row.setdefault("projected_occupied_qty", "")
            row.setdefault("remaining_capacity_qty", "")
            row.setdefault("active", True)
            row.setdefault("note", "")
        rows.append(row)
    return rows


def _read_task_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    accepted_headers = TASK_HEADERS
    legacy_optional_headers = (
        ("target_inventory",),
        ("required_by",),
        ("expected_old_price",),
        ("target_inventory", "required_by"),
        ("target_inventory", "expected_old_price"),
        ("expected_old_price", "required_by"),
        ("target_inventory", "expected_old_price", "required_by"),
    )
    legacy_origin_headers = (
        "origin_type",
        "origin_ref_id",
        "approval_policy",
        "policy_version",
    )
    legacy_header_variants = [
        [header for header in TASK_HEADERS if header not in omitted]
        for omitted in legacy_optional_headers
    ]
    legacy_header_variants.extend(
        [
            [
                header
                for header in TASK_HEADERS
                if header not in set(omitted) | set(legacy_origin_headers)
            ]
            for omitted in ((), *legacy_optional_headers)
        ]
    )
    if header_row in legacy_header_variants:
        accepted_headers = header_row
    elif header_row != TASK_HEADERS:
        raise ValidationError(f"{path.name}: invalid headers. Expected {TASK_HEADERS}, got {header_row}")

    rows: list[dict[str, object]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        if all(_is_blank_cell(value) for value in raw_row):
            continue
        row = dict(zip(accepted_headers, raw_row, strict=True))
        row.setdefault("expected_old_price", None)
        row.setdefault("target_inventory", None)
        row.setdefault("required_by", None)
        row.setdefault("origin_type", TaskOriginType.LEGACY.value)
        row.setdefault("origin_ref_id", None)
        row.setdefault("approval_policy", "UNSPECIFIED")
        row.setdefault("policy_version", None)
        rows.append(row)
    return rows


def _required_text(value: object, field_name: str, row_number: int) -> str:
    if value is None or str(value).strip() == "":
        raise ValidationError(f"row {row_number}: {field_name} is required")
    return str(value).strip()


def _parse_int_or_default(value: object, *, field_name: str, default_value: int) -> int:
    if value in (None, ""):
        return default_value
    return parse_int(value, field_name)


def _parse_non_negative_int_or_default(value: object, *, field_name: str, default_value: int) -> int:
    parsed = _parse_int_or_default(value, field_name=field_name, default_value=default_value)
    if parsed < 0:
        raise ValidationError(f"{field_name} must be greater than or equal to 0")
    return parsed


def _is_blank_cell(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False

