from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.repositories.workbook_repository import PRICE_RULE_HEADERS


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PRICE_RULES = ROOT / "data" / "samples" / "price_rules.xlsx"
DEFAULT_BACKUP = ROOT / "data" / "samples" / "price_rules_backup_before_scope_refactor.xlsx"


def migrate_price_rules_scope(path: Path = DEFAULT_PRICE_RULES, backup_path: Path = DEFAULT_BACKUP) -> None:
    """One-off migration from scope_type/scope_value to three price rule filters."""
    if not path.exists():
        raise FileNotFoundError(f"price rule workbook not found: {path}")

    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]

    if {"variety_filter", "grade_filter", "platform_filter"}.issubset(set(headers)):
        return

    if not {"scope_type", "scope_value"}.issubset(set(headers)):
        raise ValueError("workbook has neither new filter fields nor legacy scope_type/scope_value fields")

    if not backup_path.exists():
        shutil.copy2(path, backup_path)

    rows = []
    for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {header: value for header, value in zip(headers, row_values)}
        scope_type = str(row.get("scope_type") or "").strip().lower()
        scope_value = str(row.get("scope_value") or "").strip()
        variety_filter, grade_filter, platform_filter = _convert_scope(scope_type, scope_value, row_index)
        rows.append(
            {
                "rule_id": row.get("rule_id") or "",
                "rule_name": row.get("rule_name") or "",
                "variety_filter": variety_filter,
                "grade_filter": grade_filter,
                "platform_filter": platform_filter,
                "pricing_method": row.get("pricing_method") or "",
                "markup_value": row.get("markup_value") if row.get("markup_value") is not None else "",
                "min_price": row.get("min_price") if row.get("min_price") is not None else "",
                "rounding_rule": row.get("rounding_rule") or "",
                "rounding_step": row.get("rounding_step") if row.get("rounding_step") is not None else "",
                "active": row.get("active") if row.get("active") is not None else "",
                "priority": row.get("priority") if row.get("priority") is not None else "",
                "remark": row.get("remark") if row.get("remark") is not None else "",
            }
        )

    output = Workbook()
    output_sheet = output.active
    output_sheet.title = "price_rules"
    output_sheet.append(PRICE_RULE_HEADERS)
    for row in rows:
        output_sheet.append([row.get(header, "") for header in PRICE_RULE_HEADERS])
    output.save(path)


def _convert_scope(scope_type: str, scope_value: str, row_index: int) -> tuple[str, str, str]:
    if scope_type == "all":
        return "*", "*", "*"
    if scope_type == "grade":
        return "*", scope_value.upper(), "*"
    if scope_type in {"variety", "product_name", "product"}:
        return scope_value, "*", "*"
    if scope_type == "platform":
        return "*", "*", scope_value
    if scope_type == "sku":
        raise ValueError(f"row {row_index}: sku scoped price rule requires manual handling")
    raise ValueError(f"row {row_index}: unsupported legacy scope_type '{scope_type}'")


if __name__ == "__main__":
    migrate_price_rules_scope()
