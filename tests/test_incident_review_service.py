from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from threading import Event, Thread

import pytest
from openpyxl import Workbook

from app.enums import (
    IncidentEventType,
    IncidentStatus,
    ReviewTaskStatus,
    TaskActionType,
    TaskOriginType,
    TaskStatus,
)
from app.exceptions import (
    MobileReviewTransactionError,
    NotificationIdempotencyConflictError,
)
from app.models import ListingStatus, NotificationDeliveryResult, Task
from app.operations_web.auth import (
    Capability,
    Principal,
    PrincipalCapabilityBackend,
)
from app.operations_web.composition import OperationsWebPaths
from app.operations_web.presenters import render_mobile_review
from app.operations_web.queries import OperationsQueryService
from app.repositories.sqlite_runtime_repository import SQLiteRuntimeRepository
from app.repositories.workbook_repository import PRODUCT_HEADERS
from app.services.incident_management import (
    IncidentDetection,
    IncidentManagementService,
    IncidentNotificationService,
    IncidentReviewService,
)
from app.services.notification_outbox import (
    FeishuOutboxSender,
    NotificationOutboxService,
)
from app.services.runtime import ReviewTokenService
from app.services.review_resolution import ReviewResolutionApplicationService
from app.services.workflow import (
    _read_authoritative_product_cost_snapshot,
    resolve_mobile_review,
)
from tests.test_incident_management import detection

NOW = datetime(2026, 8, 2, 12, 0, tzinfo=timezone.utc)


def stable_product_snapshot() -> tuple[Decimal, str]:
    return Decimal("10.00"), "products.xlsx:sha256:synthetic"


@pytest.fixture
def runtime_repository(tmp_path, monkeypatch):
    monkeypatch.setenv(
        "REVIEW_TOKEN_SECRET", "incident-review-test-secret-at-least-32-bytes"
    )
    monkeypatch.setenv("MOBILE_REVIEW_BASE_URL", "https://example.test")
    monkeypatch.setenv("DEFAULT_NOTIFICATION_CHANNEL", "fake")
    repository = SQLiteRuntimeRepository(tmp_path / "runtime.db")
    repository.init_schema()
    return repository


def create_s4_incident(repository: SQLiteRuntimeRepository):
    incident_detection: IncidentDetection = replace(
        detection("detect-s4", occurred_at=NOW),
        severity="S4",
        subject_type="internal_sku",
        subject_key="SKU-SYNTHETIC-001",
        title="Synthetic extreme price",
    )
    return IncidentManagementService(repository).detect(incident_detection).incident


def mark_notification(
    repository: SQLiteRuntimeRepository,
    notification_id: str,
    *,
    now: datetime,
    classification: str = "SUCCESS",
):
    claimed_rows = repository.claim_notification_outbox(now=now, channel="fake")
    claimed = next(
        row for row in claimed_rows if row.notification_id == notification_id
    )
    attempt = repository.begin_notification_delivery(
        claimed.notification_id,
        owner_token=claimed.lease_owner_token,
        lease_version=claimed.lease_version,
        request_fingerprint="synthetic-request",
        now=now,
    )
    return repository.complete_notification_delivery(
        claimed.notification_id,
        attempt.delivery_attempt_id,
        owner_token=claimed.lease_owner_token,
        lease_version=claimed.lease_version,
        result=NotificationDeliveryResult(
            classification=classification,
            provider_message_id="synthetic-provider-message",
            error_code=("" if classification == "SUCCESS" else "SYNTHETIC_FAILURE"),
            error_message=("" if classification == "SUCCESS" else "synthetic failure"),
        ),
        now=now,
    )


def create_review_with_listing(repository: SQLiteRuntimeRepository):
    incident = create_s4_incident(repository)
    repository.upsert_listing_status(
        ListingStatus(
            listing_status_id="listing-synthetic-001",
            platform_name="synthetic-platform",
            internal_sku="SKU-SYNTHETIC-001",
            variety="Synthetic",
            grade="B",
            current_price=Decimal("8.00"),
            online_status="online",
            price_source="shadowbot_read_only",
            price_observed_at=NOW,
            price_source_attempt_id="read-attempt-001",
            updated_at=NOW,
        )
    )
    result = IncidentReviewService(repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    return incident, result


def insert_pulse_observation_facts(
    repository: SQLiteRuntimeRepository,
    *,
    suffix: str,
    scheduled_for: datetime,
    scan_started_at: datetime,
    scan_completed_at: datetime,
    observed_price: str = "6.00",
    observed_online: bool = True,
    mapping_status: str = "VERIFIED",
    scope_complete: bool = True,
    end_marker_verified: bool = True,
    merged_coverage: bool = False,
) -> tuple[str, str]:
    pulse_run_id = f"pulse-{suffix}"
    observation_run_id = (
        f"listing-coverage-{suffix}" if merged_coverage else pulse_run_id
    )
    batch_id = f"batch-{suffix}"
    item_id = f"observation-{suffix}"
    with repository.connect_write() as connection:
        policy_version = str(
            connection.execute(
                "SELECT policy_version FROM operational_time_policies ORDER BY effective_from LIMIT 1"
            ).fetchone()[0]
        )
        for job_id, job_type in (
            (f"job-pulse-{suffix}", "ONLINE_PULSE"),
            (f"job-listing-{suffix}", "LISTING_STATUS_SCAN"),
        ):
            connection.execute(
                """
                INSERT OR IGNORE INTO automation_jobs(
                    job_id, job_type, display_name, enabled, schedule_kind,
                    schedule_expression, priority, config_json, created_at, updated_at
                ) VALUES(?, ?, ?, 1, 'INTERVAL_MINUTES', '10', 60, '{}', ?, ?)
                """,
                (
                    job_id,
                    job_type,
                    f"Synthetic {job_type}",
                    NOW.isoformat(),
                    NOW.isoformat(),
                ),
            )
        connection.execute(
            """
            INSERT INTO automation_runs(
                run_id, job_id, job_type, logical_run_key, run_status,
                platform_name, platform_trade_date, seller_operation_date,
                seller_phase, time_policy_version, scheduled_for, started_at,
                finished_at, created_at, updated_at
            ) VALUES(?, ?, 'ONLINE_PULSE', ?, ?, 'synthetic-platform',
                     '2026-08-02', '2026-08-02', 'NORMAL_SALES', ?, ?, ?, ?, ?, ?)
            """,
            (
                pulse_run_id,
                f"job-pulse-{suffix}",
                f"logical-pulse-{suffix}",
                "MERGED" if merged_coverage else "SUCCESS",
                policy_version,
                scheduled_for.isoformat(),
                scan_started_at.isoformat(),
                scan_completed_at.isoformat(),
                scheduled_for.isoformat(),
                scan_completed_at.isoformat(),
            ),
        )
        if merged_coverage:
            connection.execute(
                """
                INSERT INTO automation_runs(
                    run_id, job_id, job_type, logical_run_key, run_status,
                    platform_name, platform_trade_date, seller_operation_date,
                    seller_phase, time_policy_version, scheduled_for, started_at,
                    finished_at, created_at, updated_at
                ) VALUES(?, ?, 'LISTING_STATUS_SCAN', ?, 'SUCCESS',
                         'synthetic-platform', '2026-08-02', '2026-08-02',
                         'NORMAL_SALES', ?, ?, ?, ?, ?, ?)
                """,
                (
                    observation_run_id,
                    f"job-listing-{suffix}",
                    f"logical-listing-{suffix}",
                    policy_version,
                    scheduled_for.isoformat(),
                    scan_started_at.isoformat(),
                    scan_completed_at.isoformat(),
                    scheduled_for.isoformat(),
                    scan_completed_at.isoformat(),
                ),
            )
            connection.execute(
                """
                INSERT INTO automation_run_links(
                    parent_run_id, child_run_id, relation_type, created_at
                ) VALUES(?, ?, 'MERGED_RUN', ?)
                """,
                (
                    observation_run_id,
                    pulse_run_id,
                    scan_completed_at.isoformat(),
                ),
            )
        connection.execute(
            """
            INSERT INTO product_observation_batches(
                observation_batch_id, automation_run_id, platform_name,
                scan_type, batch_status, scan_started_at, scan_completed_at,
                requested_scope_json, scope_complete, end_marker_verified,
                content_sha256, time_policy_version, created_at
            ) VALUES(?, ?, 'synthetic-platform', 'LISTING_STATUS_SCAN',
                     'ACCEPTED', ?, ?, '{}', ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                observation_run_id,
                scan_started_at.isoformat(),
                scan_completed_at.isoformat(),
                int(scope_complete),
                int(end_marker_verified),
                f"sha256:{suffix}",
                policy_version,
                scan_completed_at.isoformat(),
            ),
        )
        connection.execute(
            """
            INSERT INTO product_observation_items(
                observation_item_id, observation_batch_id, internal_sku,
                platform_product_name, grade, observed_price, observed_inventory,
                observed_online, observed_at, platform_trade_date,
                seller_operation_date, seller_phase, page_identity_key,
                mapping_status, mapping_version, evidence_sha256
            ) VALUES(?, ?, 'SKU-SYNTHETIC-001', 'Synthetic', 'B', ?, 10, ?, ?,
                     '2026-08-02', '2026-08-02', 'NORMAL_SALES', ?, ?,
                     'mapping-v1', ?)
            """,
            (
                item_id,
                batch_id,
                observed_price,
                int(observed_online),
                scan_completed_at.isoformat(),
                f"page-{suffix}",
                mapping_status,
                f"sha256:evidence-{suffix}",
            ),
        )
    return pulse_run_id, item_id


def test_incident_review_token_and_outbox_commit_atomically(runtime_repository):
    incident = create_s4_incident(runtime_repository)
    service = IncidentReviewService(runtime_repository)

    result = service.create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )

    assert result.replayed is False
    assert result.incident.incident_status is IncidentStatus.WAITING_HUMAN
    assert result.review_task.source_task_id is None
    assert result.review_task.review_type == "emergency_protection"
    assert result.review_token.allowed_actions == ["adjusted", "approved", "rejected"]
    assert result.notification.related_task_id is None
    assert (
        result.notification.related_review_task_id == result.review_task.review_task_id
    )
    assert "token=" in result.mobile_review_url

    stored_token = runtime_repository.get_review_token(result.review_token.token_id)
    stored_outbox = runtime_repository.get_notification_outbox(
        result.notification.notification_id
    )
    assert stored_token is not None
    assert stored_token.token_hash != result.raw_token
    assert stored_outbox is not None
    assert "mobile_review_url" not in stored_outbox.payload
    assert result.raw_token not in str(stored_outbox.payload)
    assert stored_outbox.payload["title"] == "极端低价，请立即处理"
    assert "极端低价，请立即处理" in stored_outbox.payload["message"]
    assert "emergency_protection" not in stored_outbox.payload["message"]

    events = IncidentManagementService(runtime_repository).list_events(
        incident.incident_id
    )
    assert events[-1].event_type is IncidentEventType.REVIEW_RECORDED
    assert events[-1].source_ref_id == result.review_task.review_task_id


def test_incident_review_creation_replays_same_bundle(runtime_repository):
    incident = create_s4_incident(runtime_repository)
    service = IncidentReviewService(runtime_repository)
    first = service.create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )

    replay = service.create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=3),
        created_at=NOW + timedelta(minutes=2),
    )

    assert replay.replayed is True
    assert replay.review_task.review_task_id == first.review_task.review_task_id
    assert replay.review_token.token_id == first.review_token.token_id
    assert replay.notification.notification_id == first.notification.notification_id
    assert replay.raw_token == first.raw_token
    assert (
        len(
            runtime_repository.list_review_tokens_by_review_task_id(
                first.review_task.review_task_id
            )
        )
        == 1
    )
    assert (
        len(
            runtime_repository.list_notification_outbox(
                related_review_task_id=first.review_task.review_task_id
            )
        )
        == 1
    )


def test_feishu_delivery_reuses_precreated_incident_token_without_persisting_url(
    runtime_repository,
    monkeypatch,
):
    monkeypatch.setenv("DEFAULT_NOTIFICATION_CHANNEL", "feishu")
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    outbox_service = NotificationOutboxService(
        runtime_repository,
        clock=lambda: NOW + timedelta(minutes=2),
    )

    prepared, token_id = outbox_service._prepare_delivery_notification(
        review.notification,
        FeishuOutboxSender(),
    )

    assert token_id == review.review_token.token_id
    assert prepared.payload["mobile_review_url"] == review.mobile_review_url
    stored = runtime_repository.get_notification_outbox(
        review.notification.notification_id
    )
    assert stored is not None
    assert "mobile_review_url" not in stored.payload
    assert (
        len(
            runtime_repository.list_review_tokens_by_review_task_id(
                review.review_task.review_task_id
            )
        )
        == 1
    )


def test_mobile_review_page_shows_three_business_actions_and_price_input(
    runtime_repository,
):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    model = OperationsQueryService(
        runtime_repository,
        OperationsWebPaths(
            runtime_db=runtime_repository.db_path,
            products_workbook=runtime_repository.db_path.parent / "products.xlsx",
            price_rules_workbook=runtime_repository.db_path.parent / "price_rules.xlsx",
            listing_rules_workbook=runtime_repository.db_path.parent / "listing_rules.xlsx",
            queue_root=runtime_repository.db_path.parent / "queue",
        ),
        now_provider=lambda: NOW + timedelta(minutes=2),
    ).mobile_review(review.review_task.review_task_id, review.raw_token)

    html = render_mobile_review(model)

    assert "改价到" in html
    assert "立即下架" in html
    assert "我来处理" in html
    assert 'name="target_price"' in html
    assert "cancelled" not in html
    assert "极端低价处理" in html
    assert "emergency_protection" not in html


def test_initial_sent_at_starts_window_and_midpoint_is_queued_only_once(
    runtime_repository,
):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    assert review.notification.payload["title"] == "极端低价，请立即处理"
    assert "极端低价，请立即处理" in review.notification.payload["message"]
    notifications = IncidentNotificationService(runtime_repository)

    waiting = notifications.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )
    assert waiting.escalation_state == "WAITING_INITIAL_DELIVERY"
    assert waiting.decision_window_started_at is None
    assert waiting.automatic_inference_allowed is False

    sent_at = NOW + timedelta(minutes=2)
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=sent_at,
    )
    started = notifications.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )
    assert started.escalation_state == "INITIAL_SENT"
    assert started.decision_window_started_at == sent_at
    assert started.next_notification_at == sent_at + timedelta(minutes=5)
    assert started.automatic_inference_allowed is True

    early = notifications.enqueue_midpoint_if_due(
        incident.incident_id,
        review.review_task.review_task_id,
        now=sent_at + timedelta(minutes=4),
    )
    assert early.status == "NOT_DUE"
    due = notifications.enqueue_midpoint_if_due(
        incident.incident_id,
        review.review_task.review_task_id,
        now=sent_at + timedelta(minutes=5),
    )
    replay = notifications.enqueue_midpoint_if_due(
        incident.incident_id,
        review.review_task.review_task_id,
        now=sent_at + timedelta(minutes=6),
    )
    assert due.status == "MIDPOINT_QUEUED"
    assert due.notification is not None
    assert due.notification.payload["message"].startswith("极端低价仍未处理")
    assert replay.status == "MIDPOINT_ALREADY_QUEUED"
    assert replay.notification is not None
    assert replay.notification.notification_id == due.notification.notification_id
    assert (
        len(
            runtime_repository.list_notification_outbox(
                related_review_task_id=review.review_task.review_task_id
            )
        )
        == 2
    )

    midpoint_sent_at = sent_at + timedelta(minutes=7)
    mark_notification(
        runtime_repository,
        due.notification.notification_id,
        now=midpoint_sent_at,
    )
    completed = notifications.sync_midpoint_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
        due.notification.notification_id,
    )
    assert completed.escalation_state == "MIDPOINT_SENT"
    assert completed.notification_count == 2
    assert completed.decision_window_started_at == sent_at


def test_ack_suppresses_s4_midpoint_reminder(runtime_repository):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    sent_at = NOW + timedelta(minutes=2)
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=sent_at,
    )
    notifications = IncidentNotificationService(runtime_repository)
    notifications.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )
    IncidentManagementService(runtime_repository).acknowledge(
        incident.incident_id,
        event_key="ack-midpoint-suppression",
        occurred_at=sent_at + timedelta(minutes=1),
        actor="farm-owner",
    )

    result = notifications.enqueue_midpoint_if_due(
        incident.incident_id,
        review.review_task.review_task_id,
        now=sent_at + timedelta(minutes=5),
    )

    assert result.status == "MIDPOINT_ACK_SUPPRESSED"
    assert result.notification is None
    assert (
        len(
            runtime_repository.list_notification_outbox(
                related_review_task_id=review.review_task.review_task_id
            )
        )
        == 1
    )


def test_review_result_suppresses_s4_midpoint_reminder(runtime_repository):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    sent_at = NOW + timedelta(minutes=2)
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=sent_at,
    )
    notifications = IncidentNotificationService(runtime_repository)
    notifications.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )
    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )
    runtime_repository.resolve_mobile_review_atomic(
        review_task_id=review.review_task.review_task_id,
        token_hash=token_hash,
        status=ReviewTaskStatus.REJECTED,
        actor_source="mobile_review_token",
        now=sent_at + timedelta(minutes=1),
    )

    result = notifications.enqueue_midpoint_if_due(
        incident.incident_id,
        review.review_task.review_task_id,
        now=sent_at + timedelta(minutes=5),
    )

    assert result.status == "MIDPOINT_REVIEW_RESOLVED"
    assert result.notification is None


def test_s3_initial_delivery_has_no_midpoint_schedule(runtime_repository):
    incident_detection = replace(
        detection("detect-s3", occurred_at=NOW),
        severity="S3",
        subject_type="internal_sku",
        subject_key="SKU-SYNTHETIC-001",
        title="Synthetic low price",
    )
    incident = (
        IncidentManagementService(runtime_repository)
        .detect(incident_detection)
        .incident
    )
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    assert review.notification.payload["title"] == "价格异常，请立即处理"
    assert "价格异常，请立即处理" in review.notification.payload["message"]
    sent_at = NOW + timedelta(minutes=2)
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=sent_at,
    )
    notifications = IncidentNotificationService(runtime_repository)

    timing = notifications.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )
    reminder = notifications.enqueue_midpoint_if_due(
        incident.incident_id,
        review.review_task.review_task_id,
        now=sent_at + timedelta(minutes=30),
    )

    assert timing.next_notification_at is None
    assert timing.automatic_inference_allowed is False
    assert reminder.status == "NOT_DUE"


def test_initial_delivery_failure_never_starts_decision_window(runtime_repository):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=NOW + timedelta(minutes=2),
        classification="PERM_FAILED",
    )

    timing = IncidentNotificationService(runtime_repository).sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )

    assert timing.escalation_state == "INITIAL_DELIVERY_BLOCKED"
    assert timing.decision_window_started_at is None
    assert timing.automatic_inference_allowed is False


def test_recovery_notification_uses_stable_outbox_identity(runtime_repository):
    incident = create_s4_incident(runtime_repository)
    notifications = IncidentNotificationService(runtime_repository)

    first = notifications.enqueue_status_notification(
        incident.incident_id,
        notification_kind="worker_recovered",
        source_event_key="recovery-event-001",
        message="Worker 已恢复",
    )
    replay = notifications.enqueue_status_notification(
        incident.incident_id,
        notification_kind="worker_recovered",
        source_event_key="recovery-event-001",
        message="Worker 已恢复",
    )

    assert replay.notification_id == first.notification_id
    assert replay.notification_key == first.notification_key
    assert len(runtime_repository.list_notification_outbox()) == 1

    with pytest.raises(NotificationIdempotencyConflictError):
        notifications.enqueue_status_notification(
            incident.incident_id,
            notification_kind="worker_recovered",
            source_event_key="recovery-event-001",
            message="different immutable content",
        )


@pytest.mark.parametrize("merged_coverage", [False, True])
def test_online_pulse_eligibility_requires_delivery_then_complete_imported_fact(
    runtime_repository,
    merged_coverage,
):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    _, initial_observation_id = insert_pulse_observation_facts(
        runtime_repository,
        suffix=f"initial-{int(merged_coverage)}",
        scheduled_for=NOW,
        scan_started_at=NOW,
        scan_completed_at=NOW + timedelta(seconds=30),
        merged_coverage=merged_coverage,
    )
    service = IncidentNotificationService(runtime_repository)

    before_delivery = service.evaluate_online_pulse_eligibility(
        incident.incident_id,
        review.review_task.review_task_id,
        initial_observation_id=initial_observation_id,
    )
    assert before_delivery.eligible is False
    assert before_delivery.reason == "INITIAL_DELIVERY_NOT_CONFIRMED"

    sent_at = NOW + timedelta(minutes=2)
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=sent_at,
    )
    service.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )
    insert_pulse_observation_facts(
        runtime_repository,
        suffix=f"started-before-delivery-{int(merged_coverage)}",
        scheduled_for=NOW + timedelta(minutes=10),
        scan_started_at=sent_at - timedelta(seconds=1),
        scan_completed_at=NOW + timedelta(minutes=11),
        merged_coverage=merged_coverage,
    )

    waiting = service.evaluate_online_pulse_eligibility(
        incident.incident_id,
        review.review_task.review_task_id,
        initial_observation_id=initial_observation_id,
    )
    assert waiting.eligible is False
    assert waiting.reason == "WAITING_FOR_QUALIFIED_ONLINE_PULSE"

    pulse_run_id, qualified_observation_id = insert_pulse_observation_facts(
        runtime_repository,
        suffix=f"qualified-{int(merged_coverage)}",
        scheduled_for=NOW + timedelta(minutes=20),
        scan_started_at=NOW + timedelta(minutes=20),
        scan_completed_at=NOW + timedelta(minutes=21),
        merged_coverage=merged_coverage,
    )
    eligible = service.evaluate_online_pulse_eligibility(
        incident.incident_id,
        review.review_task.review_task_id,
        initial_observation_id=initial_observation_id,
    )

    assert eligible.eligible is True
    assert eligible.reason == "QUALIFIED_ONLINE_PULSE_IMPORTED"
    assert eligible.pulse_run_id == pulse_run_id
    assert eligible.observation_item_id == qualified_observation_id
    assert eligible.pulse_scheduled_for == NOW + timedelta(minutes=20)
    assert eligible.automatic_eligibility_reached_at == NOW + timedelta(minutes=21)
    assert eligible.observed_price == Decimal("6.00")


def test_online_pulse_incomplete_unmapped_or_offline_fact_only_delays(
    runtime_repository,
):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    _, initial_observation_id = insert_pulse_observation_facts(
        runtime_repository,
        suffix="initial-invalid-matrix",
        scheduled_for=NOW,
        scan_started_at=NOW,
        scan_completed_at=NOW + timedelta(seconds=30),
    )
    sent_at = NOW + timedelta(minutes=2)
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=sent_at,
    )
    service = IncidentNotificationService(runtime_repository)
    service.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )
    invalid_cases = (
        {"scope_complete": False},
        {"end_marker_verified": False},
        {"mapping_status": "UNMAPPED"},
        {"observed_online": False},
    )
    for index, values in enumerate(invalid_cases, start=1):
        insert_pulse_observation_facts(
            runtime_repository,
            suffix=f"invalid-{index}",
            scheduled_for=NOW + timedelta(minutes=10 * index),
            scan_started_at=NOW + timedelta(minutes=10 * index),
            scan_completed_at=NOW + timedelta(minutes=10 * index + 1),
            **values,
        )

    result = service.evaluate_online_pulse_eligibility(
        incident.incident_id,
        review.review_task.review_task_id,
        initial_observation_id=initial_observation_id,
    )

    assert result.eligible is False
    assert result.reason == "WAITING_FOR_QUALIFIED_ONLINE_PULSE"


def test_online_pulse_eligibility_rejects_wrong_initial_binding_and_review_result(
    runtime_repository,
):
    incident = create_s4_incident(runtime_repository)
    review = IncidentReviewService(runtime_repository).create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    _, initial_observation_id = insert_pulse_observation_facts(
        runtime_repository,
        suffix="initial-review-result",
        scheduled_for=NOW,
        scan_started_at=NOW,
        scan_completed_at=NOW + timedelta(seconds=30),
    )
    sent_at = NOW + timedelta(minutes=2)
    mark_notification(
        runtime_repository,
        review.notification.notification_id,
        now=sent_at,
    )
    service = IncidentNotificationService(runtime_repository)
    service.sync_initial_delivery(
        incident.incident_id,
        review.review_task.review_task_id,
    )

    wrong_binding = service.evaluate_online_pulse_eligibility(
        incident.incident_id,
        review.review_task.review_task_id,
        initial_observation_id="missing-observation",
    )
    assert wrong_binding.reason == "INITIAL_OBSERVATION_NOT_BOUND"

    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )
    runtime_repository.resolve_mobile_review_atomic(
        review_task_id=review.review_task.review_task_id,
        token_hash=token_hash,
        status=ReviewTaskStatus.REJECTED,
        actor_source="mobile_review_token",
        now=sent_at + timedelta(minutes=1),
    )
    resolved = service.evaluate_online_pulse_eligibility(
        incident.incident_id,
        review.review_task.review_task_id,
        initial_observation_id=initial_observation_id,
    )
    assert resolved.eligible is False
    assert resolved.reason == "REVIEW_RESOLVED"


@pytest.mark.parametrize(
    "failure_point",
    [
        "after_review_insert",
        "after_review_token_insert",
        "after_outbox_insert",
        "after_incident_review_event_insert",
    ],
)
def test_incident_review_bundle_rolls_back_on_database_failure(
    runtime_repository,
    failure_point,
):
    incident = create_s4_incident(runtime_repository)
    service = IncidentReviewService(runtime_repository)

    def fail(point: str) -> None:
        if point == failure_point:
            raise RuntimeError("synthetic review bundle failure")

    with pytest.raises(RuntimeError, match="synthetic review bundle failure"):
        service.create_initial_review(
            incident.incident_id,
            required_by=NOW + timedelta(hours=2),
            created_at=NOW + timedelta(minutes=1),
            failure_injector=fail,
        )

    current = IncidentManagementService(runtime_repository).get(incident.incident_id)
    assert current is not None
    assert current.incident_status is IncidentStatus.OPEN
    with runtime_repository.connect_read() as connection:
        assert (
            connection.execute("SELECT COUNT(*) FROM review_tasks").fetchone()[0] == 0
        )
        assert (
            connection.execute("SELECT COUNT(*) FROM review_tokens").fetchone()[0] == 0
        )
        assert (
            connection.execute("SELECT COUNT(*) FROM notification_outbox").fetchone()[0]
            == 0
        )
        assert (
            connection.execute(
                "SELECT COUNT(*) FROM operational_incident_events WHERE event_type = 'REVIEW_RECORDED'"
            ).fetchone()[0]
            == 0
        )


@pytest.mark.parametrize(
    ("status", "payload", "expected_action"),
    [
        (
            ReviewTaskStatus.ADJUSTED,
            {"adjustment": {"target_price": "12.50"}},
            TaskActionType.UPDATE_PRICE,
        ),
        (ReviewTaskStatus.APPROVED, {}, TaskActionType.SET_OFFLINE),
    ],
)
def test_incident_mobile_review_creates_manual_v4_or_v5_task_atomically(
    runtime_repository,
    status,
    payload,
    expected_action,
):
    incident, review = create_review_with_listing(runtime_repository)
    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )

    result = runtime_repository.resolve_mobile_review_atomic(
        review_task_id=review.review_task.review_task_id,
        token_hash=token_hash,
        status=status,
        actor_source="mobile_review_token",
        resolution_payload=payload,
        emergency_base_cost=Decimal("10.00"),
        emergency_base_cost_source_ref="products.xlsx:sha256:synthetic",
        emergency_product_snapshot_verifier=stable_product_snapshot,
        now=NOW + timedelta(minutes=2),
    )

    assert result.review_task.review_status is status
    assert result.review_token.used_at == NOW + timedelta(minutes=2)
    assert result.source_task is not None
    assert result.source_task.action_type is expected_action
    assert result.source_task.task_status is TaskStatus.PENDING
    assert result.source_task.origin_type is TaskOriginType.MANUAL
    assert result.source_task.origin_ref_id == (
        f"incident-review:{review.review_task.review_task_id}"
    )
    assert result.source_task.expected_old_price == Decimal("8.00")
    if status is ReviewTaskStatus.ADJUSTED:
        assert result.source_task.target_price == Decimal("12.50")
    else:
        assert result.source_task.target_status == "offline"
    events = IncidentManagementService(runtime_repository).list_events(
        incident.incident_id
    )
    assert {event.event_type for event in events[-2:]} == {
        IncidentEventType.REVIEW_RECORDED,
        IncidentEventType.TASK_RECORDED,
    }


def test_authenticated_web_review_reuses_incident_atomic_path(
    runtime_repository,
    tmp_path,
):
    _, review = create_review_with_listing(runtime_repository)
    products_path = tmp_path / "products.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(PRODUCT_HEADERS)
    sheet.append(
        [
            "SKU-SYNTHETIC-001",
            "Synthetic flower",
            "B",
            "60",
            "bundle",
            10,
            50,
            True,
            8,
            12,
            "",
            "synthetic",
            "green",
        ]
    )
    workbook.save(products_path)
    service = ReviewResolutionApplicationService(
        runtime_repository,
        PrincipalCapabilityBackend(),
        products_path=products_path,
    )

    result = service.resolve(
        Principal("operator", frozenset({Capability.HANDLE_REVIEW})),
        review_task_id=review.review_task.review_task_id,
        action=ReviewTaskStatus.ADJUSTED.value,
        target_price="12.50",
        note="raise to safe price",
    )

    assert result.review_status == ReviewTaskStatus.ADJUSTED.value
    assert result.created_task_id
    stored_review = runtime_repository.get_review_task(result.review_task_id)
    assert stored_review is not None
    assert stored_review.resolved_by == "operator"
    created_task = runtime_repository.get_task(result.created_task_id)
    assert created_task is not None
    assert created_task.action_type is TaskActionType.UPDATE_PRICE
    assert created_task.target_price == Decimal("12.50")
    stored_token = runtime_repository.get_review_token(review.review_token.token_id)
    assert stored_token is not None
    assert stored_token.used_at is None


def test_authenticated_incident_review_rolls_back_every_table_on_failure(
    runtime_repository,
):
    incident, review = create_review_with_listing(runtime_repository)
    events_before = IncidentManagementService(runtime_repository).list_events(
        incident.incident_id
    )

    def fail_after_task(point: str) -> None:
        if point == "after_incident_task_insert":
            raise RuntimeError("synthetic failure")

    with pytest.raises(RuntimeError, match="synthetic failure"):
        runtime_repository.resolve_authenticated_incident_review_atomic(
            review_task_id=review.review_task.review_task_id,
            status=ReviewTaskStatus.ADJUSTED,
            actor_source="authenticated_web",
            actor="operator",
            resolution_payload={"adjustment": {"target_price": "12.50"}},
            emergency_base_cost=Decimal("10.00"),
            emergency_base_cost_source_ref="products.xlsx:sha256:synthetic",
            emergency_product_snapshot_verifier=stable_product_snapshot,
            now=NOW + timedelta(minutes=2),
            failure_injector=fail_after_task,
        )

    stored_review = runtime_repository.get_review_task(
        review.review_task.review_task_id
    )
    assert stored_review is not None
    assert stored_review.review_status is ReviewTaskStatus.PENDING
    assert runtime_repository.list_tasks() == []
    events = IncidentManagementService(runtime_repository).list_events(
        incident.incident_id
    )
    assert events == events_before


def test_incident_mobile_review_rejects_price_below_base_cost_without_consuming_token(
    runtime_repository,
):
    _, review = create_review_with_listing(runtime_repository)
    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )

    with pytest.raises(
        MobileReviewTransactionError,
        match="目标价格不得低于商品基础成本",
    ):
        runtime_repository.resolve_mobile_review_atomic(
            review_task_id=review.review_task.review_task_id,
            token_hash=token_hash,
            status=ReviewTaskStatus.ADJUSTED,
            actor_source="mobile_review_token",
            resolution_payload={"adjustment": {"target_price": "9.99"}},
            emergency_base_cost=Decimal("10.00"),
            emergency_base_cost_source_ref="products.xlsx:sha256:synthetic",
            emergency_product_snapshot_verifier=stable_product_snapshot,
            now=NOW + timedelta(minutes=2),
        )

    stored_review = runtime_repository.get_review_task(
        review.review_task.review_task_id
    )
    stored_token = runtime_repository.get_review_token(review.review_token.token_id)
    assert stored_review is not None
    assert stored_review.review_status is ReviewTaskStatus.PENDING
    assert stored_token is not None
    assert stored_token.used_at is None
    assert runtime_repository.list_tasks() == []


def test_incident_mobile_review_human_handling_creates_no_platform_task(
    runtime_repository,
):
    incident, review = create_review_with_listing(runtime_repository)
    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )

    result = runtime_repository.resolve_mobile_review_atomic(
        review_task_id=review.review_task.review_task_id,
        token_hash=token_hash,
        status=ReviewTaskStatus.REJECTED,
        actor_source="mobile_review_token",
        note="I will handle it",
        now=NOW + timedelta(minutes=2),
    )

    assert result.review_task.review_status is ReviewTaskStatus.REJECTED
    assert result.source_task is None
    assert runtime_repository.list_tasks() == []
    current = IncidentManagementService(runtime_repository).get(incident.incident_id)
    assert current is not None
    assert current.incident_status is IncidentStatus.WAITING_HUMAN


def test_formal_mobile_review_preempts_auto_protecting_before_side_effect(
    runtime_repository,
):
    incident, review = create_review_with_listing(runtime_repository)
    automatic_task = Task(
        task_id="TASK-EMERGENCY-SYNTHETIC",
        internal_sku=review.review_task.internal_sku,
        platform_name=review.review_task.platform_name,
        action_type=TaskActionType.SET_OFFLINE,
        priority=1,
        task_status=TaskStatus.PENDING,
        created_at=NOW + timedelta(minutes=1),
        target_status="offline",
        decision_trace={
            "incident_id": incident.incident_id,
            "review_task_id": review.review_task.review_task_id,
        },
        origin_type=TaskOriginType.SYSTEM_EMERGENCY,
        origin_ref_id="emergency:synthetic",
        expires_at=NOW + timedelta(hours=1),
    )
    with runtime_repository.connect_write() as connection, connection:
        SQLiteRuntimeRepository._insert_tasks_on_connection(
            connection,
            [automatic_task],
        )
        connection.execute(
            "UPDATE operational_incidents SET incident_status = 'AUTO_PROTECTING' "
            "WHERE incident_id = ?",
            (incident.incident_id,),
        )
    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )

    result = runtime_repository.resolve_mobile_review_atomic(
        review_task_id=review.review_task.review_task_id,
        token_hash=token_hash,
        status=ReviewTaskStatus.APPROVED,
        actor_source="mobile_review_token",
        resolution_payload={},
        emergency_base_cost=Decimal("10.00"),
        emergency_base_cost_source_ref="products.xlsx:sha256:synthetic",
        emergency_product_snapshot_verifier=stable_product_snapshot,
        now=NOW + timedelta(minutes=2),
    )

    current = IncidentManagementService(runtime_repository).get(incident.incident_id)
    cancelled = runtime_repository.get_task(automatic_task.task_id)
    assert current is not None
    assert current.incident_status is IncidentStatus.WAITING_HUMAN
    assert cancelled is not None
    assert cancelled.task_status is TaskStatus.CANCELLED
    assert result.source_task is not None
    assert result.source_task.priority == 0


def test_incident_mobile_review_task_and_token_roll_back_together(runtime_repository):
    _, review = create_review_with_listing(runtime_repository)
    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )

    def fail(point: str) -> None:
        if point == "after_incident_task_insert":
            raise RuntimeError("synthetic decision failure")

    with pytest.raises(RuntimeError, match="synthetic decision failure"):
        runtime_repository.resolve_mobile_review_atomic(
            review_task_id=review.review_task.review_task_id,
            token_hash=token_hash,
            status=ReviewTaskStatus.APPROVED,
            actor_source="mobile_review_token",
            emergency_base_cost=Decimal("10.00"),
            emergency_base_cost_source_ref="products.xlsx:sha256:synthetic",
            emergency_product_snapshot_verifier=stable_product_snapshot,
            now=NOW + timedelta(minutes=2),
            failure_injector=fail,
        )

    stored_review = runtime_repository.get_review_task(
        review.review_task.review_task_id
    )
    stored_token = runtime_repository.get_review_token(review.review_token.token_id)
    assert stored_review is not None
    assert stored_review.review_status is ReviewTaskStatus.PENDING
    assert stored_token is not None
    assert stored_token.used_at is None
    assert runtime_repository.list_tasks() == []


def test_incident_review_reissues_expired_human_entry_and_restarts_window(
    runtime_repository,
):
    incident = create_s4_incident(runtime_repository)
    service = IncidentReviewService(runtime_repository)
    original = service.create_initial_review(
        incident.incident_id,
        required_by=NOW + timedelta(hours=2),
        created_at=NOW + timedelta(minutes=1),
    )
    mark_notification(
        runtime_repository,
        original.notification.notification_id,
        now=NOW + timedelta(minutes=2),
    )
    delivered = IncidentNotificationService(runtime_repository).sync_initial_delivery(
        incident.incident_id,
        original.review_task.review_task_id,
    )
    assert delivered.decision_window_started_at == NOW + timedelta(minutes=2)
    reissue_at = NOW + timedelta(days=2)
    with runtime_repository.connect_write() as connection, connection:
        connection.execute(
            "UPDATE review_tokens SET expires_at = ? WHERE token_id = ?",
            (
                (reissue_at - timedelta(seconds=1)).isoformat(),
                original.review_token.token_id,
            ),
        )

    reissued = service.ensure_usable_human_entry(
        original.review_task.review_task_id,
        now=reissue_at,
    )

    assert reissued is not None
    assert reissued.raw_token != original.raw_token
    assert reissued.review_task.required_by == reissue_at + timedelta(
        hours=1, minutes=59
    )
    tokens = runtime_repository.list_review_tokens_by_review_task_id(
        original.review_task.review_task_id
    )
    assert len(tokens) == 2
    old_token = next(
        token for token in tokens if token.token_id == original.review_token.token_id
    )
    new_token = next(
        token for token in tokens if token.token_id == reissued.review_token.token_id
    )
    assert old_token.revoked_at == reissue_at
    assert new_token.expires_at == reissue_at + timedelta(hours=24)
    outboxes = runtime_repository.list_notification_outbox(
        related_review_task_id=original.review_task.review_task_id
    )
    assert len(outboxes) == 2
    assert outboxes[-1].notification_id == reissued.notification.notification_id
    timing = IncidentNotificationService(runtime_repository).sync_initial_delivery(
        incident.incident_id,
        original.review_task.review_task_id,
    )
    assert timing.escalation_state == "WAITING_INITIAL_DELIVERY"
    assert timing.decision_window_started_at is None


def test_manual_review_cost_snapshot_fails_closed_after_database_lock_wait(
    runtime_repository,
    tmp_path,
):
    _, review = create_review_with_listing(runtime_repository)
    products_path = tmp_path / "products-lock-wait.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(PRODUCT_HEADERS)
    sheet.append(
        [
            "SKU-SYNTHETIC-001",
            "Synthetic flower",
            "B",
            "60",
            "bundle",
            10,
            50,
            True,
            8,
            12,
            "",
            "synthetic",
            "green",
        ]
    )
    workbook.save(products_path)
    snapshot = _read_authoritative_product_cost_snapshot(
        products_path,
        internal_sku="SKU-SYNTHETIC-001",
    )
    token_hash = ReviewTokenService(runtime_repository)._hash_raw_token(
        review.raw_token
    )
    second_repository = SQLiteRuntimeRepository(runtime_repository.db_path)
    started = Event()
    errors: list[BaseException] = []
    blocker = runtime_repository.connect_write()
    blocker.execute("BEGIN IMMEDIATE")

    def submit_review() -> None:
        started.set()
        try:
            second_repository.resolve_mobile_review_atomic(
                review_task_id=review.review_task.review_task_id,
                token_hash=token_hash,
                status=ReviewTaskStatus.APPROVED,
                actor_source="mobile_review_token",
                emergency_base_cost=snapshot[0],
                emergency_base_cost_source_ref=snapshot[1],
                emergency_product_snapshot_verifier=lambda: (
                    _read_authoritative_product_cost_snapshot(
                        products_path,
                        internal_sku="SKU-SYNTHETIC-001",
                    )
                ),
                now=NOW + timedelta(minutes=2),
            )
        except BaseException as exc:  # noqa: BLE001 - thread assertion capture
            errors.append(exc)

    thread = Thread(target=submit_review)
    thread.start()
    assert started.wait(timeout=2)
    sheet.cell(row=2, column=6, value=11)
    workbook.save(products_path)
    blocker.commit()
    blocker.close()
    thread.join(timeout=5)

    assert not thread.is_alive()
    assert len(errors) == 1
    assert isinstance(errors[0], MobileReviewTransactionError)
    assert "商品主数据在复核提交前发生变化" in str(errors[0])
    stored_review = runtime_repository.get_review_task(
        review.review_task.review_task_id
    )
    assert stored_review is not None
    assert stored_review.review_status is ReviewTaskStatus.PENDING
    assert runtime_repository.list_tasks() == []


def test_workflow_rereads_authoritative_product_base_cost_for_incident_review(
    runtime_repository,
    tmp_path,
):
    _, review = create_review_with_listing(runtime_repository)
    products_path = tmp_path / "products.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(PRODUCT_HEADERS)
    sheet.append(
        [
            "SKU-SYNTHETIC-001",
            "Synthetic flower",
            "B",
            "60",
            "bundle",
            10,
            50,
            True,
            8,
            12,
            "",
            "synthetic",
            "green",
        ]
    )
    workbook.save(products_path)

    result = resolve_mobile_review(
        runtime_repository.db_path,
        review.review_task.review_task_id,
        review.raw_token,
        ReviewTaskStatus.ADJUSTED.value,
        resolution_payload={"adjustment": {"target_price": "10.00"}},
        products_path=products_path,
        now=NOW + timedelta(minutes=2),
    )

    assert result.source_task is not None
    assert result.source_task.target_price == Decimal("10.00")
    assert result.source_task.decision_trace["base_cost"] == "10.00"
    assert str(result.source_task.decision_trace["base_cost_source_ref"]).startswith(
        "products.xlsx:sha256:"
    )


def test_workflow_human_handling_does_not_require_product_workbook(
    runtime_repository,
):
    _, review = create_review_with_listing(runtime_repository)

    result = resolve_mobile_review(
        runtime_repository.db_path,
        review.review_task.review_task_id,
        review.raw_token,
        ReviewTaskStatus.REJECTED.value,
        note="I will handle it",
        now=NOW + timedelta(minutes=2),
    )

    assert result.review_task.review_status is ReviewTaskStatus.REJECTED
    assert result.source_task is None
