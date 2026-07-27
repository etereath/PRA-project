from __future__ import annotations

import io
import json
import os
import re
import shutil
import tempfile
import unittest
from contextlib import closing, contextmanager
from dataclasses import replace
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from urllib.parse import unquote, urlencode, urlparse
from unittest.mock import patch
from uuid import uuid4

from app.enums import (
    ListingStrategy,
    NotificationSendStatus,
    ReviewTaskStatus,
    TaskActionType,
    TaskStatus,
)
from app.models import (
    ExecutionLog,
    ListingStatus,
    MockPlatformProductState,
    NotificationLog,
    ReviewTask,
    ScriptRun,
    ScriptRunItem,
    ShadowBotOperationLedger,
    Task,
)
from app.repositories.workbook_repository import (
    load_capacity_plans,
    load_cold_storage_statuses,
    load_listing_rules,
    load_price_rules,
    load_products,
    load_table_records,
    load_tasks,
    save_table_records,
)
from app.repositories.sqlite_runtime_repository import SQLiteRuntimeRepository
from app.repositories.mock_platform_repository import MockPlatformRepository
from app.services.runtime import ReviewTaskService, ReviewTokenService, RuntimeTaskService
from app.services.shadowbot_executor import FileDropShadowBotTaskRunner
from app.web import (
    DISPLAY_TIMEZONE,
    TABLE_OPTIONS,
    _RUNTIME_SESSIONS,
    _build_task_group_statuses,
    _parse_task_group_required_by,
    _redact_request_log_value,
    _resolve_table_path,
    application,
    default_dashboard_state,
    default_execution_state,
    default_manual_state,
    default_runtime_state,
    default_table_editor_state,
    render_dashboard_page,
    render_execution_page,
    render_execution_logs_page,
    render_manual_intervention_page,
    render_runtime_page,
    render_table_editor_page,
    render_tasks_page,
)


def test_task_group_deadline_is_interpreted_in_beijing_timezone() -> None:
    deadline = _parse_task_group_required_by("2099-07-27T12:30")

    assert deadline.tzinfo == DISPLAY_TIMEZONE
    assert deadline.utcoffset() == timedelta(hours=8)


def test_request_log_redacts_mobile_review_token() -> None:
    value = (
        "GET /mobile/review/REVIEW-1?token=secret-token&next=1 "
        "HTTP/1.1"
    )
    redacted = _redact_request_log_value(value)
    assert "secret-token" not in redacted
    assert "token=[REDACTED]&next=1" in redacted


@contextmanager
def _workspace_temp_dir(name: str):
    root = Path.cwd() / "test_runtime_tmp" / name
    root.mkdir(parents=True, exist_ok=True)
    path = root / f"case_{uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def _runtime_task(
    task_id: str,
    *,
    status: TaskStatus = TaskStatus.PENDING,
    trade_date: date = date(2026, 5, 4),
    action_type: TaskActionType = TaskActionType.MANUAL_PRICE_REVIEW,
    required_by: datetime | None = None,
) -> Task:
    return Task(
        task_id=task_id,
        internal_sku=None,
        platform_name=None,
        action_type=action_type,
        priority=2,
        task_status=status,
        created_at=datetime(2026, 5, 4, 9, 0),
        trade_date=trade_date,
        scope_type="global",
        scope_key=trade_date.isoformat(),
        dedupe_key=f"{trade_date.isoformat()}|global|{trade_date.isoformat()}|{action_type.value}|{task_id}",
        decision_trace={"reason": "manual review"},
        result_message="需要人工复核",
        required_by=required_by,
    )


class WebTests(unittest.TestCase):
    def setUp(self) -> None:
        _RUNTIME_SESSIONS.clear()
        self._allowed_dirs_patcher = patch.dict(
            "os.environ",
            {
                "PRA_ALLOWED_DATA_DIRS": os.pathsep.join(
                    (tempfile.gettempdir(), str(Path.cwd()))
                )
            },
            clear=False,
        )
        self._allowed_dirs_patcher.start()
        self.addCleanup(self._allowed_dirs_patcher.stop)

    def _existing_runtime_db(self, label: str) -> Path:
        path = Path(tempfile.gettempdir()) / f"pra_{label}_{uuid4().hex}.sqlite3"
        SQLiteRuntimeRepository(path).init_schema()
        self.addCleanup(lambda: path.unlink(missing_ok=True))
        return path

    def test_health_ignores_request_runtime_db_path(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            healthy_db = root / "trusted.sqlite3"
            attacker_db = root / "attacker.sqlite3"
            SQLiteRuntimeRepository(healthy_db).init_schema()

            with patch("app.web.DEFAULT_RUNTIME_DB", healthy_db):
                status, _, body = self._call_app(
                    path="/health",
                    query=urlencode({"runtime_db": str(attacker_db)}),
                )
            self.assertEqual(status, "200 OK")
            self.assertEqual(body, "ok")
            self.assertFalse(attacker_db.exists())

            with patch("app.web.DEFAULT_RUNTIME_DB", attacker_db):
                status, _, body = self._call_app(
                    path="/health",
                    query=urlencode({"runtime_db": str(healthy_db)}),
                )
            self.assertEqual(status, "503 Service Unavailable")
            self.assertIn("unhealthy", body)

    def _runtime_login(self, db_path: Path, *, username: str = "admin", password: str = "secret") -> str:
        _, login_headers, login_body = self._call_app(
            path="/runtime/login",
            query=urlencode({"runtime_db": str(db_path.resolve())}),
        )
        login_match = re.search(r'name="csrf_token" value="([^"]+)"', login_body)
        self.assertIsNotNone(login_match)
        preauth_cookie = login_headers["Set-Cookie"].split(";", 1)[0]
        status, headers, _ = self._call_app(
            path="/runtime/login",
            method="POST",
            cookie=preauth_cookie,
            body=urlencode(
                {
                    "runtime_db": str(db_path.resolve()),
                    "username": username,
                    "password": password,
                    "csrf_token": login_match.group(1),
                }
            ),
        )
        self.assertEqual(status, "303 See Other")
        return headers["Set-Cookie"].split(";", 1)[0]

    def _call_app(
        self,
        *,
        path: str,
        method: str = "GET",
        query: str = "",
        body: str = "",
        cookie: str = "",
        environ_overrides: dict[str, str] | None = None,
    ) -> tuple[str, dict[str, str], str]:
        captured: dict[str, object] = {}

        def start_response(status, headers):
            captured["status"] = status
            captured["headers"] = headers

        payload = body.encode("utf-8")
        environ = {
            "REQUEST_METHOD": method,
            "PATH_INFO": path,
            "QUERY_STRING": query,
            "CONTENT_LENGTH": str(len(payload)),
            "CONTENT_TYPE": "application/x-www-form-urlencoded",
            "wsgi.input": io.BytesIO(payload),
        }
        if cookie:
            environ["HTTP_COOKIE"] = cookie
        if environ_overrides:
            environ.update(environ_overrides)
        if (
            cookie
            and method.upper() in {"POST", "PUT", "PATCH", "DELETE"}
            and not path.startswith("/mobile/review/")
            and path != "/runtime/login"
            and "csrf_token=" not in body
        ):
            session_id = cookie.split("=", 1)[1].split(";", 1)[0]
            session = _RUNTIME_SESSIONS.get(session_id)
            csrf_token = str(session.get("csrf_token", "")) if session else ""
            if csrf_token:
                body = f"{body}&{urlencode({'csrf_token': csrf_token})}" if body else urlencode(
                    {"csrf_token": csrf_token}
                )
                payload = body.encode("utf-8")
                environ["CONTENT_LENGTH"] = str(len(payload))
                environ["wsgi.input"] = io.BytesIO(payload)
        response = application(environ, start_response)
        response_body = b"".join(response).decode("utf-8")
        headers: dict[str, str] = {}
        for name, value in captured["headers"]:
            headers.setdefault(name, value)
        return str(captured["status"]), headers, response_body

    def test_render_dashboard_page_contains_console_title(self) -> None:
        html = render_dashboard_page(
            params=default_dashboard_state(),
            message="ok",
            message_level="success",
            validation_summary=None,
            generation_summary=None,
            preview_ready=False,
        )
        self.assertIn("PRA 运行态运营后台", html)
        self.assertIn("任务生成", html)
        self.assertIn("库存策略", html)
        self.assertIn("业务数据", html)
        self.assertIn("value='single_rule' selected", html)
        self.assertNotIn('id="platform"', html)
        self.assertNotIn('id="task_group_id"', html)
        self.assertNotIn('id="products"', html)
        self.assertNotIn('value="validate"', html)
        self.assertNotIn("内置资源", html)
        deadline_match = re.search(r'id="required_by"[^>]*value="([^"]+)"', html)
        self.assertIsNotNone(deadline_match)
        deadline = datetime.fromisoformat(deadline_match.group(1))
        current_beijing = datetime.now(DISPLAY_TIMEZONE).replace(tzinfo=None)
        self.assertGreater(deadline, current_beijing + timedelta(minutes=28))
        self.assertLess(deadline, current_beijing + timedelta(minutes=31))

    def test_render_table_editor_contains_management_ui(self) -> None:
        html = render_table_editor_page(
            params=default_table_editor_state(),
            headers=["internal_sku", "product_name"],
            records=[{"internal_sku": "SKU-001", "product_name": "rose"}],
            message="ok",
            message_level="success",
            table_issues=[],
        )
        self.assertIn("Excel", html)
        self.assertIn("保存当前修改", html)
        self.assertIn("内部 SKU", html)
        self.assertIn("商品名称", html)
        self.assertIn("SKU-001", html)

    def test_new_predictive_tables_have_chinese_labels(self) -> None:
        self.assertIn("harvest_forecasts", TABLE_OPTIONS)
        html = render_table_editor_page(
            params={"table_name": "harvest_forecasts", "table_path": "data/samples/harvest_forecasts.xlsx"},
            headers=["forecast_id", "target_trade_date", "predicted_harvest_qty"],
            records=[],
            message="ok",
            message_level="success",
            table_issues=[],
        )
        self.assertIn("产量预测表", html)
        self.assertIn("目标交易日", html)
        self.assertIn("预测采收量", html)

    def test_switching_table_uses_new_default_path_when_old_default_was_posted(self) -> None:
        resolved = _resolve_table_path(
            table_name="listing_rules",
            previous_table_name="products",
            posted_path="D:/PRA project/data/samples/products.xlsx",
        )
        self.assertEqual(Path(resolved).parts[-3:], ("data", "samples", "listing_rules.xlsx"))

    def test_render_execution_page_contains_form(self) -> None:
        html = render_execution_page(
            params=default_execution_state(),
            message="ok",
            message_level="success",
            execution_summary=None,
        )
        self.assertIn("模拟执行", html)
        self.assertIn("mock_executor", html)
        self.assertIn("mock 执行/旧回写兼容", html)

    def test_manual_intervention_page_is_read_only_compatible(self) -> None:
        html = render_manual_intervention_page(
            params=default_manual_state(),
            message="ok",
            message_level="success",
            tasks=[],
        )
        self.assertIn("人工介入", html)
        self.assertIn("只读兼容状态", html)
        self.assertIn("正式复核请进入复核中心", html)
        self.assertNotIn("name='decision'", html)
        self.assertNotIn("value='resolve'", html)

    def test_render_runtime_page_requires_login_when_session_missing(self) -> None:
        html = render_runtime_page(
            params=default_runtime_state(),
            message="ok",
            message_level="success",
            tasks=[],
            reviews=[],
            notifications=[],
            session_user=None,
            selected_review=None,
            selected_task=None,
            selected_notification=None,
            task_history=[],
        )
        self.assertIn("SQLite", html)
        self.assertIn("运行态登录", html)
        self.assertIn("后台密码", html)
        self.assertIn("任务中心", html)

    def test_runtime_review_flow_requires_login_and_records_session_user(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-1", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                status, _, body = self._call_app(
                    path="/runtime",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("需要先登录", body)

                cookie = self._runtime_login(db_path)

                resolve_status, resolve_headers, _ = self._call_app(
                    path="/runtime",
                    method="POST",
                    cookie=cookie,
                    body=urlencode(
                        {
                            "action": "resolve_review",
                            "runtime_db": str(db_path),
                            "review_task_id": review.review_task_id,
                            "task_id": task.task_id,
                            "review_status": "approved",
                            "reviewer_code": "R-001",
                            "resolution_note": "web approved",
                            "resolution_payload_json": '{"adjustment":{"target_price":"8.8"}}',
                        }
                    ),
                )
                self.assertEqual(resolve_status, "303 See Other")
                location = resolve_headers["Location"]
                self.assertIn("review_task_id", location)

                detail = urlparse(location)
                detail_status, _, detail_body = self._call_app(
                    path=detail.path,
                    method="GET",
                    query=detail.query,
                    cookie=cookie,
                )
                self.assertEqual(detail_status, "200 OK")
                self.assertIn("admin", detail_body)
                self.assertIn("任务状态历史", detail_body)
                self.assertIn("notification_id", detail_body)
                self.assertIn("operations", detail_body)

                resolved_review = review_service.get_review_task(review.review_task_id)
                resolved_task = task_service.get_task(task.task_id)
                self.assertEqual(resolved_review.resolved_by, "admin")
                self.assertEqual(resolved_task.task_status, TaskStatus.PENDING)

    def test_runtime_review_repeat_submit_fails_after_first_resolution(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-1", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                resolve_body = urlencode(
                    {
                        "action": "resolve_review",
                        "runtime_db": str(db_path),
                        "review_task_id": review.review_task_id,
                        "task_id": task.task_id,
                        "review_status": "cancelled",
                        "resolution_note": "stop it",
                        "resolution_payload_json": "{}",
                    }
                )
                first_status, _, _ = self._call_app(
                    path="/runtime",
                    method="POST",
                    cookie=cookie,
                    body=resolve_body,
                )
                self.assertEqual(first_status, "303 See Other")

                second_status, _, second_body = self._call_app(
                    path="/runtime",
                    method="POST",
                    cookie=cookie,
                    body=resolve_body,
                )
                self.assertEqual(second_status, "200 OK")
                self.assertIn("已处理", second_body)

    def test_runtime_page_supports_minimal_filters_and_notification_detail(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task_a = _runtime_task("TASK-A", status=TaskStatus.MANUAL_REVIEW, trade_date=date(2026, 5, 4))
            task_b = _runtime_task("TASK-B", status=TaskStatus.PENDING, trade_date=date(2026, 5, 5))
            task_service.create_tasks([task_a, task_b])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task_a, task_b])
            review = review_service.list_review_tasks(status=ReviewTaskStatus.PENDING)[0]
            notification = repository.list_notification_logs(related_review_task_id=review.review_task_id)[0]

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/runtime",
                    method="GET",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "task_trade_date": "2026-05-04",
                            "task_status": "manual_review",
                            "review_trade_date": "2026-05-04",
                            "review_status": "pending",
                            "notification_related_review_task_id": review.review_task_id,
                            "notification_send_status": "success",
                            "notification_id": notification.notification_id,
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("任务筛选", body)
                self.assertIn("复核筛选", body)
                self.assertIn("通知筛选", body)
                self.assertIn("通知详情", body)
                self.assertIn(notification.notification_id, body)
                self.assertIn("send_status", body)

    def test_ops_pages_require_login_and_do_not_expose_runtime_data(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            routes = [
                "/dashboard",
                "/tasks",
                "/reviews",
                "/notifications",
                "/execution-logs",
                "/business-inputs",
                "/task-generator",
                "/system",
            ]
            for path in routes:
                with self.subTest(path=path):
                    status, _, body = self._call_app(
                        path=path,
                        method="GET",
                        query=urlencode({"runtime_db": str(db_path)}),
                    )
                    self.assertEqual(status, "200 OK")
                    self.assertIn("需要先登录", body)
                    self.assertNotIn("TASK-OPS-1", body)

    def test_task_generator_is_available_on_new_authenticated_route(self) -> None:
        db_path = self._existing_runtime_db("task-generator")
        with patch.dict(
            "os.environ",
            {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
            clear=False,
        ):
            cookie = self._runtime_login(db_path)
            status, _, body = self._call_app(
                path="/task-generator",
                method="GET",
                cookie=cookie,
            )
            self.assertEqual(status, "200 OK")
            self.assertIn("任务生成", body)
            self.assertNotIn("先校验数据", body)
            self.assertIn("预览时会自动完成数据校验", body)
            self.assertIn("预览任务", body)
            self.assertIn("/task-generator", body)

            status, _, body = self._call_app(
                path="/task-generator",
                method="POST",
                body=urlencode({"action": "validate", "selected_rule": "price:PRICE-TEST-B"}),
                cookie=cookie,
            )
            self.assertEqual(status, "200 OK")
            self.assertNotIn("CSRF_REJECTED", body)
            self.assertIn("任务生成", body)
            self.assertIn("未知任务生成操作", body)

            status, _, body = self._call_app(
                path="/task-generator",
                method="POST",
                body=urlencode({"action": "preview", "csrf_token": "invalid"}),
                cookie=cookie,
            )
            self.assertEqual(status, "403 Forbidden")
            self.assertIn("CSRF_REJECTED", body)

    def test_task_generator_previews_and_exports_one_selected_rule(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            output_path = root / "selected_rule_tasks.xlsx"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            for sku, variety, price, online_status, platform_stock_qty in (
                ("SKU-AISHA-A", "艾莎", "20", "online", 20),
                ("SKU-CAPPUCCINO-A", "卡布奇诺", "30", "online", 30),
                ("SKU-OFFLINE", "已下架品种", "40", "offline", 40),
                ("SKU-ZERO-STOCK", "零库存品种", "50", "online", 0),
            ):
                repository.upsert_listing_status(
                    ListingStatus(
                        listing_status_id=f"LISTING-{sku}",
                        platform_name="测试平台",
                        internal_sku=sku,
                        variety=variety,
                        current_price=Decimal(price),
                        grade="A",
                        online_status=online_status,
                        platform_stock_qty=platform_stock_qty,
                    )
                )
            save_table_records(
                "products",
                products_path,
                [
                    {
                        "internal_sku": "SKU-AISHA-A",
                        "product_name": "艾莎",
                        "grade": "A",
                        "stem_length": "70",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "8",
                        "sale_enabled": "True",
                    },
                    {
                        "internal_sku": "SKU-CAPPUCCINO-A",
                        "product_name": "卡布奇诺",
                        "grade": "A",
                        "stem_length": "70",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "8",
                        "sale_enabled": "True",
                    },
                    {
                        "internal_sku": "SKU-OFFLINE",
                        "product_name": "已下架品种",
                        "grade": "A",
                        "stem_length": "70",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "8",
                        "sale_enabled": "True",
                    },
                    {
                        "internal_sku": "SKU-ZERO-STOCK",
                        "product_name": "零库存品种",
                        "grade": "A",
                        "stem_length": "70",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "8",
                        "sale_enabled": "True",
                    },
                    {
                        "internal_sku": "SKU-WITHOUT-LISTING",
                        "product_name": "未上架品种",
                        "grade": "A",
                        "stem_length": "70",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "8",
                        "sale_enabled": "True",
                    },
                ],
            )
            save_table_records(
                "price_rules",
                price_rules_path,
                [
                    {
                        "rule_id": "PRICE-AISHA",
                        "rule_name": "全部商品加价",
                        "variety_filter": "*",
                        "grade_filter": "*",
                        "platform_filter": "*",
                        "pricing_method": "fixed_markup",
                        "markup_value": "5",
                        "min_price": "",
                        "rounding_rule": "none",
                        "rounding_step": "",
                        "active": "True",
                        "priority": "1",
                        "remark": "",
                    }
                ],
            )
            save_table_records("listing_rules", listing_rules_path, [])
            payload = {
                "products": str(products_path),
                "price_rules": str(price_rules_path),
                "listing_rules": str(listing_rules_path),
                "output": str(output_path),
                "inventory_strategy": "conservative_v1",
                "selected_rule": "price:PRICE-AISHA",
            }

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, preview_body = self._call_app(
                    path="/task-generator",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(payload | {"action": "preview"}),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("已预览 2 条任务", preview_body)
                self.assertIn("数据摘要", preview_body)
                self.assertIn("SKU-AISHA-A", preview_body)
                self.assertIn("SKU-CAPPUCCINO-A", preview_body)
                self.assertIn("已忽略商品（3）", preview_body)
                self.assertIn("SKU-OFFLINE", preview_body)
                self.assertIn("当前商品未上架，改价任务已忽略", preview_body)
                self.assertIn("SKU-ZERO-STOCK", preview_body)
                self.assertIn("平台库存为 0，商品不参与改价任务", preview_body)
                self.assertIn("SKU-WITHOUT-LISTING", preview_body)
                self.assertIn("未找到当前平台上架状态，改价任务已忽略", preview_body)
                self.assertIn("原任务类型", preview_body)
                self.assertIn("当前平台状态", preview_body)
                self.assertIn("忽略原因", preview_body)
                group_match = re.search(r'name="task_group_id" value="([^"]+)"', preview_body)
                deadline_match = re.search(r'name="required_by" value="([^"]+)"', preview_body)
                self.assertIsNotNone(group_match)
                self.assertIsNotNone(deadline_match)
                generated_group_id = group_match.group(1)
                generated_deadline = deadline_match.group(1)
                self.assertTrue(generated_group_id.startswith("RULE-GROUP-"))
                self.assertIn("测试平台", preview_body)
                self.assertIn("确认生成并进入任务中心", preview_body)
                self.assertFalse(output_path.exists())

                status, _, generated_body = self._call_app(
                    path="/task-generator",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(payload | {
                        "action": "confirm_generate",
                        "task_group_id": generated_group_id,
                        "required_by": generated_deadline,
                    }),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("已生成 2 条任务", generated_body)
                self.assertIn("2 条新任务已进入任务中心", generated_body)

                status, _, tasks_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("SKU-AISHA-A", tasks_body)
                self.assertIn("SKU-CAPPUCCINO-A", tasks_body)
                self.assertNotIn("SKU-OFFLINE", tasks_body)
                self.assertNotIn("SKU-ZERO-STOCK", tasks_body)
                self.assertNotIn("SKU-WITHOUT-LISTING", tasks_body)
                self.assertIn(generated_group_id, tasks_body)
                self.assertIn("任务组状态", tasks_body)
                self.assertIn("预期原价", tasks_body)
                self.assertIn("old_price=20", tasks_body)

                status, _, duplicate_body = self._call_app(
                    path="/task-generator",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(payload | {
                        "action": "confirm_generate",
                        "task_group_id": generated_group_id,
                        "required_by": generated_deadline,
                    }),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("0 条新任务已进入任务中心", duplicate_body)
                self.assertIn("2 条因重复未再次入库", duplicate_body)

            tasks = load_tasks(output_path)
            self.assertEqual(
                [task.internal_sku for task in tasks],
                ["SKU-AISHA-A", "SKU-CAPPUCCINO-A"],
            )
            self.assertTrue(all(task.decision_trace["selected_rule_id"] == "PRICE-AISHA" for task in tasks))
            self.assertEqual(len({task.decision_trace["task_group_id"] for task in tasks}), 1)
            self.assertEqual(len({task.required_by for task in tasks}), 1)
            self.assertEqual({task.platform_name for task in tasks}, {"测试平台"})
            self.assertEqual({task.expected_old_price for task in tasks}, {Decimal("20"), Decimal("30")})
            self.assertEqual({task.target_price for task in tasks}, {Decimal("25"), Decimal("35")})
            runtime_tasks = RuntimeTaskService(SQLiteRuntimeRepository(db_path)).list_tasks()
            self.assertEqual(
                sorted(task.internal_sku for task in runtime_tasks),
                ["SKU-AISHA-A", "SKU-CAPPUCCINO-A"],
            )
            self.assertEqual(len({task.decision_trace["task_group_id"] for task in runtime_tasks}), 1)
            self.assertEqual(len({task.required_by for task in runtime_tasks}), 1)
            self.assertEqual(
                {task.expected_old_price for task in runtime_tasks},
                {Decimal("20"), Decimal("30")},
            )
            self.assertEqual({task.task_status for task in runtime_tasks}, {TaskStatus.PENDING})

    def test_listing_task_requires_editable_price_and_inventory_with_snapshot_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            output_path = root / "listing_tasks.xlsx"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            observed_at = datetime(2026, 7, 25, 8, 0, tzinfo=timezone.utc)
            repository.upsert_listing_status(
                ListingStatus(
                    listing_status_id="LISTING-ROSE-A",
                    platform_name="测试平台",
                    internal_sku="SKU-ROSE-A",
                    variety="测试玫瑰",
                    current_price=Decimal("18.60"),
                    grade="A",
                    online_status="offline",
                    platform_stock_qty=7,
                    price_source="shadowbot_read",
                    price_observed_at=observed_at,
                    price_source_attempt_id="ATTEMPT-SNAPSHOT-1",
                    inventory_source="shadowbot_read",
                    inventory_observed_at=observed_at,
                    inventory_source_attempt_id="ATTEMPT-SNAPSHOT-1",
                )
            )
            save_table_records(
                "products",
                products_path,
                [
                    {
                        "internal_sku": "SKU-ROSE-A",
                        "product_name": "测试玫瑰",
                        "grade": "A",
                        "stem_length": "70",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "30",
                        "sale_enabled": "True",
                    }
                ],
            )
            save_table_records("price_rules", price_rules_path, [])
            save_table_records(
                "listing_rules",
                listing_rules_path,
                [
                    {
                        "rule_id": "LIST-ONLINE",
                        "rule_name": "库存充足时上架",
                        "variety_filter": "测试玫瑰",
                        "grade_filter": "A",
                        "platform_filter": "测试平台",
                        "stock_threshold": "10",
                        "listing_strategy": "stock_above_online",
                        "active": "True",
                        "priority": "1",
                        "remark": "",
                    }
                ],
            )
            payload = {
                "products": str(products_path),
                "price_rules": str(price_rules_path),
                "listing_rules": str(listing_rules_path),
                "output": str(output_path),
                "inventory_strategy": "conservative_v1",
                "generation_mode": "single_rule",
                "selected_rule": "listing:LIST-ONLINE",
            }

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, preview_body = self._call_app(
                    path="/task-generator",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(payload | {"action": "preview"}),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("已预览 1 条任务", preview_body)
                self.assertIn("上架任务的目标价格和目标库存必须在创建前确认", preview_body)
                self.assertEqual(
                    preview_body.count('class="task-target-number"'),
                    2,
                )
                self.assertIn(
                    ".task-target-number::-webkit-inner-spin-button",
                    preview_body,
                )
                self.assertIn("min-width: 120px", preview_body)
                self.assertRegex(
                    preview_body,
                    r'name="listing_target_price_0"[\s\S]+?value="18\.60"',
                )
                self.assertRegex(
                    preview_body,
                    r'name="listing_target_inventory_0"[\s\S]+?value="7"',
                )
                group_match = re.search(
                    r'name="task_group_id" value="([^"]+)"',
                    preview_body,
                )
                deadline_match = re.search(
                    r'name="required_by" value="([^"]+)"',
                    preview_body,
                )
                self.assertIsNotNone(group_match)
                self.assertIsNotNone(deadline_match)
                group_id = group_match.group(1)
                required_by = deadline_match.group(1)

                status, _, missing_body = self._call_app(
                    path="/task-generator",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(
                        payload
                        | {
                            "action": "confirm_generate",
                            "task_group_id": group_id,
                            "required_by": required_by,
                            "listing_override_count": "0",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("上架任务必须输入目标价格和目标库存", missing_body)
                self.assertFalse(output_path.exists())
                self.assertEqual(RuntimeTaskService(repository).list_tasks(), [])

                override_key = json.dumps(
                    ["测试平台", "SKU-ROSE-A", "set_online"],
                    ensure_ascii=False,
                    separators=(",", ":"),
                )
                status, _, generated_body = self._call_app(
                    path="/task-generator",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(
                        payload
                        | {
                            "action": "confirm_generate",
                            "task_group_id": group_id,
                            "required_by": required_by,
                            "listing_override_count": "1",
                            "listing_override_key_0": override_key,
                            "listing_target_price_0": "19.80",
                            "listing_target_inventory_0": "9",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("已生成 1 条任务", generated_body)
                self.assertIn("1 条新任务已进入任务中心", generated_body)

            runtime_tasks = RuntimeTaskService(repository).list_tasks()
            self.assertEqual(len(runtime_tasks), 1)
            self.assertEqual(runtime_tasks[0].target_price, Decimal("19.80"))
            self.assertEqual(runtime_tasks[0].target_inventory, 9)
            self.assertEqual(
                runtime_tasks[0].decision_trace["listing_target_input"],
                {
                    "source": "operator_confirm_form",
                    "default_source": "platform_snapshot",
                    "target_price": "19.80",
                    "target_inventory": 9,
                },
            )
            exported_tasks = load_tasks(output_path)
            self.assertEqual(exported_tasks[0].target_price, Decimal("19.80"))
            self.assertEqual(exported_tasks[0].target_inventory, 9)

    def test_new_product_without_snapshot_is_recognized_for_set_online_preview(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            repository.upsert_listing_status(
                ListingStatus(
                    listing_status_id="LISTING-AISHA-A",
                    platform_name="蚂蚁花团供应商",
                    internal_sku="AISHA-A-70-Z",
                    variety="艾莎",
                    grade="A",
                    current_price=Decimal("18.00"),
                    online_status="online",
                    platform_stock_qty=10,
                    source="shadowbot_read",
                )
            )
            save_table_records(
                "products",
                products_path,
                [
                    {
                        "internal_sku": "AISHA-E-45-Z",
                        "product_name": "艾莎",
                        "grade": "E",
                        "stem_length": "45",
                        "unit": "扎",
                        "base_cost": "4",
                        "current_stock": "20",
                        "sale_enabled": "True",
                    }
                ],
            )
            save_table_records("price_rules", price_rules_path, [])
            save_table_records(
                "listing_rules",
                listing_rules_path,
                [
                    {
                        "rule_id": "LIST-AISHA-E-ONLINE",
                        "rule_name": "蚂蚁艾莎E级上架",
                        "variety_filter": "艾莎",
                        "grade_filter": "E",
                        "platform_filter": "蚂蚁",
                        "stock_threshold": "0",
                        "listing_strategy": "allow_online",
                        "active": "True",
                        "priority": "1",
                        "remark": "",
                    }
                ],
            )
            payload = {
                "products": str(products_path),
                "price_rules": str(price_rules_path),
                "listing_rules": str(listing_rules_path),
                "inventory_strategy": "conservative_v1",
                "generation_mode": "single_rule",
                "selected_rule": "listing:LIST-AISHA-E-ONLINE",
            }

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, preview_body = self._call_app(
                    path="/task-generator",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(payload | {"action": "preview"}),
                    cookie=cookie,
                )

            self.assertEqual(status, "200 OK")
            self.assertIn("已预览 1 条任务", preview_body)
            self.assertIn("AISHA-E-45-Z", preview_body)
            self.assertIn(
                "expected_old_price 和目标价格默认使用基础成本",
                preview_body,
            )
            self.assertIn(
                'title="新商品无平台快照，默认值来自商品基础成本"',
                preview_body,
            )
            self.assertIn(
                'title="新商品无平台快照，默认值来自商品当前库存"',
                preview_body,
            )
            self.assertRegex(
                preview_body,
                r"<td>4</td>[\s\S]+?name=\"listing_target_price_0\"[\s\S]+?value=\"4\.00\"",
            )
            self.assertRegex(
                preview_body,
                r'name="listing_target_inventory_0"[\s\S]+?value="20"',
            )

    def test_task_group_status_is_shared_and_derived_from_item_statuses(self) -> None:
        group_trace = {"task_group_id": "RULE-GROUP-STATUS-001"}
        pending_a = replace(_runtime_task("TASK-GROUP-A"), decision_trace=group_trace)
        pending_b = replace(_runtime_task("TASK-GROUP-B"), decision_trace=group_trace)
        self.assertEqual(
            _build_task_group_statuses([pending_a, pending_b]),
            {"RULE-GROUP-STATUS-001": "待处理"},
        )
        partial = replace(pending_b, task_status=TaskStatus.SUCCESS)
        self.assertEqual(
            _build_task_group_statuses([pending_a, partial]),
            {"RULE-GROUP-STATUS-001": "处理中（部分完成）"},
        )
        failed = replace(pending_b, task_status=TaskStatus.FAILED)
        self.assertEqual(
            _build_task_group_statuses([partial, failed]),
            {"RULE-GROUP-STATUS-001": "部分失败"},
        )

    def test_ops_pages_render_lists_and_empty_states_after_login(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            empty_db_path = Path(temp_dir) / "empty.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            task = _runtime_task("TASK-OPS-1", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service.create_from_tasks([task])
            RuntimeTaskService(SQLiteRuntimeRepository(empty_db_path)).init_schema()

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                empty_cookie = self._runtime_login(empty_db_path)

                tasks_status, _, tasks_body = self._call_app(
                    path="/dashboard",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                self.assertEqual(tasks_status, "200 OK")
                self.assertIn("首页总览", tasks_body)
                self.assertIn("任务中心", tasks_body)

                tasks_status, _, tasks_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                self.assertEqual(tasks_status, "200 OK")
                self.assertIn("TASK-OPS-1", tasks_body)
                self.assertIn("处理对象", tasks_body)
                self.assertIn("等待人工确认", tasks_body)

                reviews_status, _, reviews_body = self._call_app(
                    path="/reviews",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                self.assertEqual(reviews_status, "200 OK")
                self.assertIn("TASK-OPS-1", reviews_body)
                self.assertIn("/reviews?review_task_id=", reviews_body)
                self.assertNotIn("resolution_payload_json", reviews_body)

                notifications_status, _, notifications_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                self.assertEqual(notifications_status, "200 OK")
                self.assertIn("发送时间", notifications_body)
                self.assertIn("unconfigured", notifications_body)
                self.assertIn("待发送", notifications_body)
                self.assertNotIn("token=", notifications_body)

                execution_status, _, execution_body = self._call_app(
                    path="/execution-logs",
                    method="GET",
                    query=urlencode({"runtime_db": str(empty_db_path)}),
                    cookie=empty_cookie,
                )
                self.assertEqual(execution_status, "200 OK")
                self.assertIn("当前还没有执行器回写结果", execution_body)

                business_inputs_status, _, business_inputs_body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                self.assertEqual(business_inputs_status, "200 OK")
                self.assertIn("商品资料与库存录入", business_inputs_body)

                _, _, empty_tasks_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(empty_db_path)}),
                    cookie=empty_cookie,
                )
                self.assertIn("当前没有待执行或待处理任务", empty_tasks_body)
                _, _, empty_reviews_body = self._call_app(
                    path="/reviews",
                    method="GET",
                    query=urlencode({"runtime_db": str(empty_db_path)}),
                    cookie=empty_cookie,
                )
                self.assertIn("当前没有需要人工确认的事项", empty_reviews_body)
                _, _, empty_notifications_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(empty_db_path)}),
                    cookie=empty_cookie,
                )
                self.assertIn("当前还没有飞书或系统通知记录", empty_notifications_body)

    def test_listing_status_page_is_read_only_but_debug_post_is_available(self) -> None:
        with _workspace_temp_dir("web_listing_status") as temp_dir:
            db_path = temp_dir / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            repository.upsert_listing_status(
                ListingStatus(
                    listing_status_id="LISTING-VISIBLE",
                    platform_name="蚂蚁花团供应商",
                    internal_sku="",
                    variety="页面可见品种",
                    grade="A",
                    current_price=Decimal("20"),
                    platform_stock_qty=12,
                    online_status="online",
                )
            )
            repository.upsert_listing_status(
                ListingStatus(
                    listing_status_id="LISTING-ZERO-STOCK",
                    platform_name="蚂蚁花团供应商",
                    internal_sku="",
                    variety="页面隐藏零库存品种",
                    grade="A",
                    current_price=Decimal("30"),
                    platform_stock_qty=0,
                    online_status="online",
                )
            )

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "input_tab": "listing_status"}),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("本页面只读，不提供人工录入或修改", body)
                self.assertIn("页面可见品种", body)
                self.assertNotIn("页面隐藏零库存品种", body)
                self.assertNotIn('name="action" value="save_listing_status"', body)
                self.assertNotIn('id="listing_platform"', body)

                post_status, headers, _ = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    cookie=cookie,
                    body=urlencode({
                        "action": "save_listing_status",
                        "runtime_db": str(db_path),
                        "input_tab": "listing_status",
                        "platform_name": "ant_flower_wechat",
                        "variety": "艾莎",
                        "grade": "C级",
                        "current_price": "19.50",
                        "sold_qty": "2",
                        "online_status": "online",
                    }),
                )

            self.assertEqual(post_status, "303 See Other")
            self.assertIn("调试用上架状态已保存", unquote(headers["Location"]))
            saved = repository.get_listing_status("ant_flower_wechat", "艾莎", "C")
            self.assertIsNotNone(saved)
            assert saved is not None
            self.assertEqual(saved.current_price, Decimal("19.50"))
            self.assertEqual(saved.source, "debug_web_request")

    def test_tasks_automation_tab_requires_login_and_shows_script_runs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            repository.insert_script_run(
                ScriptRun(
                    script_run_id="SCRIPT-RUN-1",
                    evaluator_id="capacity_warning",
                    evaluator_name="包装产能预警",
                    description="dry-run test",
                    run_mode="dry-run",
                    run_status="success",
                    trade_date=date(2026, 5, 8),
                    started_at=datetime(2026, 5, 7, 10, 0),
                    finished_at=datetime(2026, 5, 7, 10, 1),
                    summary={"proposals_count": 1, "inserted_review_tasks_count": 0},
                    created_by="test",
                )
            )
            repository.insert_script_run_items(
                [
                    ScriptRunItem(
                        item_id="ITEM-1",
                        script_run_id="SCRIPT-RUN-1",
                        proposal_type="review_task",
                        dedupe_key="dedupe-1",
                        severity="warning",
                        item_status="previewed",
                        message="capacity warning preview",
                        payload={"action_type": "capacity_warning"},
                        decision_trace={"run_mode": "dry-run"},
                        created_at=datetime(2026, 5, 7, 10, 0),
                    )
                ]
            )
            status, _, body = self._call_app(
                path="/tasks",
                method="GET",
                query=urlencode({"runtime_db": str(db_path), "task_tab": "automation"}),
            )
            self.assertEqual(status, "200 OK")
            self.assertNotIn("SCRIPT-RUN-1", body)

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "task_tab": "automation",
                            "script_run_id": "SCRIPT-RUN-1",
                        }
                    ),
                    cookie=cookie,
                )
            self.assertEqual(status, "200 OK")
            self.assertIn("SCRIPT-RUN-1", body)
            self.assertIn("dry-run", body)
            self.assertIn("ITEM-1", body)
            self.assertNotIn("token=", body)
            self.assertNotIn("FEISHU_WEBHOOK_SECRET", body)

    def test_tasks_mock_platform_tab_requires_login_and_shows_state(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            mock_db_path = Path(temp_dir) / "mock_platform.sqlite3"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            mock_repository = MockPlatformRepository(mock_db_path)
            mock_repository.init_schema()
            mock_repository.upsert_product_states(
                [
                    MockPlatformProductState(
                        platform_name="default_platform",
                        internal_sku="SKU-001",
                        platform_sku="MP-SKU-001",
                        product_name="艾莎",
                        grade="A",
                        platform_price=Decimal("18"),
                        platform_online_status="online",
                        platform_stock_qty=12,
                        last_platform_update_at=datetime(2026, 5, 7, 10, 0),
                    )
                ]
            )
            query = urlencode(
                {
                    "runtime_db": str(db_path),
                    "task_tab": "mock_platform",
                    "mock_platform_db": str(mock_db_path),
                }
            )
            status, _, body = self._call_app(path="/tasks", method="GET", query=query)
            self.assertEqual(status, "200 OK")
            self.assertNotIn("SKU-001", body)

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=query,
                    cookie=cookie,
                )
            self.assertEqual(status, "200 OK")
            self.assertIn("Mock 平台状态", body)
            self.assertIn("SKU-001", body)
            self.assertIn("已上架", body)
            self.assertNotIn("token=", body)
            self.assertNotIn(str(mock_db_path), body)

    def test_business_inputs_adds_inventory_to_products_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records(
                "products",
                products_path,
                [
                    {
                        "internal_sku": "AISHA-B-FG-Z",
                        "product_name": "艾莎",
                        "grade": "B",
                        "stem_length": "跟随等级",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "5",
                        "sale_enabled": "True",
                        "last_price": "",
                        "recommended_price": "",
                        "remark": "",
                        "feature_season": "",
                        "feature_color": "",
                    }
                ],
            )

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "products_path": str(products_path)}),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("录入库存", body)
                self.assertIn("价格规则管理", body)
                self.assertNotIn('id="price_rule_name"', body)
                self.assertIn("是", body)

                post_body = urlencode(
                    {
                        "products_path": str(products_path),
                        "action": "add_inventory",
                        "product_name": "艾莎 ",
                        "grade": "b",
                        "stem_length": "FG",
                        "unit": "扎",
                        "base_cost": "12",
                        "quantity": "4",
                        "sale_enabled": "false",
                    }
                )
                status, headers, _ = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path), "products_path": str(products_path)}),
                    body=post_body,
                    cookie=cookie,
                )
                self.assertEqual(status, "303 See Other")
                self.assertIn("business-inputs", headers["Location"])
                self.assertIn("input_tab=inventory", headers["Location"])
                location = urlparse(headers["Location"])
                status, _, body = self._call_app(
                    path=location.path,
                    query=location.query,
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("inventory_feedback_dialog", body)
                self.assertIn("保存成功", body)
                self.assertIn("已补充库存", body)

            products = load_products(products_path)
            self.assertEqual(len(products), 1)
            self.assertEqual(products[0].internal_sku, "AISHA-B-FG-Z")
            self.assertEqual(products[0].stem_length, "60")
            self.assertEqual(products[0].current_stock, 9)
            self.assertEqual(products[0].base_cost, 12)
            self.assertFalse(products[0].sale_enabled)

    def test_business_inputs_adds_price_rule_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            platform_mappings_path = root / "platform_mappings.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records(
                "products",
                products_path,
                [
                    {
                        "internal_sku": "SKU-001",
                        "product_name": "艾莎",
                        "grade": "A",
                        "stem_length": "65",
                        "unit": "扎",
                        "base_cost": "10",
                        "current_stock": "8",
                        "sale_enabled": "True",
                    },
                    {
                        "internal_sku": "SKU-002",
                        "product_name": "卡布奇诺",
                        "grade": "B",
                        "stem_length": "60",
                        "unit": "扎",
                        "base_cost": "12",
                        "current_stock": "6",
                        "sale_enabled": "True",
                    },
                ],
            )
            save_table_records("price_rules", price_rules_path, [])
            save_table_records(
                "platform_mappings",
                platform_mappings_path,
                [
                    {
                        "mapping_id": "PLATFORM-CUSTOM",
                        "platform_name": "测试平台",
                        "mapping_status": "active",
                        "search_keyword": "测试平台",
                        "remark": "unit test",
                    }
                ],
            )

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "products_path": str(products_path),
                            "price_rules_path": str(price_rules_path),
                            "platform_mappings_path": str(platform_mappings_path),
                            "input_tab": "price_rules",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("价格规则管理", body)
                self.assertIn("新增价格规则", body)
                self.assertIn('id="price_rule_name"', body)
                self.assertNotIn('id="quantity"', body)
                self.assertIn('name="variety_filter"', body)
                self.assertIn('name="grade_filter"', body)
                self.assertIn('name="platform_filter"', body)
                self.assertIn("不限制", body)
                self.assertIn("艾莎", body)
                self.assertIn("寻梦", body)
                self.assertIn("测试平台", body)

                post_body = urlencode(
                    {
                        "price_rules_path": str(price_rules_path),
                        "action": "add_price_rule",
                        "rule_name": "全局固定加价",
                        "variety_filter": "*",
                        "grade_filter": "*",
                        "platform_filter": "*",
                        "pricing_method": "fixed_markup",
                        "markup_value": "3",
                        "min_price": "5",
                        "rounding_rule": "round",
                        "rounding_step": "",
                        "active": "true",
                        "priority": "10",
                        "remark": "Web 表单录入",
                    }
                )
                status, headers, _ = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "products_path": str(products_path),
                            "price_rules_path": str(price_rules_path),
                            "platform_mappings_path": str(platform_mappings_path),
                        }
                    ),
                    body=post_body,
                    cookie=cookie,
                )
                self.assertEqual(status, "303 See Other")
                self.assertIn("price_rules_path", headers["Location"])
                self.assertIn("input_tab=price_rules", headers["Location"])

            rules = load_price_rules(price_rules_path)
            self.assertEqual(len(rules), 1)
            self.assertEqual(rules[0].rule_name, "全局固定加价")
            self.assertEqual(rules[0].variety_filter, "*")
            self.assertEqual(rules[0].grade_filter, "*")
            self.assertEqual(rules[0].platform_filter, "*")
            self.assertEqual(rules[0].markup_value, Decimal("3"))

    def test_business_inputs_adds_platform_mapping_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            platform_mappings_path = root / "platform_mappings.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records("products", products_path, [])
            save_table_records("price_rules", price_rules_path, [])
            save_table_records("platform_mappings", platform_mappings_path, [])

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "products_path": str(products_path),
                            "price_rules_path": str(price_rules_path),
                            "platform_mappings_path": str(platform_mappings_path),
                            "input_tab": "inventory",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("新增平台", body)

                status, headers, _ = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "products_path": str(products_path),
                            "price_rules_path": str(price_rules_path),
                            "platform_mappings_path": str(platform_mappings_path),
                            "input_tab": "inventory",
                        }
                    ),
                    body=urlencode(
                        {
                            "platform_mappings_path": str(platform_mappings_path),
                            "input_tab": "inventory",
                            "action": "add_platform",
                            "platform_name": "新测试平台",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "303 See Other")
                self.assertIn("input_tab=inventory", headers["Location"])

            rows = load_table_records("platform_mappings", platform_mappings_path)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["platform_name"], "新测试平台")
            self.assertEqual(rows[0]["mapping_status"], "active")

    def test_business_inputs_price_rules_falls_back_when_platform_mapping_file_missing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            platform_mappings_path = root / "missing_platform_mappings.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records("products", products_path, [])
            save_table_records("price_rules", price_rules_path, [])

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "products_path": str(products_path),
                            "price_rules_path": str(price_rules_path),
                            "platform_mappings_path": str(platform_mappings_path),
                            "input_tab": "price_rules",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("蚂蚁", body)
                self.assertIn('name="platform_filter"', body)

    def test_business_inputs_adds_direct_set_offline_listing_rule(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records("products", products_path, [])
            save_table_records("price_rules", price_rules_path, [])
            save_table_records("listing_rules", listing_rules_path, [])

            query = urlencode(
                {
                    "runtime_db": str(db_path),
                    "products_path": str(products_path),
                    "price_rules_path": str(price_rules_path),
                    "listing_rules_path": str(listing_rules_path),
                    "input_tab": "listing_rules",
                }
            )
            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=query,
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("value='set_offline'", body)
                self.assertIn("直接下架（set_offline）", body)

                status, headers, _ = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=query,
                    body=urlencode(
                        {
                            "listing_rules_path": str(listing_rules_path),
                            "input_tab": "listing_rules",
                            "action": "add_listing_rule",
                            "rule_name": "全部商品直接下架",
                            "variety_filter": "*",
                            "grade_filter": "*",
                            "platform_filter": "*",
                            "stock_threshold": "0",
                            "listing_strategy": "set_offline",
                            "active": "true",
                            "priority": "1",
                            "remark": "Web 表单录入",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "303 See Other")
                self.assertIn("input_tab=listing_rules", headers["Location"])

            rules = load_listing_rules(listing_rules_path)
            self.assertEqual(len(rules), 1)
            self.assertEqual(rules[0].listing_strategy, ListingStrategy.SET_OFFLINE)

    def test_business_inputs_adds_capacity_plan_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            capacity_plans_path = root / "capacity_plans.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records("products", products_path, [])
            save_table_records("price_rules", price_rules_path, [])
            save_table_records("listing_rules", listing_rules_path, [])
            from openpyxl import Workbook
            from app.repositories.workbook_repository import CAPACITY_PLAN_HEADERS

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "data"
            sheet.append(CAPACITY_PLAN_HEADERS)
            workbook.save(capacity_plans_path)

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                query = urlencode(
                    {
                        "runtime_db": str(db_path),
                        "products_path": str(products_path),
                        "price_rules_path": str(price_rules_path),
                        "listing_rules_path": str(listing_rules_path),
                        "capacity_plans_path": str(capacity_plans_path),
                        "input_tab": "capacity_plans",
                    }
                )
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=query,
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("包装产能计划", body)
                self.assertIn('name="confirmed_packing_capacity_qty"', body)
                self.assertIn("新增包装产能计划", body)

                status, headers, _ = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=query,
                    body=urlencode(
                        {
                            "capacity_plans_path": str(capacity_plans_path),
                            "input_tab": "capacity_plans",
                            "action": "add_capacity_plan",
                            "trade_date": "2026-05-08",
                            "normal_packing_capacity_qty": "250",
                            "confirmed_temp_worker_count": "2",
                            "temp_worker_capacity_qty": "100",
                            "confirmed_packing_capacity_qty": "460",
                            "active": "true",
                            "note": "Web 表单录入",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "303 See Other")
                self.assertIn("input_tab=capacity_plans", headers["Location"])

            plans = load_capacity_plans(capacity_plans_path)
            self.assertEqual(len(plans), 1)
            self.assertEqual(plans[0].trade_date.isoformat(), "2026-05-08")
            self.assertEqual(plans[0].confirmed_packing_capacity_qty, 460)

    def test_business_inputs_rejects_duplicate_active_capacity_plan(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            capacity_plans_path = root / "capacity_plans.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records("products", products_path, [])
            save_table_records("price_rules", price_rules_path, [])
            save_table_records("listing_rules", listing_rules_path, [])
            save_table_records(
                "capacity_plans",
                capacity_plans_path,
                [
                    {
                        "trade_date": "2026-05-08",
                        "normal_packing_capacity_qty": 250,
                        "temp_worker_capacity_qty": 100,
                        "confirmed_temp_worker_count": 0,
                        "confirmed_packing_capacity_qty": 250,
                        "allocation_rule": "proportional_by_forecast",
                        "active": True,
                        "note": "",
                    }
                ],
            )

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                query = urlencode(
                    {
                        "runtime_db": str(db_path),
                        "products_path": str(products_path),
                        "price_rules_path": str(price_rules_path),
                        "listing_rules_path": str(listing_rules_path),
                        "capacity_plans_path": str(capacity_plans_path),
                        "input_tab": "capacity_plans",
                    }
                )
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=query,
                    body=urlencode(
                        {
                            "capacity_plans_path": str(capacity_plans_path),
                            "input_tab": "capacity_plans",
                            "action": "add_capacity_plan",
                            "trade_date": "2026-05-08",
                            "normal_packing_capacity_qty": "250",
                            "confirmed_temp_worker_count": "1",
                            "temp_worker_capacity_qty": "100",
                            "confirmed_packing_capacity_qty": "350",
                            "active": "true",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("同一业务日期已经存在启用", body)

    def test_business_inputs_adds_cold_storage_status_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            capacity_plans_path = root / "capacity_plans.xlsx"
            cold_storage_status_path = root / "cold_storage_status.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records("products", products_path, [])
            save_table_records("price_rules", price_rules_path, [])
            save_table_records("listing_rules", listing_rules_path, [])
            from openpyxl import Workbook
            from app.repositories.workbook_repository import CAPACITY_PLAN_HEADERS, COLD_STORAGE_STATUS_HEADERS

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "data"
            sheet.append(CAPACITY_PLAN_HEADERS)
            workbook.save(capacity_plans_path)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "data"
            sheet.append(COLD_STORAGE_STATUS_HEADERS)
            workbook.save(cold_storage_status_path)

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                query = urlencode(
                    {
                        "runtime_db": str(db_path),
                        "products_path": str(products_path),
                        "price_rules_path": str(price_rules_path),
                        "listing_rules_path": str(listing_rules_path),
                        "capacity_plans_path": str(capacity_plans_path),
                        "cold_storage_status_path": str(cold_storage_status_path),
                        "input_tab": "cold_storage_status",
                    }
                )
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="GET",
                    query=query,
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("冷库状态", body)
                self.assertIn('name="projected_occupied_qty"', body)
                self.assertIn("新增冷库状态", body)

                status, headers, _ = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=query,
                    body=urlencode(
                        {
                            "cold_storage_status_path": str(cold_storage_status_path),
                            "input_tab": "cold_storage_status",
                            "action": "add_cold_storage_status",
                            "trade_date": "2026-05-08",
                            "total_capacity_qty": "500",
                            "current_occupied_qty": "120",
                            "expected_inbound_qty": "80",
                            "expected_outbound_qty": "20",
                            "warning_threshold_qty": "50",
                            "projected_occupied_qty": "",
                            "remaining_capacity_qty": "",
                            "active": "true",
                            "note": "Web 表单录入",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "303 See Other")
                self.assertIn("input_tab=cold_storage_status", headers["Location"])

            statuses = load_cold_storage_statuses(cold_storage_status_path)
            self.assertEqual(len(statuses), 1)
            self.assertEqual(statuses[0].trade_date.isoformat(), "2026-05-08")
            self.assertEqual(statuses[0].projected_occupied_qty, 180)
            self.assertEqual(statuses[0].remaining_capacity_qty, 320)

    def test_business_inputs_rejects_duplicate_active_cold_storage_status(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            db_path = root / "runtime.sqlite3"
            products_path = root / "products.xlsx"
            price_rules_path = root / "price_rules.xlsx"
            listing_rules_path = root / "listing_rules.xlsx"
            capacity_plans_path = root / "capacity_plans.xlsx"
            cold_storage_status_path = root / "cold_storage_status.xlsx"
            RuntimeTaskService(SQLiteRuntimeRepository(db_path)).init_schema()
            save_table_records("products", products_path, [])
            save_table_records("price_rules", price_rules_path, [])
            save_table_records("listing_rules", listing_rules_path, [])
            from openpyxl import Workbook
            from app.repositories.workbook_repository import CAPACITY_PLAN_HEADERS

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "data"
            sheet.append(CAPACITY_PLAN_HEADERS)
            workbook.save(capacity_plans_path)
            save_table_records(
                "cold_storage_status",
                cold_storage_status_path,
                [
                    {
                        "trade_date": "2026-05-08",
                        "total_capacity_qty": 500,
                        "current_occupied_qty": 120,
                        "expected_inbound_qty": 0,
                        "expected_outbound_qty": 0,
                        "warning_threshold_qty": 50,
                        "projected_occupied_qty": 120,
                        "remaining_capacity_qty": 380,
                        "active": True,
                        "note": "",
                    }
                ],
            )

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                query = urlencode(
                    {
                        "runtime_db": str(db_path),
                        "products_path": str(products_path),
                        "price_rules_path": str(price_rules_path),
                        "listing_rules_path": str(listing_rules_path),
                        "capacity_plans_path": str(capacity_plans_path),
                        "cold_storage_status_path": str(cold_storage_status_path),
                        "input_tab": "cold_storage_status",
                    }
                )
                status, _, body = self._call_app(
                    path="/business-inputs",
                    method="POST",
                    query=query,
                    body=urlencode(
                        {
                            "cold_storage_status_path": str(cold_storage_status_path),
                            "input_tab": "cold_storage_status",
                            "action": "add_cold_storage_status",
                            "trade_date": "2026-05-08",
                            "total_capacity_qty": "500",
                            "current_occupied_qty": "200",
                            "expected_inbound_qty": "0",
                            "expected_outbound_qty": "0",
                            "warning_threshold_qty": "50",
                            "active": "true",
                        }
                    ),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("同一业务日期已经存在启用", body)

    def test_dashboard_metrics_and_filtered_list_links(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            now = datetime.now()
            pending_task = _runtime_task("TASK-PENDING", status=TaskStatus.PENDING)
            expired_task = _runtime_task("TASK-EXPIRED", status=TaskStatus.EXPIRED)
            failed_task = _runtime_task("TASK-FAILED", status=TaskStatus.FAILED)
            review_due = _runtime_task(
                "TASK-REVIEW-DUE",
                status=TaskStatus.MANUAL_REVIEW,
                required_by=now + timedelta(minutes=45),
            )
            review_due_utc = _runtime_task(
                "TASK-REVIEW-DUE-UTC",
                status=TaskStatus.MANUAL_REVIEW,
                required_by=(now + timedelta(minutes=45))
                .replace(tzinfo=timezone(timedelta(hours=8)))
                .astimezone(timezone.utc),
            )
            review_later = _runtime_task(
                "TASK-REVIEW-LATER",
                status=TaskStatus.MANUAL_REVIEW,
                required_by=now + timedelta(hours=3),
            )
            review_overdue = _runtime_task(
                "TASK-REVIEW-OVERDUE",
                status=TaskStatus.MANUAL_REVIEW,
                required_by=now - timedelta(minutes=5),
            )
            review_expired_source = _runtime_task("TASK-REVIEW-EXPIRED", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks(
                [
                    pending_task,
                    expired_task,
                    failed_task,
                    review_due,
                    review_due_utc,
                    review_later,
                    review_overdue,
                    review_expired_source,
                ]
            )
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks(
                [review_due, review_due_utc, review_later, review_overdue, review_expired_source]
            )
            expired_review = repository.list_review_tasks(status=ReviewTaskStatus.PENDING)[-1]
            repository.update_review_task(
                replace(
                    expired_review,
                    review_status=ReviewTaskStatus.EXPIRED,
                    updated_at=now,
                    resolved_at=now,
                    resolved_by="system",
                )
            )
            repository.insert_notification_logs(
                [
                    NotificationLog(
                        notification_id="NOTIFY-FAILED",
                        related_task_id=None,
                        related_review_task_id=None,
                        recipient_type="role",
                        recipient="operations",
                        channel="mock",
                        sent_at=now,
                        send_status=NotificationSendStatus.FAILED.value,
                        dedupe_key="notify-failed",
                        message="failed notification",
                        error_message="failed",
                        created_at=now,
                    )
                ]
            )

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, _, dashboard_body = self._call_app(
                    path="/dashboard",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                self.assertEqual(status, "200 OK")
                self.assertIn("待复核", dashboard_body)
                self.assertIn("即将超时复核", dashboard_body)
                self.assertIn("失败通知", dashboard_body)
                self.assertIn("待执行任务", dashboard_body)
                self.assertIn("已过期", dashboard_body)
                self.assertIn("任务 1，复核 1", dashboard_body)
                self.assertIn("review_status=pending", dashboard_body)
                self.assertIn("due=soon", dashboard_body)
                self.assertIn("send_status=failed", dashboard_body)
                self.assertIn("task_status=pending", dashboard_body)
                self.assertIn("task_status=expired", dashboard_body)
                self.assertNotIn(str(db_path), dashboard_body)
                self.assertNotIn("token=", dashboard_body)

                _, _, pending_tasks_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "task_status": "pending"}),
                    cookie=cookie,
                )
                self.assertIn("TASK-PENDING", pending_tasks_body)
                self.assertNotIn("TASK-FAILED", pending_tasks_body)
                self.assertNotIn("TASK-EXPIRED", pending_tasks_body)

                _, _, due_reviews_body = self._call_app(
                    path="/reviews",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "review_status": "pending", "due": "soon"}),
                    cookie=cookie,
                )
                self.assertIn("TASK-REVIEW-DUE", due_reviews_body)
                self.assertIn("TASK-REVIEW-DUE-UTC", due_reviews_body)
                self.assertNotIn("TASK-REVIEW-LATER", due_reviews_body)
                self.assertNotIn("TASK-REVIEW-OVERDUE", due_reviews_body)

                _, _, failed_notifications_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "send_status": "failed"}),
                    cookie=cookie,
                )
                self.assertIn("NOTIFY-FAILED", failed_notifications_body)
                self.assertNotIn("token=", failed_notifications_body)

    def test_tasks_page_sorts_open_newest_first_and_paginates(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            tasks = [
                replace(
                    _runtime_task(f"TASK-PAGE-{index:03d}", status=TaskStatus.PENDING),
                    created_at=datetime(2026, 5, 4, 9, 0) + timedelta(minutes=index),
                )
                for index in range(52)
            ]
            processed_newer = replace(
                _runtime_task("TASK-PROCESSED-NEWER", status=TaskStatus.SKIPPED),
                created_at=datetime(2026, 5, 8, 9, 0),
            )
            task_service.create_tasks(tasks + [processed_newer])

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                _, _, page_one = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                _, _, page_two = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "page": "2"}),
                    cookie=cookie,
                )

            self.assertIn("显示 1-50 / 53", page_one)
            self.assertIn("下一页", page_one)
            self.assertIn("TASK-PAGE-051", page_one)
            self.assertIn("TASK-PAGE-002", page_one)
            self.assertNotIn("TASK-PAGE-001", page_one)
            self.assertNotIn("TASK-PROCESSED-NEWER", page_one)
            self.assertIn("显示 51-53 / 53", page_two)
            self.assertIn("TASK-PAGE-001", page_two)
            self.assertIn("TASK-PAGE-000", page_two)
            self.assertIn("TASK-PROCESSED-NEWER", page_two)

    def test_reviews_center_detail_shows_context_without_token_or_large_json(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = replace(
                _runtime_task("TASK-REVIEW-DETAIL", status=TaskStatus.MANUAL_REVIEW),
                decision_trace={"safe": "visible", "huge": "x" * 5000},
            )
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            with patch.dict("os.environ", {"DEFAULT_NOTIFICATION_CHANNEL": "mock"}, clear=False):
                review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                    "REVIEW_TOKEN_SECRET": "unit-test-secret",
                },
                clear=False,
            ):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                )
                cookie = self._runtime_login(db_path)
                status, _, body = self._call_app(
                    path="/reviews",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "review_task_id": review.review_task_id}),
                    cookie=cookie,
                )

            self.assertEqual(status, "200 OK")
            self.assertIn(review.review_task_id, body)
            self.assertIn(task.task_id, body)
            self.assertIn(token_result.review_token.token_id, body)
            self.assertIn("mobile_reviewer", body)
            self.assertIn("safe", body)
            self.assertIn("已截断", body)
            self.assertNotIn("token_hash", body)
            self.assertNotIn(token_result.raw_token, body)
            self.assertNotIn("token=", body)
            self.assertNotIn("reviewer_code", body)
            self.assertNotIn("x" * 5000, body)

    def test_reviews_center_resolves_pending_review_and_prevents_repeat_submit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-REVIEW-RESOLVE", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            with patch.dict("os.environ", {"DEFAULT_NOTIFICATION_CHANNEL": "mock"}, clear=False):
                review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                body = urlencode(
                    {
                        "action": "resolve_review",
                        "review_task_id": review.review_task_id,
                        "review_status": "approved",
                        "resolution_note": "approved from reviews",
                        "resolution_payload_json": '{"source":"reviews"}',
                    }
                )
                first_status, first_headers, _ = self._call_app(
                    path="/reviews",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=body,
                    cookie=cookie,
                )
                second_status, _, second_body = self._call_app(
                    path="/reviews",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=body,
                    cookie=cookie,
                )
                _, _, detail_body = self._call_app(
                    path="/reviews",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "review_task_id": review.review_task_id}),
                    cookie=cookie,
                )

            self.assertEqual(first_status, "303 See Other")
            self.assertIn("/reviews", first_headers["Location"])
            self.assertIn(review.review_task_id, first_headers["Location"])
            self.assertEqual(second_status, "200 OK")
            self.assertIn("已处理", second_body)
            self.assertIn("已处理", detail_body)
            self.assertNotIn("reviewer_code", detail_body)
            resolved_review = review_service.get_review_task(review.review_task_id)
            resolved_task = task_service.get_task(task.task_id)
            history = task_service.list_status_history(task.task_id)
            self.assertEqual(resolved_review.review_status, ReviewTaskStatus.APPROVED)
            self.assertEqual(resolved_review.resolved_by, "admin")
            self.assertEqual(resolved_task.task_status, TaskStatus.PENDING)
            self.assertTrue(any(item.metadata.get("actor_source") == "session_user" for item in history))

    def test_reviews_center_rejects_expired_action_and_invalid_payload(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-REVIEW-INVALID", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            with patch.dict("os.environ", {"DEFAULT_NOTIFICATION_CHANNEL": "mock"}, clear=False):
                review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                expired_status, _, expired_body = self._call_app(
                    path="/reviews",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(
                        {
                            "action": "resolve_review",
                            "review_task_id": review.review_task_id,
                            "review_status": "expired",
                            "resolution_payload_json": "{}",
                        }
                    ),
                    cookie=cookie,
                )
                invalid_status, _, invalid_body = self._call_app(
                    path="/reviews",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(
                        {
                            "action": "resolve_review",
                            "review_task_id": review.review_task_id,
                            "review_status": "approved",
                            "resolution_payload_json": "[]",
                        }
                    ),
                    cookie=cookie,
                )
                huge_status, _, huge_body = self._call_app(
                    path="/reviews",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                    body=urlencode(
                        {
                            "action": "resolve_review",
                            "review_task_id": review.review_task_id,
                            "review_status": "approved",
                            "resolution_payload_json": '{"huge":"' + ("x" * 5000) + '"}',
                        }
                    ),
                    cookie=cookie,
                )

            self.assertEqual(expired_status, "200 OK")
            self.assertIn("expired", expired_body)
            self.assertEqual(invalid_status, "200 OK")
            self.assertIn("JSON object", invalid_body)
            self.assertEqual(huge_status, "200 OK")
            self.assertIn("4096", huge_body)
            self.assertEqual(review_service.get_review_task(review.review_task_id).review_status, ReviewTaskStatus.PENDING)

    def test_notifications_center_filters_detail_and_redacts_sensitive_links(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-NOTIFY-DETAIL", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            with patch.dict("os.environ", {"DEFAULT_NOTIFICATION_CHANNEL": "mock"}, clear=False):
                review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]
            now = datetime.now()
            repository.insert_notification_logs(
                [
                    NotificationLog(
                        notification_id="NOTIFY-FEISHU-FAILED",
                        related_task_id=task.task_id,
                        related_review_task_id=review.review_task_id,
                        recipient_type="role",
                        recipient="operations",
                        channel="feishu",
                        sent_at=None,
                        send_status=NotificationSendStatus.FAILED.value,
                        dedupe_key="notify-feishu-failed",
                        message=(
                            "mobile_review_url=https://mobile.example.com/mobile/review/"
                            f"{review.review_task_id}?token=raw-secret-token "
                            "webhook=https://open.feishu.cn/open-apis/bot/v2/hook/secret"
                        ),
                        error_message=(
                            "failed via https://open.feishu.cn/open-apis/bot/v2/hook/secret "
                            "/mobile/review/abc?token=another-secret " + ("x" * 240)
                        ),
                        created_at=now,
                    ),
                    NotificationLog(
                        notification_id="NOTIFY-MOCK-SUCCESS",
                        related_task_id=None,
                        related_review_task_id=None,
                        recipient_type="role",
                        recipient="operations",
                        channel="mock",
                        sent_at=now,
                        send_status=NotificationSendStatus.SUCCESS.value,
                        dedupe_key="notify-mock-success",
                        message="mock success",
                        created_at=now,
                    ),
                ]
            )

            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                    "FEISHU_MESSAGE_TYPE": "post",
                },
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                _, _, failed_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "send_status": "failed"}),
                    cookie=cookie,
                )
                _, _, channel_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "channel": "feishu"}),
                    cookie=cookie,
                )
                _, _, related_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "related_review_task_id": review.review_task_id}),
                    cookie=cookie,
                )
                _, _, detail_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "notification_id": "NOTIFY-FEISHU-FAILED"}),
                    cookie=cookie,
                )
                _, _, missing_body = self._call_app(
                    path="/notifications",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "notification_id": "MISSING-NOTIFY"}),
                    cookie=cookie,
                )

            self.assertIn("NOTIFY-FEISHU-FAILED", failed_body)
            self.assertNotIn("NOTIFY-MOCK-SUCCESS", failed_body)
            self.assertIn("NOTIFY-FEISHU-FAILED", channel_body)
            self.assertNotIn("NOTIFY-MOCK-SUCCESS", channel_body)
            self.assertIn("NOTIFY-FEISHU-FAILED", related_body)
            self.assertIn(review.review_task_id, detail_body)
            self.assertIn(task.task_id, detail_body)
            self.assertIn("当前飞书消息类型", detail_body)
            self.assertIn("post", detail_body)
            self.assertIn("/reviews?review_task_id=", detail_body)
            self.assertIn("未找到对应通知", missing_body)
            for body in (failed_body, channel_body, related_body, detail_body):
                self.assertNotIn("raw-secret-token", body)
                self.assertNotIn("another-secret", body)
                self.assertNotIn("token=", body)
                self.assertNotIn("https://mobile.example.com/mobile/review", body)
                self.assertNotIn("https://open.feishu.cn/open-apis/bot/v2/hook/secret", body)
                self.assertIn("[mobile_review_url_redacted]", body)
                self.assertIn("[webhook_redacted]", body)
            self.assertIn("已截断", detail_body)

    def test_tasks_center_filters_detail_and_related_runtime_records(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = replace(
                _runtime_task(
                    "TASK-CENTER-DETAIL",
                    status=TaskStatus.MANUAL_REVIEW,
                    action_type=TaskActionType.MANUAL_REVIEW,
                ),
                internal_sku="SKU-DETAIL",
                platform_name="demo-platform",
                target_status="manual_review",
                decision_trace={"safe": "visible", "huge": "x" * 5000},
                scope_type="sku",
                scope_key="SKU-DETAIL",
            )
            other_task = _runtime_task(
                "TASK-CENTER-OTHER",
                status=TaskStatus.PENDING,
                trade_date=date(2026, 5, 5),
                action_type=TaskActionType.UPDATE_PRICE,
            )
            task_service.create_tasks([task, other_task])
            task_service.change_status(
                task_id=task.task_id,
                to_status=TaskStatus.PENDING,
                changed_by="unit-test",
                reason="task center history",
                metadata={"actor_source": "test", "huge": "y" * 5000},
            )
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            with patch.dict("os.environ", {"DEFAULT_NOTIFICATION_CHANNEL": "mock"}, clear=False):
                review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]
            repository.insert_notification_logs(
                [
                    NotificationLog(
                        notification_id="TASK-DIRECT-NOTIFY",
                        related_task_id=task.task_id,
                        related_review_task_id=None,
                        recipient_type="role",
                        recipient="operations",
                        channel="mock",
                        sent_at=datetime.now(),
                        send_status=NotificationSendStatus.SUCCESS.value,
                        dedupe_key="task-direct-notify",
                        message="direct /mobile/review/abc?token=direct-secret",
                        created_at=datetime.now(),
                    ),
                    NotificationLog(
                        notification_id="TASK-REVIEW-NOTIFY",
                        related_task_id=None,
                        related_review_task_id=review.review_task_id,
                        recipient_type="role",
                        recipient="operations",
                        channel="mock",
                        sent_at=datetime.now(),
                        send_status=NotificationSendStatus.SUCCESS.value,
                        dedupe_key="task-review-notify",
                        message="review linked notification",
                        created_at=datetime.now(),
                    )
                ]
            )
            repository.insert_execution_logs(
                [
                    ExecutionLog(
                        log_id="EXEC-TASK-DETAIL",
                        task_id=task.task_id,
                        executor_name="mock_executor",
                        start_time=datetime.now(),
                        end_time=datetime.now(),
                        success_flag=False,
                        error_message="execution failed https://open.feishu.cn/open-apis/bot/v2/hook/secret",
                        raw_output="raw token=secret-output " + ("z" * 5000),
                        created_at=datetime.now(),
                    )
                ]
            )

            with patch.dict(
                "os.environ",
                {"RUNTIME_ADMIN_USER": "admin", "RUNTIME_ADMIN_PASSWORD": "secret"},
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                _, _, filtered_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode(
                        {
                            "runtime_db": str(db_path),
                            "trade_date": "2026-05-04",
                            "action_type": "manual_review",
                            "scope_type": "sku",
                            "scope_key": "SKU-DETAIL",
                        }
                    ),
                    cookie=cookie,
                )
                _, _, invalid_filter_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "trade_date": "not-a-date"}),
                    cookie=cookie,
                )
                _, _, detail_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "task_id": task.task_id}),
                    cookie=cookie,
                )
                _, _, missing_body = self._call_app(
                    path="/tasks",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path), "task_id": "MISSING-TASK"}),
                    cookie=cookie,
                )

            self.assertIn(task.task_id, filtered_body)
            self.assertNotIn(other_task.task_id, filtered_body)
            self.assertIn(task.task_id, invalid_filter_body)
            self.assertIn(other_task.task_id, invalid_filter_body)
            self.assertIn("任务详情", detail_body)
            self.assertIn("状态历史", detail_body)
            self.assertIn("关联复核", detail_body)
            self.assertIn("关联通知", detail_body)
            self.assertIn("关联执行日志", detail_body)
            self.assertIn(review.review_task_id, detail_body)
            self.assertIn("/reviews?review_task_id=", detail_body)
            self.assertIn("TASK-DIRECT-NOTIFY", detail_body)
            self.assertIn("/notifications?notification_id=", detail_body)
            self.assertIn("EXEC-TASK-DETAIL", detail_body)
            self.assertIn("通过复核关联", detail_body)
            self.assertIn("直接关联", detail_body)
            self.assertIn("已截断", detail_body)
            self.assertIn("未找到对应任务", missing_body)
            self.assertNotIn("token=", detail_body)
            self.assertNotIn("secret-output", detail_body)
            self.assertNotIn("direct-secret", detail_body)
            self.assertNotIn("https://open.feishu.cn/open-apis/bot/v2/hook/secret", detail_body)
            self.assertNotIn("z" * 5000, detail_body)

    def test_system_page_masks_secret_values(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "super-secret-password",
                    "REVIEW_TOKEN_SECRET": "review-secret-value",
                    "FEISHU_WEBHOOK_URL": "https://open.feishu.cn/webhook/abc123",
                    "FEISHU_WEBHOOK_SECRET": "feishu-secret-value",
                    "MOBILE_REVIEW_BASE_URL": "https://mobile-review.example.com",
                    "DEFAULT_NOTIFICATION_CHANNEL": "feishu",
                    "FEISHU_MESSAGE_TYPE": "post",
                },
                clear=False,
            ):
                cookie = self._runtime_login(db_path, password="super-secret-password")
                status, _, body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
            self.assertEqual(status, "200 OK")
            self.assertIn("已配置", body)
            self.assertIn(db_path.name, body)
            self.assertNotIn(str(db_path), body)
            self.assertNotIn("super-secret-password", body)
            self.assertNotIn("review-secret-value", body)
            self.assertNotIn("feishu-secret-value", body)
            self.assertNotIn("https://open.feishu.cn/webhook/abc123", body)

    def test_system_page_reports_counts_and_config_statuses(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            pending_task = _runtime_task("SYSTEM-PENDING", status=TaskStatus.PENDING)
            expired_task = _runtime_task("SYSTEM-EXPIRED", status=TaskStatus.EXPIRED)
            review_task_source = _runtime_task("SYSTEM-REVIEW-SOURCE", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([pending_task, expired_task, review_task_source])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            with patch.dict("os.environ", {"DEFAULT_NOTIFICATION_CHANNEL": "mock", "DEV_MODE": "true"}, clear=False):
                review_service.create_from_tasks([review_task_source])
            pending_review = review_service.list_review_tasks()[0]
            repository.insert_review_tasks(
                [
                    ReviewTask(
                        review_task_id="SYSTEM-EXPIRED-REVIEW",
                        trade_date=date(2026, 5, 4),
                        scope_type="global",
                        scope_key="system",
                        dedupe_key="system-expired-review",
                        source_task_id=None,
                        review_type="manual_review",
                        review_status=ReviewTaskStatus.EXPIRED,
                        reason="expired review",
                        created_at=datetime.now(),
                        updated_at=datetime.now(),
                    )
                ]
            )
            repository.insert_notification_logs(
                [
                    NotificationLog(
                        notification_id="SYSTEM-FAILED-NOTIFY",
                        related_task_id=None,
                        related_review_task_id=pending_review.review_task_id,
                        recipient_type="role",
                        recipient="operations",
                        channel="mock",
                        sent_at=None,
                        send_status=NotificationSendStatus.FAILED.value,
                        dedupe_key="system-failed-notify",
                        message="failed notification",
                        error_message="failure summary",
                        created_at=datetime.now(),
                    )
                ]
            )
            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "super-secret-password",
                    "REVIEW_TOKEN_SECRET": "review-token-secret-long-enough-for-tests",
                    "DEFAULT_NOTIFICATION_CHANNEL": "mock",
                    "DEV_MODE": "true",
                    "FEISHU_MESSAGE_TYPE": "post",
                },
                clear=False,
            ):
                ReviewTokenService(repository).create_token(pending_review.review_task_id)
                cookie = self._runtime_login(db_path, password="super-secret-password")
                status, _, body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )

        self.assertEqual(status, "200 OK")
        self.assertIn("DEFAULT_NOTIFICATION_CHANNEL", body)
        self.assertIn("DEV_MODE", body)
        self.assertIn("info", body)
        self.assertIn("结构版本", body)
        self.assertIn("review_tokens", body)
        self.assertNotIn("SYSTEM-PENDING", body)
        self.assertIn("2", body)
        self.assertNotIn(str(db_path), body)
        self.assertNotIn("review-token-secret-long-enough-for-tests", body)
        self.assertNotIn("super-secret-password", body)
        self.assertNotIn("token=", body)

    def test_system_page_handles_invalid_config_and_partial_or_missing_db(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "partial.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            connection = repository.connect()
            try:
                connection.execute("DROP TABLE review_tokens")
                connection.commit()
            finally:
                connection.close()
            missing_db_path = Path(temp_dir) / "missing.sqlite3"
            missing_db_path.touch()
            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                    "DEFAULT_NOTIFICATION_CHANNEL": "feishu",
                    "DEV_MODE": "false",
                    "FEISHU_MESSAGE_TYPE": "invalid",
                    "REVIEW_TOKEN_SECRET": "short",
                    "PRA_ALLOWED_DATA_DIRS": str(Path(temp_dir)),
                },
                clear=True,
            ):
                cookie = self._runtime_login(db_path)
                missing_cookie = self._runtime_login(missing_db_path)
                status, _, partial_body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                missing_status, _, missing_body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlencode({"runtime_db": str(missing_db_path)}),
                    cookie=missing_cookie,
                )

        self.assertEqual(status, "200 OK")
        self.assertEqual(missing_status, "200 OK")
        self.assertIn("error", partial_body)
        self.assertIn("review_tokens", partial_body)
        self.assertIn("OperationalError", partial_body)
        self.assertIn("FEISHU_WEBHOOK_URL", partial_body)
        self.assertIn("MOBILE_REVIEW_BASE_URL", partial_body)
        self.assertIn("FEISHU_MESSAGE_TYPE", partial_body)
        self.assertIn("error", missing_body)
        self.assertIn("missing.sqlite3", missing_body)
        self.assertNotIn(str(missing_db_path), missing_body)
        self.assertNotIn("token=", partial_body)

    def test_system_feishu_test_requires_login_and_respects_channel(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                    "DEFAULT_NOTIFICATION_CHANNEL": "mock",
                    "FEISHU_WEBHOOK_URL": "https://open.feishu.cn/open-apis/bot/v2/hook/secret",
                },
                clear=False,
            ), patch("app.services.runtime.urlopen") as fake_urlopen:
                status, _, body = self._call_app(
                    path="/system/test-feishu-notification",
                    method="POST",
                    query=urlencode({"runtime_db": str(db_path)}),
                )
                cookie = self._runtime_login(db_path)
                post_status, headers, _ = self._call_app(
                    path="/system/test-feishu-notification",
                    method="POST",
                    body=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                redirected_status, _, redirected_body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlparse(headers["Location"]).query,
                    cookie=cookie,
                )

            self.assertEqual(status, "200 OK")
            self.assertIn("需要先登录", body)
            self.assertEqual(post_status, "303 See Other")
            self.assertEqual(redirected_status, "200 OK")
            self.assertIn("当前通知渠道不是 feishu", redirected_body)
            self.assertFalse(fake_urlopen.called)
            self.assertEqual(repository.list_notification_logs(), [])

    def test_system_feishu_test_logs_failed_missing_webhook_without_500(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                    "DEFAULT_NOTIFICATION_CHANNEL": "feishu",
                    "FEISHU_WEBHOOK_URL": "",
                    "FEISHU_MESSAGE_TYPE": "post",
                },
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, headers, _ = self._call_app(
                    path="/system/test-feishu-notification",
                    method="POST",
                    body=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                _, _, body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlparse(headers["Location"]).query,
                    cookie=cookie,
                )

            logs = repository.list_notification_logs()
            self.assertEqual(status, "303 See Other")
            self.assertIn("发送失败", body)
            self.assertIn("FEISHU_WEBHOOK_URL", body)
            self.assertEqual(len(logs), 1)
            self.assertEqual(logs[0].recipient_type, "system")
            self.assertEqual(logs[0].recipient, "system_test")
            self.assertIsNone(logs[0].related_task_id)
            self.assertIsNone(logs[0].related_review_task_id)
            self.assertEqual(logs[0].message, "PRA 系统测试通知")
            self.assertEqual(logs[0].send_status, NotificationSendStatus.FAILED.value)
            self.assertNotIn("token=", body)
            self.assertNotIn(str(db_path), body)

    def test_system_feishu_test_sends_via_sender_logs_success_and_prg(self) -> None:
        captured: dict[str, object] = {}

        class FakeResponse:
            def getcode(self):
                return 200

            def read(self):
                return b'{"code":0,"msg":"success","request_id":"SYSTEM-REQ"}'

            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

        def fake_urlopen(request, timeout):
            captured["body"] = json.loads(request.data.decode("utf-8"))
            captured["url"] = request.full_url
            return FakeResponse()

        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            RuntimeTaskService(repository).init_schema()
            with patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                    "DEFAULT_NOTIFICATION_CHANNEL": "feishu",
                    "FEISHU_WEBHOOK_URL": "https://open.feishu.cn/open-apis/bot/v2/hook/secret-webhook",
                    "FEISHU_WEBHOOK_SECRET": "sign-secret",
                    "FEISHU_MESSAGE_TYPE": "post",
                },
                clear=False,
            ), patch("app.services.runtime.urlopen", side_effect=fake_urlopen):
                cookie = self._runtime_login(db_path)
                status, headers, _ = self._call_app(
                    path="/system/test-feishu-notification",
                    method="POST",
                    body=urlencode({"runtime_db": str(db_path)}),
                    cookie=cookie,
                )
                _, _, body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlparse(headers["Location"]).query,
                    cookie=cookie,
                )
                refresh_status, _, refresh_body = self._call_app(
                    path="/system",
                    method="GET",
                    query=urlparse(headers["Location"]).query,
                    cookie=cookie,
                )

            logs = repository.list_notification_logs()

        self.assertEqual(status, "303 See Other")
        self.assertEqual(refresh_status, "200 OK")
        self.assertIn("飞书测试通知发送成功", body)
        self.assertIn("飞书测试通知发送成功", refresh_body)
        self.assertEqual(len(logs), 1)
        self.assertEqual(logs[0].recipient_type, "system")
        self.assertEqual(logs[0].recipient, "system_test")
        self.assertIsNone(logs[0].related_task_id)
        self.assertIsNone(logs[0].related_review_task_id)
        self.assertEqual(logs[0].channel, "feishu")
        self.assertEqual(logs[0].send_status, NotificationSendStatus.SUCCESS.value)
        self.assertIn("system_test", logs[0].dedupe_key)
        self.assertEqual(logs[0].message, "PRA 系统测试通知")
        self.assertNotIn("token=", logs[0].message)
        self.assertNotIn("mobile_review_url", logs[0].message)
        request_body = captured["body"]
        self.assertEqual(request_body["msg_type"], "post")
        self.assertEqual(request_body["content"]["post"]["zh_cn"]["title"], "PRA 系统测试通知")
        flattened = str(request_body)
        self.assertIn("/system", flattened)
        self.assertNotIn("token=", flattened)
        self.assertNotIn("mobile_review_url", flattened)
        self.assertNotIn("secret-webhook", flattened)
        self.assertNotIn("sign-secret", flattened)
        self.assertNotIn(str(db_path), flattened)
        self.assertNotIn("token=", body)
        self.assertNotIn("https://open.feishu.cn/open-apis/bot/v2/hook/secret-webhook", body)
        self.assertNotIn("sign-secret", body)

    def test_legacy_routes_are_fail_closed_by_default(self) -> None:
        with patch.dict(
            "os.environ",
            {
                "PRA_ENV": "production",
                "PRA_ENABLE_LEGACY_WEB": "",
                "PRA_LEGACY_ACCESS_MODE": "",
                "PRA_PROXY_MODE": "",
            },
            clear=False,
        ):
            for path in ("/", "/tables", "/execution", "/manual-intervention"):
                with self.subTest(path=path):
                    status, _, body = self._call_app(
                        path=path,
                        method="GET",
                        environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
                    )
                    self.assertEqual(status, "403 Forbidden")
                    self.assertIn("旧版 Web 路由当前已安全关闭", body)

    def test_legacy_routes_require_session_after_explicit_loopback_enable(self) -> None:
        with patch.dict(
            "os.environ",
            {
                "PRA_ENV": "production",
                "PRA_ENABLE_LEGACY_WEB": "1",
                "PRA_LEGACY_ACCESS_MODE": "direct_loopback",
                "PRA_PROXY_MODE": "none",
            },
            clear=False,
        ):
            with patch("app.web._WEB_LISTEN_HOST", "127.0.0.1"):
                status, headers, _ = self._call_app(
                    path="/",
                    method="GET",
                    environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
                )
            self.assertEqual(status, "303 See Other")
            self.assertEqual(urlparse(headers["Location"]).path, "/runtime/login")
            self.assertEqual(urlparse(headers["Location"]).query, "next=%2F")

    def test_legacy_routes_allow_authenticated_direct_loopback_only(self) -> None:
        with patch.dict(
            "os.environ",
            {
                "PRA_ENV": "production",
                "PRA_ENABLE_LEGACY_WEB": "1",
                "PRA_LEGACY_ACCESS_MODE": "direct_loopback",
                "PRA_PROXY_MODE": "none",
                "RUNTIME_ADMIN_USER": "admin",
                "RUNTIME_ADMIN_PASSWORD": "secret",
            },
            clear=False,
        ):
            cookie = self._runtime_login(self._existing_runtime_db("legacy"))
            with patch("app.web._WEB_LISTEN_HOST", "127.0.0.1"):
                for path in ("/", "/tables", "/execution", "/manual-intervention"):
                    with self.subTest(path=path):
                        status, _, body = self._call_app(
                            path=path,
                            method="GET",
                            cookie=cookie,
                            environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
                        )
                        self.assertEqual(status, "200 OK")
                        self.assertNotIn("旧版 Web 路由当前已安全关闭", body)

    def test_legacy_routes_ignore_request_bind_overrides(self) -> None:
        with patch.dict(
            "os.environ",
            {
                "PRA_ENV": "production",
                "PRA_ENABLE_LEGACY_WEB": "1",
                "PRA_LEGACY_ACCESS_MODE": "direct_loopback",
                "PRA_PROXY_MODE": "none",
                "RUNTIME_ADMIN_USER": "admin",
                "RUNTIME_ADMIN_PASSWORD": "secret",
            },
            clear=False,
        ):
            cookie = self._runtime_login(self._existing_runtime_db("legacy"))

            # A wildcard startup bind remains public even if request-scoped
            # fields claim that it is loopback-only.
            with patch("app.web._WEB_LISTEN_HOST", "0.0.0.0"):
                status, _, body = self._call_app(
                    path="/",
                    method="GET",
                    cookie=cookie,
                    environ_overrides={
                        "REMOTE_ADDR": "127.0.0.1",
                        "PRA_LISTEN_HOST": "127.0.0.1",
                        "SERVER_ADDR": "127.0.0.1",
                    },
                )
            self.assertEqual(status, "403 Forbidden")
            self.assertIn("0.0.0.0", body)

            # The IPv6 wildcard bind is equally public even when request
            # fields claim that the service is bound only to ::1.
            with patch("app.web._WEB_LISTEN_HOST", "::"):
                status, _, body = self._call_app(
                    path="/",
                    method="GET",
                    cookie=cookie,
                    environ_overrides={
                        "REMOTE_ADDR": "::1",
                        "PRA_LISTEN_HOST": "::1",
                        "SERVER_ADDR": "::1",
                    },
                )
            self.assertEqual(status, "403 Forbidden")
            self.assertIn("服务监听地址 ::", body)

            # Without a startup binding context, request fields cannot prove
            # that the service is loopback-only; the gate fails closed.
            with patch("app.web._WEB_LISTEN_HOST", None):
                status, _, body = self._call_app(
                    path="/",
                    method="GET",
                    cookie=cookie,
                    environ_overrides={
                        "REMOTE_ADDR": "127.0.0.1",
                        "PRA_LISTEN_HOST": "127.0.0.1",
                        "SERVER_ADDR": "127.0.0.1",
                    },
                )
            self.assertEqual(status, "403 Forbidden")
            self.assertIn("服务监听地址", body)

    def test_legacy_routes_reject_proxy_headers_and_non_loopback_topology(self) -> None:
        with patch.dict(
            "os.environ",
            {
                "PRA_ENV": "production",
                "PRA_ENABLE_LEGACY_WEB": "1",
                "PRA_LEGACY_ACCESS_MODE": "direct_loopback",
                "PRA_PROXY_MODE": "none",
                "RUNTIME_ADMIN_USER": "admin",
                "RUNTIME_ADMIN_PASSWORD": "secret",
            },
            clear=False,
        ):
            cookie = self._runtime_login(self._existing_runtime_db("legacy"))
            with patch("app.web._WEB_LISTEN_HOST", "127.0.0.1"):
                cases = [
                    {"REMOTE_ADDR": "127.0.0.1", "HTTP_X_FORWARDED_FOR": "127.0.0.1"},
                    {"REMOTE_ADDR": "203.0.113.10", "HTTP_X_FORWARDED_FOR": "127.0.0.1"},
                    {"REMOTE_ADDR": "127.0.0.1", "HTTP_FORWARDED": "for=127.0.0.1"},
                    {"REMOTE_ADDR": "127.0.0.1", "HTTP_X_REAL_IP": "127.0.0.1"},
                ]
                for overrides in cases:
                    with self.subTest(overrides=overrides):
                        status, _, _ = self._call_app(
                            path="/",
                            method="GET",
                            cookie=cookie,
                            environ_overrides=overrides,
                        )
                        self.assertEqual(status, "403 Forbidden")

                with patch.dict("os.environ", {"PRA_PROXY_MODE": "reverse_proxy"}, clear=False):
                    status, _, _ = self._call_app(
                        path="/",
                        method="GET",
                        cookie=cookie,
                        environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
                    )
                self.assertEqual(status, "403 Forbidden")

                with patch.dict("os.environ", {"PRA_PROXY_MODE": ""}, clear=False):
                    status, _, _ = self._call_app(
                        path="/",
                        method="GET",
                        cookie=cookie,
                        environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
                    )
                self.assertEqual(status, "403 Forbidden")

            with patch("app.web._WEB_LISTEN_HOST", "::1"):
                ipv6_status, _, _ = self._call_app(
                    path="/",
                    method="GET",
                    cookie=cookie,
                    environ_overrides={"REMOTE_ADDR": "::1"},
                )
            self.assertEqual(ipv6_status, "200 OK")

    def test_mobile_review_get_records_detail_access_only(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-M1", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                )
                status, _, body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}",
                    method="GET",
                    query=urlencode({"runtime_db": str(Path(temp_dir) / 'attacker.sqlite3'), "token": token_result.raw_token}),
                )

            self.assertEqual(status, "200 OK")
            self.assertIn("manual_price_review", body)
            self.assertIn("TASK-M1", body)
            self.assertIn("approved", body)
            stored_token = repository.get_review_token(token_result.review_token.token_id)
            self.assertIsNone(stored_token.used_at)
            self.assertIsNotNone(stored_token.last_used_at)
            self.assertFalse((Path(temp_dir) / "attacker.sqlite3").exists())

    def test_mobile_review_ignores_runtime_db_query_parameter(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            attacker_db = Path(temp_dir) / "attacker.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-M1-SAFE", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(review.review_task_id)
                status, _, body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}",
                    method="GET",
                    query=urlencode({"runtime_db": str(attacker_db), "token": token_result.raw_token}),
                )

            self.assertEqual(status, "200 OK")
            self.assertIn("TASK-M1-SAFE", body)
            self.assertFalse(attacker_db.exists())

    def test_mobile_review_post_resolves_and_prevents_reuse(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-M2", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                )
                body = urlencode(
                    {
                        "token": token_result.raw_token,
                        "action": "approved",
                        "resolution_note": "mobile approved",
                        "resolution_payload_json": '{"source":"mobile"}',
                    }
                )
                first_status, first_headers, _ = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=body,
                )
                second_status, _, second_body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=body,
                )

            self.assertEqual(first_status, "303 See Other")
            self.assertIn("resolved=1", first_headers["Location"])
            self.assertEqual(second_status, "410 Gone")
            self.assertIn("链接已失效或无权访问该复核任务", second_body)

            stored_token = repository.get_review_token(token_result.review_token.token_id)
            resolved_review = review_service.get_review_task(review.review_task_id)
            resolved_task = task_service.get_task(task.task_id)
            history = task_service.list_status_history(task.task_id)
            self.assertIsNotNone(stored_token.used_at)
            self.assertEqual(resolved_review.resolved_by, "mobile_reviewer")
            self.assertEqual(resolved_task.task_status, TaskStatus.PENDING)
            self.assertTrue(any(item.metadata.get("actor_source") == "mobile_review_token" for item in history))

    def test_execution_failure_mobile_review_uses_retry_and_cancel_results(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            review_created_at = datetime.now()
            retry_task = replace(
                _runtime_task(
                    "TASK-EXECUTION-RETRY",
                    status=TaskStatus.MANUAL_REVIEW,
                    action_type=TaskActionType.UPDATE_PRICE,
                ),
                created_at=review_created_at,
                updated_at=review_created_at,
            )
            cancel_task = replace(
                _runtime_task(
                    "TASK-EXECUTION-CANCEL",
                    status=TaskStatus.MANUAL_REVIEW,
                    action_type=TaskActionType.SET_ONLINE,
                ),
                created_at=review_created_at,
                updated_at=review_created_at,
            )
            task_service.create_tasks([retry_task, cancel_task])
            review_service = ReviewTaskService(
                repository,
                runtime_task_service=task_service,
            )
            review_service.create_from_tasks([retry_task, cancel_task])
            reviews = {
                review.source_task_id: review
                for review in review_service.list_review_tasks()
            }

            with patch(
                "app.web.DEFAULT_RUNTIME_DB",
                db_path,
            ), patch.dict(
                "os.environ",
                {"REVIEW_TOKEN_SECRET": "unit-test-secret"},
                clear=False,
            ):
                retry_token = ReviewTokenService(repository).create_token(
                    reviews[retry_task.task_id].review_task_id,
                    token_subject="mobile_reviewer",
                )
                with closing(repository.connect()) as connection, connection:
                    connection.execute(
                        """
                        UPDATE review_tokens
                        SET allowed_actions = ?
                        WHERE token_id = ?
                        """,
                        (
                            '["approved","rejected","adjusted","cancelled"]',
                            retry_token.review_token.token_id,
                        ),
                    )
                page_status, _, page_body = self._call_app(
                    path=f"/mobile/review/{reviews[retry_task.task_id].review_task_id}",
                    method="GET",
                    query=urlencode({"token": retry_token.raw_token}),
                )
                retry_status, _, _ = self._call_app(
                    path=f"/mobile/review/{reviews[retry_task.task_id].review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": retry_token.raw_token,
                            "action": "approved",
                            "resolution_payload_json": "{}",
                        }
                    ),
                )

                cancel_token = ReviewTokenService(repository).create_token(
                    reviews[cancel_task.task_id].review_task_id,
                    token_subject="mobile_reviewer",
                )
                cancel_status, _, _ = self._call_app(
                    path=f"/mobile/review/{reviews[cancel_task.task_id].review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": cancel_token.raw_token,
                            "action": "cancelled",
                            "resolution_payload_json": "{}",
                        }
                    ),
                )

            self.assertEqual(page_status, "200 OK")
            self.assertIn(">重试任务</button>", page_body)
            self.assertIn(">取消任务</button>", page_body)
            self.assertNotIn(">已拒绝</button>", page_body)
            self.assertNotIn(">已调整</button>", page_body)
            self.assertEqual(retry_status, "303 See Other")
            self.assertEqual(cancel_status, "303 See Other")
            self.assertEqual(
                task_service.get_task(retry_task.task_id).task_status,
                TaskStatus.PENDING,
            )
            self.assertEqual(
                task_service.get_task(cancel_task.task_id).task_status,
                TaskStatus.CANCELLED,
            )
            self.assertEqual(
                review_service.get_review_task(
                    reviews[retry_task.task_id].review_task_id
                ).resolution_payload["decision"],
                "retry_task",
            )
            self.assertEqual(
                review_service.get_review_task(
                    reviews[cancel_task.task_id].review_task_id
                ).resolution_payload["decision"],
                "cancel_task",
            )

    def test_mobile_review_cancels_the_whole_task_group(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            review_created_at = datetime.now()
            tasks = []
            for suffix in ("A", "B"):
                task = replace(
                    _runtime_task(
                        f"TASK-GROUP-CANCEL-{suffix}",
                        status=TaskStatus.MANUAL_REVIEW,
                        action_type=TaskActionType.SET_ONLINE,
                    ),
                    created_at=review_created_at,
                    updated_at=review_created_at,
                )
                task.decision_trace["task_group_id"] = "RULE-GROUP-CANCEL-001"
                tasks.append(task)
            task_service.create_tasks(tasks)
            review_service = ReviewTaskService(
                repository,
                runtime_task_service=task_service,
            )
            summary = review_service.create_from_tasks(tasks)
            review = summary.review_tasks[0]

            with patch(
                "app.web.DEFAULT_RUNTIME_DB",
                db_path,
            ), patch.dict(
                "os.environ",
                {"REVIEW_TOKEN_SECRET": "unit-test-secret"},
                clear=False,
            ):
                token = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                )
                page_status, _, page_body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}",
                    method="GET",
                    query=urlencode({"token": token.raw_token}),
                )
                resolve_status, _, _ = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": token.raw_token,
                            "action": "cancelled",
                            "resolution_payload_json": "{}",
                        }
                    ),
                )

            self.assertEqual(summary.inserted_review_tasks_count, 1)
            self.assertEqual(page_status, "200 OK")
            self.assertIn("RULE-GROUP-CANCEL-001", page_body)
            self.assertIn("affected_task_count", page_body)
            self.assertEqual(resolve_status, "303 See Other")
            for task in tasks:
                self.assertEqual(
                    task_service.get_task(task.task_id).task_status,
                    TaskStatus.CANCELLED,
                )
            resolved = review_service.get_review_task(review.review_task_id)
            self.assertEqual(resolved.resolution_payload["decision"], "cancel_task")
            self.assertEqual(resolved.resolution_payload["affected_task_count"], 2)

    def test_mobile_review_closes_pending_operational_source_task(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task(
                "TASK-M2-OPS",
                status=TaskStatus.PENDING,
                action_type=TaskActionType.LABOR_REQUIRED,
            )
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                )
                status, _, _ = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": token_result.raw_token,
                            "action": "approved",
                            "resolution_note": "mobile approved ops",
                        }
                    ),
                )

            resolved_task = task_service.get_task(task.task_id)
            history = task_service.list_status_history(task.task_id)
            self.assertEqual(status, "303 See Other")
            self.assertEqual(resolved_task.task_status, TaskStatus.SKIPPED)
            self.assertTrue(any(item.metadata.get("actor_source") == "mobile_review_token" for item in history))

    def test_mobile_review_post_rejects_invalid_payload_and_expired_action(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-M3", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                    allowed_actions=["approved", "expired"],
                )
                invalid_json_status, _, invalid_json_body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": token_result.raw_token,
                            "action": "approved",
                            "resolution_payload_json": "[]",
                        }
                    ),
                )
                expired_status, _, expired_body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": token_result.raw_token,
                            "action": "expired",
                            "resolution_payload_json": "{}",
                        }
                    ),
                )

            self.assertEqual(invalid_json_status, "200 OK")
            self.assertIn("JSON object", invalid_json_body)
            self.assertEqual(expired_status, "422 Unprocessable Entity")
            self.assertIn("链接已失效或无权访问该复核任务", expired_body)

    def test_mobile_review_action_not_allowed_returns_forbidden_without_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-M3-ACTION", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                    allowed_actions=["approved"],
                )
                status, _, body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": token_result.raw_token,
                            "action": "rejected",
                            "resolution_payload_json": "{}",
                        }
                    ),
                )

            self.assertEqual(status, "403 Forbidden")
            self.assertIn("链接已失效或无权访问该复核任务", body)
            self.assertIsNone(repository.get_review_token(token_result.review_token.token_id).used_at)
            self.assertEqual(repository.get_review_task(review.review_task_id).review_status, ReviewTaskStatus.PENDING)
            self.assertEqual(repository.get_task(task.task_id).task_status, TaskStatus.MANUAL_REVIEW)
            self.assertEqual(repository.list_task_status_history(task.task_id), [])

    def test_mobile_review_unknown_action_returns_422_without_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-M3-UNKNOWN", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                )
                status, _, body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": token_result.raw_token,
                            "action": "totally-invalid-action",
                            "resolution_payload_json": "{}",
                        }
                    ),
                )

            self.assertEqual(status, "422 Unprocessable Entity")
            self.assertIn("链接已失效或无权访问该复核任务", body)
            self.assertIsNone(repository.get_review_token(token_result.review_token.token_id).used_at)
            self.assertEqual(repository.get_review_task(review.review_task_id).review_status, ReviewTaskStatus.PENDING)
            self.assertEqual(repository.get_task(task.task_id).task_status, TaskStatus.MANUAL_REVIEW)
            self.assertEqual(repository.list_task_status_history(task.task_id), [])

    def test_mobile_review_token_fails_after_web_resolution(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task = _runtime_task("TASK-M4", status=TaskStatus.MANUAL_REVIEW)
            task_service.create_tasks([task])
            review_service = ReviewTaskService(repository, runtime_task_service=task_service)
            review_service.create_from_tasks([task])
            review = review_service.list_review_tasks()[0]

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict("os.environ", {"REVIEW_TOKEN_SECRET": "unit-test-secret"}, clear=False):
                token_result = ReviewTokenService(repository).create_token(
                    review.review_task_id,
                    token_subject="mobile_reviewer",
                )
                review_service.resolve_review_task(
                    review_task_id=review.review_task_id,
                    status=ReviewTaskStatus.CANCELLED,
                    actor="admin",
                    actor_source="web_session",
                    source_task_status=TaskStatus.CANCELLED,
                )
                status, _, body = self._call_app(
                    path=f"/mobile/review/{review.review_task_id}/resolve",
                    method="POST",
                    body=urlencode(
                        {
                            "token": token_result.raw_token,
                            "action": "approved",
                            "resolution_payload_json": "{}",
                        }
                    ),
                )

            self.assertEqual(status, "409 Conflict")
            self.assertIn("链接已失效或无权访问该复核任务", body)

    def test_execution_logs_page_displays_shadowbot_summary_evidence_and_warnings(self) -> None:
        log = ExecutionLog(
            log_id="LOG-SB-1",
            task_id="TASK-SB-1",
            executor_name="shadowbot_executor",
            start_time=datetime(2026, 6, 24, 9, 0),
            end_time=datetime(2026, 6, 24, 9, 1),
            success_flag=None,
            error_code="SUBMIT_RESULT_UNKNOWN",
            raw_output=json.dumps(
                {
                    "operation_id": "OP-1",
                    "execution_attempt_id": "ATTEMPT-1",
                    "shadowbot_run_id": "RUN-1",
                    "execution_mode": "COMMIT",
                    "worker_id": "LAPTOP-O9O76RQV",
                    "queue_phase": "RESULT_WRITTEN",
                    "instruction_hash": "sha256:instruction",
                    "request_file_sha256": "request-sha256",
                    "status": "NEEDS_RECONCILIATION",
                    "side_effect_state": "UNKNOWN",
                    "old_price": "19.00",
                    "target_price": "19.50",
                    "actual_price": "",
                    "evidence_status": "COMPLETE",
                    "evidence": [
                        {
                            "type": "BEFORE_SUBMIT",
                            "storage_uri": r"\\LAPTOP-O9O76RQV\pra-evidence\before.png",
                            "sha256": "abc123",
                            "upload_status": "SUCCESS",
                        }
                    ],
                },
                ensure_ascii=False,
            ),
        )

        html = render_execution_logs_page(
            runtime_db="runtime.sqlite3",
            session_user="tester",
            execution_logs=[log],
            shadowbot_queue_status={
                "queue_dir": r"D:\PRA_Runtime\shadowbot_queue",
                "heartbeat": {
                    "worker_id": "LAPTOP-O9O76RQV",
                    "status": "RUNNING",
                    "updated_at": "2026-06-24T09:01:00+08:00",
                },
                "phases": [
                    {
                        "execution_attempt_id": "ATTEMPT-1",
                        "execution_mode": "COMMIT",
                        "phase": "SUBMIT_CLICKED",
                        "side_effect_state": "SUBMIT_CLICKED",
                        "updated_at": "2026-06-24T09:01:00+08:00",
                    }
                ],
                "quarantine_count": 0,
            },
        )

        self.assertIn("operation_id", html)
        self.assertIn("OP-1", html)
        self.assertIn("execution_attempt_id", html)
        self.assertIn("ATTEMPT-1", html)
        self.assertIn("shadowbot_run_id", html)
        self.assertIn("RUN-1", html)
        self.assertIn("ShadowBot 队列状态", html)
        self.assertIn("LAPTOP-O9O76RQV", html)
        self.assertIn("SUBMIT_CLICKED", html)
        self.assertIn("request_file_sha256", html)
        self.assertIn("NEEDS_RECONCILIATION", html)
        self.assertIn("启动只读对账", html)
        self.assertIn("确认人工处理完成", html)
        self.assertIn("查看证据", html)
        self.assertIn("BEFORE_SUBMIT", html)
        self.assertIn("abc123", html)
        self.assertNotIn("强制重新提交", html)

    def test_tasks_page_displays_task13_listing_action_projection(self) -> None:
        task = _runtime_task(
            "TASK-T13-WEB-PROJECTION",
            action_type=TaskActionType.SET_OFFLINE,
            status=TaskStatus.RUNNING,
        )
        html = render_tasks_page(
            runtime_db="runtime.sqlite3",
            session_user="tester",
            tasks=[task],
            selected_task_id=task.task_id,
            selected_task=task,
            listing_action_projections=[
                {
                    "action_type": "set_offline",
                    "expected_old_status": "online",
                    "target_status": "offline",
                    "operation_result": "NEEDS_RECONCILIATION",
                    "batch_id": "BATCH-T13-WEB-001",
                    "operation_id": "OP-T13-WEB-001",
                    "batch_status": "UNKNOWN",
                    "operation_status": "NEEDS_RECONCILIATION",
                    "readback_observed_at": "2026-07-27T09:30:00+00:00",
                    "error_code": "CONTROLLED_AFTER_ACTION_CLICK_UNKNOWN",
                    "attempts": [
                        {
                            "execution_attempt_id": "ATTEMPT-T13-WEB-001",
                            "execution_mode": "COMMIT",
                            "status": "UNKNOWN",
                        },
                        {
                            "execution_attempt_id": "RECONCILE-T13-WEB-001",
                            "execution_mode": "RECONCILE",
                            "status": "RUNNING",
                        },
                    ],
                }
            ],
        )

        self.assertIn("上下架运行投影", html)
        self.assertIn("set_offline", html)
        self.assertIn("online", html)
        self.assertIn("offline", html)
        self.assertIn("实际回读状态", html)
        self.assertIn("BATCH-T13-WEB-001", html)
        self.assertIn("OP-T13-WEB-001", html)
        self.assertIn("ATTEMPT-T13-WEB-001", html)
        self.assertIn("RECONCILE-T13-WEB-001", html)
        self.assertIn("NEEDS_RECONCILIATION", html)
        self.assertIn("CONTROLLED_AFTER_ACTION_CLICK_UNKNOWN", html)
        self.assertNotIn("自动批准", html)

    def test_execution_logs_post_starts_shadowbot_reconcile_attempt(self) -> None:
        with _workspace_temp_dir("web_shadowbot_tests") as temp_dir:
            db_path = temp_dir / "runtime.sqlite3"
            request_dir = temp_dir / "shadowbot_requests"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task_service.create_tasks(
                [
                    _runtime_task(
                        "TASK-SB-WEB-1",
                        action_type=TaskActionType.UPDATE_PRICE,
                        status=TaskStatus.RUNNING,
                    )
                ]
            )
            repository.insert_shadowbot_operation(
                ShadowBotOperationLedger(
                    operation_id="OP-WEB-1",
                    task_id="TASK-SB-WEB-1",
                    platform="蚂蚁花团供应商",
                    product_identity={"sku": "SKU-AISHA-C", "name": "艾莎", "grade": "C级"},
                    expected_old_price=Decimal("19.00"),
                    target_price=Decimal("19.50"),
                    status="NEEDS_RECONCILIATION",
                    approved_payload_hash="hash",
                    created_at=datetime(2026, 6, 24, 9, 0),
                    updated_at=datetime(2026, 6, 24, 9, 0),
                )
            )

            runner = FileDropShadowBotTaskRunner(request_dir=request_dir)
            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch(
                "app.web.build_shadowbot_task_runner_from_environment",
                return_value=runner,
            ) as runner_factory, patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                },
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, headers, _ = self._call_app(
                    path="/execution-logs",
                    method="POST",
                    body=urlencode(
                        {
                            "action": "start_shadowbot_reconcile",
                            "operation_id": "OP-WEB-1",
                            "execution_attempt_id": "ATTEMPT-WEB-RECONCILE-1",
                        }
                    ),
                    cookie=cookie,
                )

            attempt = repository.get_shadowbot_execution_attempt("ATTEMPT-WEB-RECONCILE-1")
            operation = repository.get_shadowbot_operation("OP-WEB-1")
            request_file = request_dir / "inbox" / "ATTEMPT-WEB-RECONCILE-1.ready.json"
            self.assertEqual(status, "303 See Other")
            self.assertIn("已启动只读对账", unquote(headers["Location"]))
            runner_factory.assert_called_once_with()
            self.assertIsNotNone(attempt)
            self.assertEqual(attempt.execution_mode, "RECONCILE")
            self.assertEqual(attempt.shadowbot_run_id, "filequeue:ATTEMPT-WEB-RECONCILE-1")
            self.assertEqual(operation.status, "RUNNING")
            self.assertTrue(request_file.exists())
            request_text = request_file.read_text(encoding="utf-8")
            self.assertIn('"execution_mode": "RECONCILE"', request_text)
            self.assertIn('"platform_sku": ""', request_text)

    def test_execution_logs_post_confirms_shadowbot_manual_handled_without_resubmit(self) -> None:
        with _workspace_temp_dir("web_shadowbot_tests") as temp_dir:
            db_path = temp_dir / "runtime.sqlite3"
            repository = SQLiteRuntimeRepository(db_path)
            task_service = RuntimeTaskService(repository)
            task_service.init_schema()
            task_service.create_tasks(
                [
                    _runtime_task(
                        "TASK-SB-WEB-2",
                        action_type=TaskActionType.UPDATE_PRICE,
                        status=TaskStatus.RUNNING,
                    )
                ]
            )
            repository.insert_shadowbot_operation(
                ShadowBotOperationLedger(
                    operation_id="OP-WEB-2",
                    task_id="TASK-SB-WEB-2",
                    platform="蚂蚁花团供应商",
                    product_identity={"sku": "SKU-AISHA-C"},
                    expected_old_price=Decimal("19.00"),
                    target_price=Decimal("19.50"),
                    status="NEEDS_RECONCILIATION",
                    approved_payload_hash="hash",
                    created_at=datetime(2026, 6, 24, 9, 0),
                    updated_at=datetime(2026, 6, 24, 9, 0),
                )
            )

            with patch("app.web.DEFAULT_RUNTIME_DB", db_path), patch.dict(
                "os.environ",
                {
                    "RUNTIME_ADMIN_USER": "admin",
                    "RUNTIME_ADMIN_PASSWORD": "secret",
                },
                clear=False,
            ):
                cookie = self._runtime_login(db_path)
                status, headers, _ = self._call_app(
                    path="/execution-logs",
                    method="POST",
                    body=urlencode(
                        {
                            "action": "confirm_shadowbot_manual_handled",
                            "operation_id": "OP-WEB-2",
                            "manual_note": "checked by operator",
                        }
                    ),
                    cookie=cookie,
                )

            operation = repository.get_shadowbot_operation("OP-WEB-2")
            attempts = repository.get_shadowbot_execution_attempt("OP-WEB-2")
            logs = repository.list_execution_logs(task_id="TASK-SB-WEB-2")
            self.assertEqual(status, "303 See Other")
            self.assertIn("已确认人工处理完成", unquote(headers["Location"]))
            self.assertEqual(operation.status, "MANUAL_HANDLED")
            self.assertIsNone(attempts)
            self.assertTrue(any('"manual_note": "checked by operator"' in log.raw_output for log in logs))
            self.assertFalse(any("COMMIT" in log.raw_output for log in logs))


if __name__ == "__main__":
    unittest.main()
