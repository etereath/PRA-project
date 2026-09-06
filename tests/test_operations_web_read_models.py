from __future__ import annotations

import hashlib
import io
import re
from dataclasses import replace
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from urllib.parse import urlencode

import pytest
from openpyxl import Workbook, load_workbook

from app.enums import (
    DataQualityLevel,
    FactSource,
    ReviewTaskStatus,
    SellerPhase,
    SummaryStatus,
    TaskActionType,
    TaskOriginType,
    TaskStatus,
)
from app.models import ReviewTask, Task
from app.operational_models import PlatformTradeDaySummary, TradeDaySummaryEvent
from app.operations_web.app import create_application
from app.operations_web.composition import (
    OperationsWebPaths,
    OperationsWebSettings,
    build_container,
)
from app.operations_web.queries import (
    _automation_job_label,
    _notification_type_label,
    _review_type_label,
)
from app.repositories.operational_summary_repository import OperationalSummaryRepository
from app.repositories.sqlite_runtime_repository import SQLiteRuntimeRepository
from app.repositories.workbook_repository import PRODUCT_HEADERS
from app.repositories.workbook_repository import load_products
from app.repositories.inventory_repository import InventoryRepository
from app.services.authoritative_inventory import (
    InventoryApplicationService,
    sqlite_logical_snapshot_sha256,
)
from app.services.runtime import ReviewTokenService
from app.services.operational_time import OperationalTimePolicy
from tests.inventory_cutover_support import insert_cutover_order_snapshot


FIXED_NOW = datetime(2026, 8, 12, 1, 0, tzinfo=timezone.utc)
TRADE_DATE = date(2026, 8, 12)


@pytest.fixture()
def read_only_web(tmp_path: Path):
    runtime_db = tmp_path / "runtime.sqlite3"
    repository = SQLiteRuntimeRepository(runtime_db)
    repository.init_schema()
    products = tmp_path / "products.xlsx"
    _write_products(products)
    price_rules = tmp_path / "price_rules.xlsx"
    listing_rules = tmp_path / "listing_rules.xlsx"
    price_rules.write_bytes(b"synthetic")
    listing_rules.write_bytes(b"synthetic")
    queue_root = tmp_path / "queue"
    for name in ("inbox", "working", "results", "archive"):
        (queue_root / name).mkdir(parents=True, exist_ok=True)
    settings = OperationsWebSettings(
        environment="development",
        public_scheme="http",
        cookie_secure=False,
        admin_username="admin",
        admin_password="synthetic-password",
        paths=OperationsWebPaths(
            runtime_db=runtime_db,
            products_workbook=products,
            price_rules_workbook=price_rules,
            listing_rules_workbook=listing_rules,
            queue_root=queue_root,
        ),
    )
    container = build_container(settings)
    app = create_application(container)
    app.queries.now_provider = lambda: FIXED_NOW
    return app, container, repository, tmp_path


def test_today_uses_trade_day_summary_and_distinguishes_trustworthy_zero(read_only_web) -> None:
    app, container, repository, _ = read_only_web
    _insert_summary(
        repository,
        summary_id="SUMMARY-ZERO",
        scope_type="PLATFORM",
        scope_key="蚂蚁花团",
        sold_qty=0,
        amount=Decimal("0"),
        quality=DataQualityLevel.ORDER_COMPLETE,
    )
    cookie = _login(app, container)

    status, _, body = _call_app(app, path="/today", cookie=cookie)

    assert status == "200 OK"
    assert "已确认暂无销量" in body
    assert "0 扎" in body
    assert "¥0.00" in body
    assert "仅统计当前允许销售的商品" in body
    assert "完整度百分比" not in body
    assert "复购" not in body
    assert "买家端" not in body
    assert "Agent 建议" not in body


@pytest.mark.parametrize(
    ("quality", "sold_qty", "updated_at", "expected_title"),
    [
        (DataQualityLevel.ORDER_COMPLETE, 3, FIXED_NOW, "销售数据已更新"),
        (DataQualityLevel.ORDER_PARTIAL, 3, FIXED_NOW, "部分销售数据可用"),
        (DataQualityLevel.SCAN_ESTIMATED_HIGH, 3, FIXED_NOW, "销售数据已更新"),
        (DataQualityLevel.SCAN_ESTIMATED_MEDIUM, 3, FIXED_NOW, "部分销售数据可用"),
        (DataQualityLevel.SCAN_ESTIMATED_LOW, 3, FIXED_NOW, "部分销售数据可用"),
        (DataQualityLevel.UNAVAILABLE, None, FIXED_NOW, "销售数据暂不可用"),
        (
            DataQualityLevel.ORDER_COMPLETE,
            3,
            FIXED_NOW - timedelta(minutes=31),
            "销售数据已更新",
        ),
    ],
)
def test_today_presents_all_frozen_quality_and_freshness_branches(
    read_only_web,
    quality: DataQualityLevel,
    sold_qty: int | None,
    updated_at: datetime,
    expected_title: str,
) -> None:
    app, container, repository, _ = read_only_web
    _insert_summary(
        repository,
        summary_id=f"SUMMARY-{quality.value}",
        scope_type="PLATFORM",
        scope_key="蚂蚁花团",
        sold_qty=sold_qty,
        amount=(Decimal("30") if sold_qty is not None else None),
        quality=quality,
        updated_at=updated_at,
    )
    cookie = _login(app, container)

    status, _, body = _call_app(app, path="/today", cookie=cookie)

    assert status == "200 OK"
    assert expected_title in body
    if updated_at < FIXED_NOW:
        assert "2026-08-12 08:29" in body
    if quality is DataQualityLevel.UNAVAILABLE:
        assert "—" in body


def test_today_sellable_inventory_excludes_disabled_products(read_only_web) -> None:
    app, _, _, tmp_path = read_only_web
    workbook_path = tmp_path / "products.xlsx"
    workbook = load_workbook(workbook_path)
    sheet = workbook["data"]
    sheet.append(
        [
            "CLOSED-B-50-Z",
            "停售商品",
            "B级",
            "50cm",
            "扎",
            8,
            999,
            False,
            12,
            12,
            "合成测试",
            "",
            "",
        ]
    )
    workbook.save(workbook_path)

    model = app.queries.today()
    sellable_inventory = next(
        metric for metric in model.metrics if metric.label == "当前可售库存"
    )

    assert sellable_inventory.value == "72 扎"
    assert "允许销售" in sellable_inventory.note


@pytest.mark.parametrize("failure_mode", ["read-failed", "no-effective-policy"])
def test_today_does_not_query_a_guessed_trade_day_when_time_policy_is_unavailable(
    read_only_web,
    monkeypatch,
    failure_mode: str,
) -> None:
    app, container, _, _ = read_only_web
    summary_queries: list[dict[str, object]] = []

    if failure_mode == "read-failed":
        def fail_policy_read():
            raise RuntimeError("synthetic policy read failure")

        monkeypatch.setattr(
            app.queries.automation,
            "load_operational_time_policies",
            fail_policy_read,
        )
        expected_title = "销售日期读取失败"
    else:
        monkeypatch.setattr(
            app.queries.automation,
            "load_operational_time_policies",
            lambda: (
                OperationalTimePolicy(
                    policy_version="CN_SINGLE_PLATFORM_2026_V2",
                    effective_from=FIXED_NOW + timedelta(days=1),
                ),
            ),
        )
        expected_title = "当前销售日期暂不可用"

    def record_summary_query(**kwargs):
        summary_queries.append(kwargs)
        return []

    monkeypatch.setattr(
        app.queries.summaries,
        "list_summaries_page",
        record_summary_query,
    )
    cookie = _login(app, container)

    status, _, body = _call_app(app, path="/today", cookie=cookie)
    model = app.queries.today()

    assert status == "200 OK"
    assert expected_title in body
    assert model.platform_trade_date == "不可用"
    assert summary_queries == []


def test_explicit_historical_trade_date_remains_readable_when_current_policy_fails(
    read_only_web,
    monkeypatch,
) -> None:
    app, _, repository, _ = read_only_web
    _insert_summary(
        repository,
        summary_id="SUMMARY-HISTORICAL",
        scope_type="PLATFORM",
        scope_key="蚂蚁花团",
        sold_qty=3,
        amount=Decimal("30"),
        quality=DataQualityLevel.ORDER_COMPLETE,
    )

    def fail_policy_read():
        raise RuntimeError("synthetic policy read failure")

    monkeypatch.setattr(
        app.queries.automation,
        "load_operational_time_policies",
        fail_policy_read,
    )

    implicit = app.queries.database(
        section="business",
        dataset="settlements",
        page=1,
        trade_date=None,
        platform_name="",
    )
    explicit = app.queries.database(
        section="business",
        dataset="settlements",
        page=1,
        trade_date=TRADE_DATE,
        platform_name="",
    )

    assert implicit.table.state.state.value == "failed"
    assert implicit.table.rows == ()
    assert len(explicit.table.rows) == 1
    assert explicit.trade_date == TRADE_DATE.isoformat()


def test_database_task_history_uses_default_25_server_page(read_only_web) -> None:
    app, container, repository, _ = read_only_web
    for index in range(30):
        repository.insert_task(
            Task(
                task_id=f"TASK-{index:02d}",
                internal_sku=f"SKU-{index:02d}",
                platform_name="蚂蚁花团",
                action_type=TaskActionType.UPDATE_PRICE,
                priority=50,
                task_status=TaskStatus.PENDING,
                created_at=FIXED_NOW + timedelta(minutes=index),
                origin_type=TaskOriginType.MANUAL,
                origin_ref_id=f"synthetic:{index}",
            )
        )
    cookie = _login(app, container)

    status, _, first = _call_app(
        app,
        path="/database/project",
        query="dataset=tasks",
        cookie=cookie,
    )
    status_two, _, second = _call_app(
        app,
        path="/database/project",
        query="dataset=tasks&page=2",
        cookie=cookie,
    )

    assert status == status_two == "200 OK"
    assert first.count("<tbody>") == 1
    assert first.count("/management/task/") == 25
    assert "SKU-29" in first and "SKU-05" in first
    assert "SKU-04" not in first
    assert "下一页" in first
    assert second.count("/management/task/") == 5
    assert "SKU-04" in second and "SKU-00" in second


@pytest.mark.parametrize(
    ("dataset", "owner_name", "method_name"),
    [
        ("reviews", "runtime", "list_review_history_page"),
        ("runs", "automation", "list_runs"),
        ("incidents", "incidents", "list_history_page"),
        ("executions", "runtime", "list_execution_logs"),
        ("notifications", "runtime", "list_notification_outbox"),
    ],
)
def test_project_datasets_request_bounded_repository_pages(
    read_only_web,
    monkeypatch,
    dataset: str,
    owner_name: str,
    method_name: str,
) -> None:
    app, _, _, _ = read_only_web
    calls: list[dict[str, int]] = []

    def fake_page(**kwargs):
        calls.append(kwargs)
        return []

    owner = getattr(app.queries, owner_name)
    monkeypatch.setattr(owner, method_name, fake_page)

    model = app.queries.database(
        section="project",
        dataset=dataset,
        page=2,
        trade_date=TRADE_DATE,
        platform_name="",
    )

    assert calls == [{"limit": 26, "offset": 25}]
    assert model.table.page == 2
    assert model.table.page_size == 25
    assert model.table.has_previous is True


def test_detail_ownership_is_unique_and_cross_owner_routes_are_absent(read_only_web) -> None:
    app, container, repository, _ = read_only_web
    repository.insert_task(
        Task(
            task_id="TASK-DETAIL",
            internal_sku="AISHA-A-50-Z",
            platform_name="蚂蚁花团",
            action_type=TaskActionType.SET_OFFLINE,
            priority=10,
            task_status=TaskStatus.PENDING,
            created_at=FIXED_NOW,
            origin_type=TaskOriginType.MANUAL,
            origin_ref_id="synthetic:detail",
        )
    )
    cookie = _login(app, container)

    assert _call_app(app, path="/management/task/TASK-DETAIL", cookie=cookie)[0] == "200 OK"
    assert _call_app(app, path="/database/task/TASK-DETAIL", cookie=cookie)[0] == "404 Not Found"
    assert _call_app(app, path="/management/product/AISHA-A-50-Z", cookie=cookie)[0] == "404 Not Found"


def test_settlement_list_shows_only_current_version_and_marks_historical_detail(
    read_only_web,
) -> None:
    app, _, repository, _ = read_only_web
    previous = _insert_summary(
        repository,
        summary_id="SUMMARY-V1",
        scope_type="PLATFORM",
        scope_key="蚂蚁花团",
        sold_qty=3,
        amount=Decimal("30"),
        quality=DataQualityLevel.ORDER_COMPLETE,
    )
    revision = replace(
        previous,
        summary_id="SUMMARY-V2",
        version_no=2,
        supersedes_summary_id=previous.summary_id,
        sold_qty=4,
        transaction_amount_total=Decimal("40"),
        summary_status=SummaryStatus.OBSERVED,
        input_manifest_sha256="b" * 64,
        updated_at=FIXED_NOW + timedelta(minutes=5),
    )
    event = TradeDaySummaryEvent(
        event_id="EVENT-SUMMARY-V2",
        summary_id=revision.summary_id,
        from_status=SummaryStatus.PROVISIONAL,
        to_status=SummaryStatus.OBSERVED,
        trigger_type="SYNTHETIC_REVISION",
        trigger_ref_id="fixture-v2",
        fact_source_before=previous.fact_source,
        fact_source_after=revision.fact_source,
        quality_level_before=previous.quality_level,
        quality_level_after=revision.quality_level,
        input_manifest_sha256=revision.input_manifest_sha256,
        changed_at=revision.updated_at,
        changed_by="test",
    )
    inserted = OperationalSummaryRepository(repository).insert_revision(
        previous=previous,
        revision=revision,
        event=event,
        inputs=(),
    )

    model = app.queries.database(
        section="business",
        dataset="settlements",
        page=1,
        trade_date=TRADE_DATE,
        platform_name="",
    )
    old_detail = app.queries.detail("settlement", previous.summary_id)
    current_detail = app.queries.detail("settlement", revision.summary_id)

    assert inserted is True
    assert len(model.table.rows) == 1
    assert len(model.table.row_urls) == 1
    assert "SUMMARY-V2" in model.table.row_urls[0]
    assert "SUMMARY-V1" not in model.table.row_urls[0]
    assert old_detail is not None
    assert old_detail.state.title == "历史版本 · 已更新"
    assert any(
        field.label == "版本状态" and field.value == "历史版本，已更新"
        for field in old_detail.fields
    )
    assert current_detail is not None
    assert any(
        field.label == "版本状态" and field.value == "当前版本"
        for field in current_detail.fields
    )


def test_mobile_review_valid_invalid_expired_and_processed_states_are_zero_write(
    read_only_web,
    monkeypatch,
) -> None:
    app, _, repository, _ = read_only_web
    monkeypatch.setenv("REVIEW_TOKEN_SECRET", "synthetic-review-token-secret")
    review = ReviewTask(
        review_task_id="REVIEW-MOBILE",
        trade_date=TRADE_DATE,
        scope_type="sku",
        scope_key="AISHA-A-50-Z",
        dedupe_key="synthetic-review",
        source_task_id=None,
        review_type="emergency_protection",
        review_status=ReviewTaskStatus.PENDING,
        internal_sku="AISHA-A-50-Z",
        platform_name="蚂蚁花团",
        reason="售价低于安全阈值",
        required_by=FIXED_NOW + timedelta(minutes=30),
        created_at=FIXED_NOW,
        updated_at=FIXED_NOW,
    )
    repository.insert_review_tasks([review])
    created = ReviewTokenService(repository).create_token(
        review.review_task_id,
        allowed_actions=["adjusted", "approved", "rejected"],
        expires_at=FIXED_NOW + timedelta(hours=1),
    )
    repository.get_review_token(created.review_token.token_id)
    before = hashlib.sha256(repository.db_path.read_bytes()).hexdigest()

    valid_status, _, valid_body = _call_app(
        app,
        path="/mobile/review/REVIEW-MOBILE",
        query=urlencode({"token": created.raw_token}),
    )
    invalid_status, _, invalid_body = _call_app(
        app,
        path="/mobile/review/REVIEW-MOBILE",
        query=urlencode({"token": "invalid-secret"}),
    )

    assert valid_status == "200 OK"
    assert "等待处理" in valid_body
    assert "改价到" in valid_body
    assert "立即下架" in valid_body
    assert created.raw_token not in valid_body
    assert invalid_status == "404 Not Found"
    assert "链接无效" in invalid_body
    assert "invalid-secret" not in invalid_body
    unchanged = repository.get_review_token(created.review_token.token_id)
    assert unchanged is not None and unchanged.last_used_at is None
    assert hashlib.sha256(repository.db_path.read_bytes()).hexdigest() == before

    app.queries.now_provider = lambda: FIXED_NOW + timedelta(hours=2)
    expired_status, _, expired_body = _call_app(
        app,
        path="/mobile/review/REVIEW-MOBILE",
        query=urlencode({"token": created.raw_token}),
    )
    assert expired_status == "410 Gone"
    assert "链接已过期" in expired_body

    app.queries.now_provider = lambda: FIXED_NOW
    review.review_status = ReviewTaskStatus.APPROVED
    review.resolution_note = "已人工确认"
    review.resolved_at = FIXED_NOW + timedelta(minutes=5)
    review.updated_at = review.resolved_at
    repository.update_review_task(review)
    processed_status, _, processed_body = _call_app(
        app,
        path="/mobile/review/REVIEW-MOBILE",
        query=urlencode({"token": created.raw_token}),
    )
    assert processed_status == "200 OK"
    assert "已经处理" in processed_body


@pytest.mark.parametrize(
    ("review_type", "with_failed_task", "requested_actions", "expected_actions"),
    [
        (
            "emergency_protection",
            False,
            ["adjusted", "approved", "rejected", "cancelled"],
            ("改价到", "立即下架", "我来处理"),
        ),
        (
            "manual_review",
            True,
            ["approved", "cancelled", "rejected"],
            ("重试任务", "取消任务"),
        ),
        (
            "manual_price_review",
            False,
            ["approved", "rejected", "adjusted", "cancelled"],
            ("通过", "拒绝", "调整", "取消"),
        ),
    ],
)
def test_mobile_review_action_labels_reuse_formal_review_policy(
    read_only_web,
    monkeypatch,
    review_type: str,
    with_failed_task: bool,
    requested_actions: list[str],
    expected_actions: tuple[str, ...],
) -> None:
    app, _, repository, _ = read_only_web
    monkeypatch.setenv("REVIEW_TOKEN_SECRET", "synthetic-review-token-secret")
    source_task_id = None
    if with_failed_task:
        source_task_id = "TASK-EXECUTION-FAILURE"
        repository.insert_task(
            Task(
                task_id=source_task_id,
                internal_sku="AISHA-A-50-Z",
                platform_name="蚂蚁花团",
                action_type=TaskActionType.UPDATE_PRICE,
                priority=10,
                task_status=TaskStatus.MANUAL_REVIEW,
                created_at=FIXED_NOW,
                origin_type=TaskOriginType.MANUAL,
                origin_ref_id="synthetic:execution-failure",
            )
        )
    review = _make_review(
        review_task_id=f"REVIEW-{review_type}",
        review_type=review_type,
        source_task_id=source_task_id,
    )
    repository.insert_review_tasks([review])
    created = ReviewTokenService(repository).create_token(
        review.review_task_id,
        allowed_actions=requested_actions,
        expires_at=FIXED_NOW + timedelta(hours=1),
    )

    model = app.queries.mobile_review(review.review_task_id, created.raw_token)

    assert model.http_status == "200 OK"
    assert model.allowed_actions == expected_actions


def test_mobile_review_revoked_and_mismatched_tokens_use_stable_get_statuses(
    read_only_web,
    monkeypatch,
) -> None:
    app, _, repository, _ = read_only_web
    monkeypatch.setenv("REVIEW_TOKEN_SECRET", "synthetic-review-token-secret")
    first = _make_review(
        review_task_id="REVIEW-TOKEN-FIRST",
        review_type="manual_price_review",
    )
    second = _make_review(
        review_task_id="REVIEW-TOKEN-SECOND",
        review_type="manual_price_review",
    )
    repository.insert_review_tasks([first, second])
    created = ReviewTokenService(repository).create_token(
        first.review_task_id,
        expires_at=FIXED_NOW + timedelta(hours=1),
    )

    mismatch_status, _, mismatch_body = _call_app(
        app,
        path=f"/mobile/review/{second.review_task_id}",
        query=urlencode({"token": created.raw_token}),
    )

    assert mismatch_status == "404 Not Found"
    assert "链接无效" in mismatch_body

    ReviewTokenService(repository).revoke_token(
        created.review_token.token_id,
        revoked_at=FIXED_NOW,
    )
    revoked_status, _, revoked_body = _call_app(
        app,
        path=f"/mobile/review/{first.review_task_id}",
        query=urlencode({"token": created.raw_token}),
    )

    assert revoked_status == "410 Gone"
    assert "链接已失效" in revoked_body


def test_system_page_only_reports_current_component_state(read_only_web) -> None:
    app, container, _, _ = read_only_web
    cookie = _login(app, container)

    status, _, body = _call_app(app, path="/system", cookie=cookie)

    assert status == "200 OK"
    assert "业务数据库" in body
    assert "商品与规则资料" in body
    assert "平台任务传递" in body
    assert "影刀执行端" in body
    assert "历史任务" not in body
    assert "订单历史" not in body
    assert "平台任务暂时不会执行" in body


def test_operator_pages_do_not_expose_internal_implementation_language(
    read_only_web,
) -> None:
    app, container, _, _ = read_only_web
    cookie = _login(app, container)
    paths = (
        "/today",
        "/database",
        "/database/sales-analysis",
        "/database/dictionary",
        "/management",
        "/system",
        "/system/notifications",
        "/system/data",
        "/system/diagnostics",
    )
    forbidden = (
        "当前 PRA",
        "后端窄查询",
        "Runtime 数据库",
        "ShadowBot Worker",
        "Importer / Archive",
        "Outbox / 通知",
        "Web Route",
        "bootstrap",
        "原子事务",
        "Agent 预测",
        "当前真实库存权威",
        "缺失不显示为 0",
        "可信空页",
    )

    for path in paths:
        status, _, body = _call_app(app, path=path, cookie=cookie)
        assert status == "200 OK"
        for phrase in forbidden:
            assert phrase not in body, (path, phrase)


def test_operator_labels_do_not_fall_back_to_developer_identifiers() -> None:
    assert _automation_job_label("FULL_MARKET_SCAN") == "完整市场扫描"
    assert _automation_job_label("UNKNOWN_INTERNAL_JOB") == "其他自动化方案"
    assert _review_type_label("emergency_protection") == "价格异常处理"
    assert _review_type_label("unknown_internal_review") == "人工复核"
    assert _notification_type_label("mobile_review_required") == "价格异常，请立即处理"
    assert _notification_type_label("unknown_internal_notice") == "其他通知"


def test_management_inventory_adjustment_is_csrf_fenced_prg_and_db_only(
    read_only_web,
) -> None:
    app, container, repository, tmp_path = read_only_web
    workbook_path = tmp_path / "products.xlsx"
    products = load_products(workbook_path)
    cutover_batch_id = insert_cutover_order_snapshot(
        repository,
        batch_id="web-inventory-cutover-empty",
        observed_at=FIXED_NOW - timedelta(minutes=1),
        platform_trade_date=TRADE_DATE,
    )
    InventoryApplicationService(repository, clock=lambda: FIXED_NOW).bootstrap(
        products,
        snapshot_sha256="sha256:" + "a" * 64,
        runtime_snapshot_sha256=sqlite_logical_snapshot_sha256(repository),
        cutover_order_observation_batch_id=cutover_batch_id,
        idempotency_key="bootstrap:web-inventory",
        actor="admin",
        freeze_validator=lambda: True,
    )
    cookie = _login(app, container)
    session = container.sessions.get(cookie)
    assert session is not None

    status, _, body = _call_app(
        app,
        path="/management",
        cookie=cookie,
    )
    assert status == "200 OK"
    assert "人工库存调整" in body
    assert "当前 72 扎" in body
    key = re.search(r'name="idempotency_key" value="([^"]+)"', body)
    assert key is not None

    form = {
        "csrf_token": "invalid",
        "idempotency_key": key.group(1),
        "internal_sku": "AISHA-A-50-Z",
        "inventory_delta": "8",
        "expected_version": "1",
        "source_type": "NEW_FLOWER_INBOUND",
        "reason": "新花入库",
    }
    status, _, _ = _call_app(
        app,
        path="/management/inventory-adjustments",
        method="POST",
        cookie=cookie,
        form=form,
    )
    assert status == "403 Forbidden"
    assert InventoryRepository(repository).get_balance("AISHA-A-50-Z").current_qty == 72

    form["csrf_token"] = session.csrf_token
    status, headers, _ = _call_app(
        app,
        path="/management/inventory-adjustments",
        method="POST",
        cookie=cookie,
        form=form,
    )
    assert status == "303 See Other"
    location = _header(headers, "Location")
    assert location.startswith("/management?inventory_transaction=")
    assert InventoryRepository(repository).get_balance("AISHA-A-50-Z").current_qty == 80
    assert load_products(workbook_path)[0].current_stock == 72

    status, _, body = _call_app(
        app,
        path="/management",
        query=location.split("?", 1)[1],
        cookie=cookie,
    )
    assert status == "200 OK"
    assert "库存调整已记录" in body
    assert "72 扎" in body and "+8 扎" in body and "80 扎" in body

    status, _, body = _call_app(
        app,
        path="/database",
        query="dataset=inventory-adjustments",
        cookie=cookie,
    )
    assert status == "200 OK"
    assert "新花入库" in body
    assert "+8 扎" in body

    invalid_form = {
        "csrf_token": session.csrf_token,
        "idempotency_key": "web-inventory:invalid-sign",
        "internal_sku": "AISHA-A-50-Z",
        "inventory_delta": "-1",
        "expected_version": "2",
        "source_type": "NEW_FLOWER_INBOUND",
        "reason": "不应进入 URL 的原始输入",
    }
    status, headers, response_body = _call_app(
        app,
        path="/management/inventory-adjustments",
        method="POST",
        cookie=cookie,
        form=invalid_form,
    )
    assert status == "303 See Other"
    assert response_body == ""
    error_location = _header(headers, "Location")
    assert error_location == "/management?inventory_error=INVALID_ADJUSTMENT"
    assert "原始输入" not in error_location
    assert InventoryRepository(repository).get_balance("AISHA-A-50-Z").current_qty == 80

    status, _, body = _call_app(
        app,
        path="/management",
        query=error_location.split("?", 1)[1],
        cookie=cookie,
    )
    assert status == "200 OK"
    assert "库存调整未完成" in body
    assert "调整值、来源或调整后库存不符合要求" in body


def _write_products(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(PRODUCT_HEADERS)
    sheet.append(
        [
            "AISHA-A-50-Z",
            "艾莎",
            "A级",
            "50cm",
            "扎",
            8,
            72,
            True,
            12,
            12,
            "合成测试",
            "",
            "",
        ]
    )
    workbook.save(path)


def _make_review(
    *,
    review_task_id: str,
    review_type: str,
    source_task_id: str | None = None,
) -> ReviewTask:
    return ReviewTask(
        review_task_id=review_task_id,
        trade_date=TRADE_DATE,
        scope_type="sku",
        scope_key="AISHA-A-50-Z",
        dedupe_key=f"synthetic:{review_task_id}",
        source_task_id=source_task_id,
        review_type=review_type,
        review_status=ReviewTaskStatus.PENDING,
        internal_sku="AISHA-A-50-Z",
        platform_name="蚂蚁花团",
        reason="需要人工确认",
        required_by=FIXED_NOW + timedelta(minutes=30),
        created_at=FIXED_NOW,
        updated_at=FIXED_NOW,
    )


def _insert_summary(
    repository: SQLiteRuntimeRepository,
    *,
    summary_id: str,
    scope_type: str,
    scope_key: str,
    sold_qty: int | None,
    amount: Decimal | None,
    quality: DataQualityLevel,
    updated_at: datetime = FIXED_NOW,
) -> PlatformTradeDaySummary:
    fact_source = (
        None
        if quality is DataQualityLevel.UNAVAILABLE
        else (
            FactSource.SCAN_ESTIMATED
            if quality.value.startswith("SCAN_ESTIMATED")
            else FactSource.ORDER_OBSERVED
        )
    )
    summary = PlatformTradeDaySummary(
        summary_id=summary_id,
        summary_series_id=f"SERIES-{summary_id}",
        version_no=1,
        supersedes_summary_id=None,
        is_current=True,
        platform_name="蚂蚁花团",
        platform_trade_date=TRADE_DATE,
        seller_operation_date=TRADE_DATE,
        seller_phase=SellerPhase.NORMAL_SALES,
        scope_type=scope_type,
        scope_key=scope_key,
        fact_source=fact_source,
        quality_level=quality,
        summary_status=SummaryStatus.PROVISIONAL,
        sold_qty=sold_qty,
        order_count=None if sold_qty is None else (0 if sold_qty == 0 else 1),
        transaction_amount_total=amount,
        quality_reason="合成完整订单事实",
        source_proportions={fact_source.value: "1"} if fact_source else {},
        input_manifest_sha256="a" * 64,
        mapping_version="mapping-v1",
        algorithm_version="synthetic-v1",
        time_policy_version="CN_SINGLE_PLATFORM_2026_V1",
        finalized_at=None,
        created_at=FIXED_NOW,
        updated_at=updated_at,
    )
    event = TradeDaySummaryEvent(
        event_id=f"EVENT-{summary_id}",
        summary_id=summary_id,
        from_status=None,
        to_status=SummaryStatus.PROVISIONAL,
        trigger_type="SYNTHETIC_TEST",
        trigger_ref_id="fixture",
        fact_source_before=None,
        fact_source_after=fact_source,
        quality_level_before=None,
        quality_level_after=quality,
        input_manifest_sha256="a" * 64,
        changed_at=FIXED_NOW,
        changed_by="test",
    )
    OperationalSummaryRepository(repository).insert_initial(summary, event, ())
    return summary


def _call_app(
    app,
    *,
    path: str,
    method: str = "GET",
    query: str = "",
    form: dict[str, str] | None = None,
    cookie: str = "",
):
    encoded = urlencode(form or {}).encode("utf-8")
    environ = {
        "REQUEST_METHOD": method,
        "PATH_INFO": path,
        "QUERY_STRING": query,
        "CONTENT_LENGTH": str(len(encoded)),
        "CONTENT_TYPE": "application/x-www-form-urlencoded",
        "REMOTE_ADDR": "127.0.0.1",
        "HTTP_COOKIE": cookie,
        "wsgi.input": io.BytesIO(encoded),
    }
    captured: dict[str, object] = {}

    def start_response(status, headers):
        captured["status"] = status
        captured["headers"] = headers

    body = b"".join(app(environ, start_response)).decode("utf-8")
    return str(captured["status"]), list(captured["headers"]), body


def _login(app, container) -> str:
    status, headers, body = _call_app(app, path="/login")
    assert status == "200 OK"
    preauth = _header(headers, "Set-Cookie").split(";", 1)[0]
    csrf = re.search(r'name="csrf_token" value="([^"]+)"', body)
    assert csrf is not None
    status, headers, _ = _call_app(
        app,
        path="/login",
        method="POST",
        cookie=preauth,
        form={
            "username": container.settings.admin_username,
            "password": container.settings.admin_password,
            "csrf_token": csrf.group(1),
        },
    )
    assert status == "303 See Other"
    return _header(headers, "Set-Cookie").split(";", 1)[0]


def _header(headers: list[tuple[str, str]], name: str) -> str:
    return next(value for key, value in headers if key.lower() == name.lower())
